#!/usr/bin/env python3
"""Strict validation for the generated stock tax workbook.

Checks:
- Required sheets are present.
- No worksheet contains both <tableParts> and worksheet-level <autoFilter>.
- Checks sheet has no ERROR severity rows.
- SELL quantity matching difference does not exceed tolerance.
- Optional headless open via LibreOffice if available.
"""
from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Dict, List, Tuple
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

EXPECTED_SHEETS = [
    "README", "Operator_Dashboard", "Settings", "Import_Log", "Raw_Yahoo",
    "Ignored_Rows", "Transactions", "Instrument_Map", "FX_Daily",
    "FX_Yearly", "Corporate_Actions", "Split_Audit", "Method_Selection",
    "Method_Plan", "Filed_Year_Reconciliation", "Locked_Years",
    "Frozen_Inventory", "Frozen_Lot_Matching", "Frozen_Snapshots",
    "Review_State", "Lots", "Lot_Matching", "Sell_Review",
    "Open_Lots_Review", "Open_Position_Check", "Yearly_Tax_Summary",
    "Method_Comparison", "Checks", "Audit_Report",
]

DEFAULT_UNMATCHED_TOLERANCE = 1e-3
NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def read_table(ws, header_row: int = 1):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < header_row:
        return []
    header = rows[header_row - 1]
    out = []
    for r in rows[header_row:]:
        if r is None or all(v is None for v in r):
            continue
        out.append({str(h): r[i] for i, h in enumerate(header) if h is not None})
    return out


def worksheet_xml_map(xlsx_path: Path) -> Dict[str, str]:
    """Map sheet title -> worksheet XML path using workbook rels."""
    rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
    wb_ns = {
        "x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rel_map: Dict[str, str] = {}
    for rel in rels_xml.findall("r:Relationship", rel_ns):
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target", "")
        if rid and target:
            norm = target.replace("\\", "/")
            if norm.startswith("/"):
                norm = norm.lstrip("/")
            if not norm.startswith("xl/"):
                norm = f"xl/{norm}"
            rel_map[rid] = norm

    mapping: Dict[str, str] = {}
    for sheet in wb_xml.findall("x:sheets/x:sheet", wb_ns):
        name = sheet.attrib.get("name", "")
        rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if name and rid and rid in rel_map:
            mapping[name] = rel_map[rid]
    return mapping


def check_table_autofilter_conflicts(xlsx_path: Path) -> Tuple[List[dict], List[dict]]:
    report: List[dict] = []
    conflicts: List[dict] = []
    xml_map = worksheet_xml_map(xlsx_path)

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        for sheet_name, xml_path in sorted(xml_map.items()):
            data = zf.read(xml_path)
            root = ET.fromstring(data)
            has_table_parts = root.find("x:tableParts", NS) is not None
            has_ws_autofilter = root.find("x:autoFilter", NS) is not None
            if has_table_parts:
                report.append({
                    "sheet": sheet_name,
                    "has_table_parts": has_table_parts,
                    "has_ws_autofilter": has_ws_autofilter,
                })
            if has_table_parts and has_ws_autofilter:
                conflicts.append({
                    "sheet": sheet_name,
                    "xml_path": xml_path,
                })

    return report, conflicts


def try_headless_open_with_libreoffice(xlsx_path: Path) -> Tuple[bool, str]:
    common_soffice = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        for p in common_soffice:
            if Path(p).exists():
                soffice = p
                break
    if not soffice:
        return False, "LibreOffice not available; headless open skipped."

    with tempfile.TemporaryDirectory() as tmpdir:
        cmd = [
            soffice,
            "--headless",
            "--nologo",
            "--nodefault",
            "--nolockcheck",
            "--convert-to",
            "ods",
            "--outdir",
            tmpdir,
            str(xlsx_path),
        ]
        res = subprocess.run(cmd, capture_output=True, text=True)
        if res.returncode != 0:
            msg = (res.stderr or res.stdout or "").strip()
            return True, f"LibreOffice headless open failed: {msg}"
    return True, "LibreOffice headless open succeeded."


def find_excel_executable() -> str | None:
    common_excel = [
        r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE",
        r"C:\Program Files\Microsoft Office\Office16\EXCEL.EXE",
        r"C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE",
    ]
    for p in common_excel:
        if Path(p).exists():
            return p
    return shutil.which("excel")


def main(path: str = "stock_tax_system.xlsx", unmatched_tolerance: float = DEFAULT_UNMATCHED_TOLERANCE) -> int:
    xlsx_path = Path(path)
    failures: List[str] = []

    if not xlsx_path.exists():
        print(f"FAIL: workbook does not exist: {xlsx_path}")
        return 1

    try:
        wb = load_workbook(xlsx_path, data_only=False)
    except Exception as exc:
        print(f"FAIL: workbook structure invalid/corrupt: {exc}")
        return 1

    print(f"Workbook: {xlsx_path}")
    print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    missing = [s for s in EXPECTED_SHEETS if s not in wb.sheetnames]
    if missing:
        failures.append(f"Missing required sheets: {missing}")

    table_report, table_conflicts = check_table_autofilter_conflicts(xlsx_path)
    print("\nSheets with tableParts and worksheet autoFilter status:")
    if not table_report:
        print("  (none)")
    for r in table_report:
        status = "YES" if r["has_ws_autofilter"] else "NO"
        print(f"  {r['sheet']}: standalone worksheet autoFilter = {status}")

    if table_conflicts:
        failures.append(
            "Worksheet table/autofilter conflicts found: "
            + ", ".join([c["sheet"] for c in table_conflicts])
        )

    # Check for ERROR rows on Checks sheet.
    checks_rows = read_table(wb["Checks"])
    checks_errors = [r for r in checks_rows if str(r.get("Severity") or "").upper() == "ERROR"]
    print(f"\nChecks rows: {len(checks_rows)}; ERROR rows: {len(checks_errors)}")
    for r in checks_errors:
        print(f"  [ERROR] {r.get('Category')}: {r.get('Detail')}")
    if checks_errors:
        failures.append("Checks sheet contains ERROR rows.")

    # Open_Position_Check must not contain ERROR.
    opc_rows = read_table(wb["Open_Position_Check"])
    opc_errors = [r for r in opc_rows if str(r.get("Status") or "").upper() == "ERROR"]
    print(f"\nOpen_Position_Check rows: {len(opc_rows)}; ERROR rows: {len(opc_errors)}")
    if opc_errors:
        failures.append("Open_Position_Check contains ERROR rows.")

    # Filed years reconciliation must be OK.
    fyr_rows = read_table(wb["Filed_Year_Reconciliation"])
    fyr_errors = [r for r in fyr_rows if str(r.get("Status") or "").upper() == "ERROR"]
    print(f"Filed_Year_Reconciliation rows: {len(fyr_rows)}; ERROR rows: {len(fyr_errors)}")
    if fyr_errors:
        failures.append("Filed_Year_Reconciliation contains ERROR rows.")

    # Method_Plan should not suggest optimization for filed/unlocked years.
    mp_rows = read_table(wb["Method_Plan"])
    mp_flags = [
        r for r in mp_rows
        if str(r.get("Filed?") or "").upper() in ("TRUE", "1")
        and str(r.get("Locked?") or "").upper() not in ("TRUE", "1")
        and "review possible optimization" in str(r.get("Action required") or "").lower()
    ]
    print(f"Method_Plan rows: {len(mp_rows)}; filed+unlocked optimization flags: {len(mp_flags)}")
    if mp_flags:
        failures.append("Method_Plan has optimization warnings for filed unlocked years.")

    # SELL quantity conservation.
    lm = read_table(wb["Lot_Matching"])
    tx = read_table(wb["Transactions"])
    sells = [t for t in tx if t.get("Side") == "SELL"]
    matched_by_sell: Dict[str, float] = {}
    for r in lm:
        sid = r.get("Sell_ID")
        matched_by_sell[sid] = matched_by_sell.get(sid, 0.0) + float(r.get("Quantity") or 0.0)

    unmatched = []
    for t in sells:
        tx_id = t.get("Tx_ID")
        qty = float(t.get("Quantity") or 0.0)
        matched = matched_by_sell.get(tx_id, 0.0)
        diff = abs(matched - qty)
        if diff > unmatched_tolerance:
            unmatched.append((tx_id, t.get("Symbol"), qty, matched, diff))

    print(f"\nSELL transactions: {len(sells)}")
    print(f"Unmatched over tolerance ({unmatched_tolerance}): {len(unmatched)}")
    for u in unmatched:
        print(f"  {u[0]} {u[1]} qty={u[2]} matched={u[3]} diff={u[4]}")
    if unmatched:
        failures.append("Unmatched SELL quantity exceeds tolerance.")

    attempted_headless, headless_msg = try_headless_open_with_libreoffice(xlsx_path)
    print("\nHeadless open check:")
    print(f"  {headless_msg}")
    if attempted_headless and "failed" in headless_msg.lower():
        failures.append("LibreOffice headless open failed.")

    excel_exe = find_excel_executable()
    if excel_exe:
        print(f"Excel detected: {excel_exe}")
    else:
        print("Excel not found in common Windows paths.")

    if failures:
        print("\nVALIDATION FAILED:")
        for f in failures:
            print(f"  - {f}")
        return 1

    print("\nVALIDATION PASSED: workbook structure and integrity checks succeeded.")
    return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("path", nargs="?", default="stock_tax_system.xlsx")
    ap.add_argument("--unmatched-tolerance", type=float,
                    default=DEFAULT_UNMATCHED_TOLERANCE)
    args = ap.parse_args()
    raise SystemExit(main(args.path, args.unmatched_tolerance))
