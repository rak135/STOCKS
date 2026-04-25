from __future__ import annotations

import json
import shutil
from datetime import date
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

import build_stock_tax_workbook as workbook_module
from stock_tax_app.backend.main import create_app
from stock_tax_app.engine import policy
from stock_tax_app.engine import core as engine_core
from stock_tax_app.engine import ui_state as ui_state_module
from stock_tax_app.engine.core import run
from stock_tax_app.state import project_store
from stock_tax_app.state.models import ProjectState


ROOT = Path(__file__).resolve().parent


def _items(payload: dict) -> list[dict]:
    return payload["items"]


def _copy_project_fixture(tmp_path: Path) -> Path:
    project = tmp_path / "project"
    project.mkdir()
    shutil.copytree(ROOT / ".csv", project / ".csv")
    return project


def _workbook_path(project: Path) -> Path:
    # Tests pre-seed a workbook at the same path the backend runtime
    # uses as its export target so the runtime reads back the seeded
    # state. The runtime default is no longer the legacy
    # `stock_tax_system.xlsx` name.
    return project / "stock_tax_export.xlsx"


def _ensure_test_workbook(project: Path) -> Path:
    workbook_path = _workbook_path(project)
    if workbook_path.exists():
        return workbook_path

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=workbook_path,
        fetch_missing_fx=False,
    )
    workbook_module.write_workbook(
        calc.output_path,
        calc.raw_rows,
        calc.txs,
        calc.ignored,
        calc.problems,
        calc.instrument_map,
        calc.fx_yearly,
        calc.fx_daily,
        calc.fx_daily_sources,
        calc.corporate_actions,
        calc.method_selection,
        calc.locked_years,
        calc.settings,
        calc.frozen_inventory,
        calc.frozen_matching,
        calc.frozen_snapshots,
        calc.fx,
        calc.lots_final,
        calc.match_lines,
        calc.sim_warnings,
        calc.yearly_summary,
        calc.method_comparison,
        calc.split_warnings,
        calc.year_end_inventory,
        calc.import_log,
        calc.review_state,
        calc.filed_reconciliation,
        fx_yearly_sources=calc.fx_yearly_sources,
    )
    return workbook_path


def _find_workbook_review_state_row(project: Path, canonical_sell_id: str) -> int | None:
    wb = load_workbook(_ensure_test_workbook(project))
    ws = wb["Review_State"]
    for row in range(2, ws.max_row + 1):
        sell_id = ws.cell(row=row, column=1).value
        if ui_state_module.canonical_sell_id(sell_id) == canonical_sell_id:
            return row
    return None


def _first_workbook_review_sell_id(project: Path) -> tuple[str, str]:
    wb = load_workbook(_ensure_test_workbook(project))
    ws = wb["Review_State"]
    raw_sell_id = str(ws.cell(row=2, column=1).value)
    return raw_sell_id, ui_state_module.canonical_sell_id(raw_sell_id)


def _write_workbook_review_state(
    project: Path,
    canonical_sell_id: str,
    *,
    review_status: str,
    note: str,
) -> str:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Review_State"]
    row = None
    raw_sell_id = None
    for idx in range(2, ws.max_row + 1):
        candidate = ws.cell(row=idx, column=1).value
        if ui_state_module.canonical_sell_id(candidate) == canonical_sell_id:
            row = idx
            raw_sell_id = str(candidate)
            break
    if row is None:
        row = ws.max_row + 1
        raw_sell_id = canonical_sell_id
    ws.cell(row=row, column=1, value=raw_sell_id)
    ws.cell(row=row, column=2, value=review_status)
    ws.cell(row=row, column=3, value=note)
    wb.save(workbook_path)
    return raw_sell_id


def _set_year_fx_method(project: Path, year: int, method: str) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == year:
            ws.cell(row=row, column=3, value=method)
            wb.save(workbook_path)
            return
    raise AssertionError(f"Year {year} not found in Settings")


def _set_locked_year(project: Path, year: int, locked: bool) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Locked_Years"]
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == year:
            ws.cell(row=row, column=2, value=locked)
            wb.save(workbook_path)
            return
    raise AssertionError(f"Year {year} not found in Locked_Years")


def _build_check_rows_for_project(project: Path) -> list[dict]:
    workbook_path = _ensure_test_workbook(project)
    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=workbook_path,
        fetch_missing_fx=False,
    )
    return workbook_module.build_check_rows(
        sim_warnings=calc.sim_warnings,
        problems=calc.problems,
        fx_yearly=calc.fx_yearly,
        fx_daily=calc.fx_daily,
        settings=calc.settings,
        locked_years=calc.locked_years,
        frozen_inventory=calc.frozen_inventory,
        split_warnings=calc.split_warnings,
        method_selection=calc.method_selection,
        yearly_summary=calc.yearly_summary,
        match_lines=calc.match_lines,
        lots_final=calc.lots_final,
        year_end_inventory=calc.year_end_inventory,
        frozen_snapshots=calc.frozen_snapshots,
        fx=calc.fx,
    )


def _clear_fx_daily_rows(project: Path) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["FX_Daily"]
    for row in range(4, ws.max_row + 1):
        for col in range(1, 4):
            ws.cell(row=row, column=col, value=None)
    wb.save(workbook_path)


def _remove_fx_yearly_row(project: Path, year: int) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["FX_Yearly"]
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == year:
            for col in range(1, 4):
                ws.cell(row=row, column=col, value=None)
            break
    wb.save(workbook_path)


def _set_fx_yearly_row(project: Path, year: int, rate: float, source_note: str = "") -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["FX_Yearly"]
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == year:
            ws.cell(row=row, column=2, value=rate)
            ws.cell(row=row, column=3, value=source_note)
            wb.save(workbook_path)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=year)
    ws.cell(row=row, column=2, value=rate)
    ws.cell(row=row, column=3, value=source_note)
    wb.save(workbook_path)


def _set_fx_daily_row(project: Path, target_date: str, rate: float, source_note: str = "") -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["FX_Daily"]
    for row in range(4, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        current = value.date().isoformat() if hasattr(value, "date") else str(value or "")
        if current == target_date:
            ws.cell(row=row, column=1, value=target_date)
            ws.cell(row=row, column=2, value=rate)
            ws.cell(row=row, column=3, value=source_note)
            wb.save(workbook_path)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=target_date)
    ws.cell(row=row, column=2, value=rate)
    ws.cell(row=row, column=3, value=source_note)
    wb.save(workbook_path)


def _set_instrument_map_row(
    project: Path,
    symbol: str,
    *,
    instrument_id: str,
    isin: str = "",
    instrument_name: str = "",
    notes: str = "",
) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Instrument_Map"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == symbol:
            ws.cell(row=row, column=2, value=instrument_id)
            ws.cell(row=row, column=3, value=isin)
            ws.cell(row=row, column=4, value=instrument_name)
            ws.cell(row=row, column=5, value=notes)
            wb.save(workbook_path)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=symbol)
    ws.cell(row=row, column=2, value=instrument_id)
    ws.cell(row=row, column=3, value=isin)
    ws.cell(row=row, column=4, value=instrument_name)
    ws.cell(row=row, column=5, value=notes)
    wb.save(workbook_path)


def _append_position_row(
    project: Path,
    *,
    symbol: str,
    quantity: float,
    csv_name: str = "Trading212.csv",
) -> int:
    csv_path = project / ".csv" / csv_name
    raw = csv_path.read_text(encoding="utf-8")
    lines = raw.splitlines()
    if not lines:
        raise AssertionError(f"CSV fixture {csv_name} is empty")

    headers = [value.strip() for value in lines[0].split(",")]
    values = {header: "" for header in headers}
    if "Symbol" in values:
        values["Symbol"] = symbol
    if "Quantity" in values:
        values["Quantity"] = f"{quantity:.8f}".rstrip("0").rstrip(".")
    if "Current Price" in values:
        values["Current Price"] = "0"
    if "Date" in values:
        values["Date"] = ""
    if "Time" in values:
        values["Time"] = ""

    row = ",".join(str(values.get(header, "")) for header in headers)
    with csv_path.open("a", encoding="utf-8", newline="\n") as fh:
        if raw and not raw.endswith("\n"):
            fh.write("\n")
        fh.write(row)
    return len(lines) + 1


def _append_position_row_with_snapshot(
    project: Path,
    *,
    symbol: str,
    quantity: float,
    snapshot_date: str,
    snapshot_time: str = "16:00 EDT",
    csv_name: str = "Trading212.csv",
) -> int:
    csv_path = project / ".csv" / csv_name
    raw = csv_path.read_text(encoding="utf-8")
    lines = raw.splitlines()
    if not lines:
        raise AssertionError(f"CSV fixture {csv_name} is empty")

    headers = [value.strip() for value in lines[0].split(",")]
    values = {header: "" for header in headers}
    values["Symbol"] = symbol
    values["Quantity"] = f"{quantity:.8f}".rstrip("0").rstrip(".")
    values["Current Price"] = "0"
    if "Date" in values:
        values["Date"] = snapshot_date
    if "Time" in values:
        values["Time"] = snapshot_time

    row = ",".join(str(values.get(header, "")) for header in headers)
    with csv_path.open("a", encoding="utf-8", newline="\n") as fh:
        if raw and not raw.endswith("\n"):
            fh.write("\n")
        fh.write(row)
    return len(lines) + 1


def _first_unknown_open_position(client: TestClient) -> dict:
    response = client.get("/api/open-positions")
    assert response.status_code == 200
    body = response.json()
    row = next(
        item
        for item in body["items"]
        if item["status"] == "unknown" and float(item["calculated_qty"]) > 0
    )
    return row


def test_engine_run_returns_engine_result(tmp_path):
    project = _copy_project_fixture(tmp_path)
    result = run(project_dir=project, write_workbook=False)

    assert result.__class__.__name__ == "EngineResult"
    assert result.tax_years.items
    assert result.sales.items
    year_2024 = next(year for year in result.tax_years.items if year.year == 2024)
    assert year_2024.method == "LIFO"
    assert year_2024.filed is True
    assert year_2024.locked is True
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)
    assert year_2025.method == "FIFO"
    assert year_2025.filed is False
    assert year_2025.locked is False


def test_api_status_import_years_and_sales(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    status = client.get("/api/status")
    assert status.status_code == 200
    status_body = status.json()
    assert status_body["next_action"] is not None
    assert status_body["next_action"]["href"] in engine_core.FRONTEND_READY_HREFS

    import_summary = client.get("/api/import")
    assert import_summary.status_code == 200
    assert len(import_summary.json()["files"]) == 5

    years = client.get("/api/years")
    assert years.status_code == 200
    year_2024 = next(year for year in _items(years.json()) if year["year"] == 2024)
    assert year_2024["method"] == "LIFO"
    assert year_2024["filed"] is True
    assert year_2024["locked"] is True
    assert year_2024["show_method_comparison"] is False

    sales = client.get("/api/sales")
    assert sales.status_code == 200
    assert len(_items(sales.json())) > 0


def test_api_runs_without_root_workbook_and_only_exports_explicitly(tmp_path):
    project = _copy_project_fixture(tmp_path)
    legacy_workbook_path = project / workbook_module.CANONICAL_OUTPUT_NAME
    assert not legacy_workbook_path.exists()

    client = TestClient(create_app(project_dir=project))
    runtime = client.app.state.runtime
    export_path = runtime.output_path
    assert export_path.name != workbook_module.CANONICAL_OUTPUT_NAME
    assert not export_path.exists()

    status = client.get("/api/status")
    years = client.get("/api/years")
    sales = client.get("/api/sales")
    recalc = client.post("/api/recalculate")

    assert status.status_code == 200
    assert years.status_code == 200
    assert sales.status_code == 200
    assert recalc.status_code == 200
    assert status.json()["global_status"] in {"ready", "needs_review", "blocked"}
    assert years.json()["truth"]["status"] in {"ready", "needs_review", "partial", "blocked"}
    assert sales.json()["truth"]["status"] in {"ready", "needs_review", "partial", "blocked"}
    assert not legacy_workbook_path.exists()
    assert not export_path.exists()

    export_result = runtime.calculate(write_workbook=True)
    assert export_result.tax_years.items
    assert export_path.exists()
    assert not legacy_workbook_path.exists()


def test_sales_list_includes_financial_fields_and_matches_detail(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    sales_response = client.get("/api/sales")
    assert sales_response.status_code == 200
    sales_items = _items(sales_response.json())
    assert sales_items

    summary = sales_items[0]
    assert "total_cost_basis_czk" in summary
    assert "total_gain_loss_czk" in summary
    assert isinstance(summary["total_cost_basis_czk"], (int, float))
    assert isinstance(summary["total_gain_loss_czk"], (int, float))

    detail_response = client.get(f"/api/sales/{summary['id']}")
    assert detail_response.status_code == 200
    detail = detail_response.json()

    assert summary["total_cost_basis_czk"] == detail["total_cost_basis_czk"]
    assert summary["total_gain_loss_czk"] == detail["total_gain_loss_czk"]


def test_policy_module_is_canonical_year_source_of_truth():
    assert policy.is_filed(2024) is True
    assert policy.is_locked(2024) is True
    assert policy.filed_method(2024) == "LIFO"
    assert policy.default_method_for(2024) == "LIFO"
    assert policy.resolved_method_for(2024, "FIFO") == "LIFO"

    assert policy.is_filed(2025) is False
    assert policy.is_locked(2025) is False
    assert policy.default_method_for(2025) == "FIFO"
    assert policy.resolved_method_for(2025) == "FIFO"
    assert policy.resolved_method_for(2025, None) == "FIFO"


def test_workbook_policy_aliases_cannot_drift_from_engine_policy():
    assert workbook_module.FILED_YEARS is policy.FILED_YEARS
    assert workbook_module.YEAR_DEFAULT_METHODS is policy.YEAR_DEFAULT_METHODS
    assert workbook_module.DEFAULT_METHOD == policy.DEFAULT_METHOD
    assert workbook_module.SUPPORTED_METHODS == policy.SUPPORTED_METHODS

    selection = workbook_module.build_method_selection(
        user_state={},
        years=[2024, 2025],
        instrument_ids=["AAPL"],
    )
    assert selection[(2024, "AAPL")] == "LIFO"
    assert selection[(2025, "AAPL")] == "FIFO"


def test_status_routing_only_uses_live_frontend_hrefs(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.get("/api/status")
    assert response.status_code == 200
    body = response.json()

    placeholder_hrefs = {
        "/audit",
        "/fx",
        "/open-positions",
        "/settings",
        "/sales-review",
    }
    allowed_hrefs = set(engine_core.FRONTEND_READY_HREFS)

    assert body["next_action"] is not None
    assert body["next_action"]["href"] in allowed_hrefs
    assert body["next_action"]["href"] not in placeholder_hrefs

    unresolved_checks = body["unresolved_checks"]
    assert unresolved_checks
    assert {check["href"] for check in unresolved_checks}.issubset(allowed_hrefs)
    assert not ({check["href"] for check in unresolved_checks} & placeholder_hrefs)


@pytest.mark.parametrize(
    ("category", "expected_href"),
    [
        ("fx_missing", "/"),
        ("remaining_position_mismatch", "/"),
        ("import_warning", "/import"),
        ("method_policy", "/tax-years"),
        ("filed_reconciliation", "/tax-years"),
        ("generic_check", "/"),
    ],
)
def test_check_href_remaps_placeholder_destinations(category, expected_href):
    assert engine_core._check_href(category) == expected_href
    assert engine_core._check_href(category) in engine_core.FRONTEND_READY_HREFS


def test_unresolved_checks_still_appear_when_href_is_remapped(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.get("/api/status")
    assert response.status_code == 200
    body = response.json()

    assert body["global_status"] == "needs_review"
    assert body["next_action"] == {"label": "Review checks", "href": "/"}
    assert body["unresolved_checks"]
    assert all(check["message"] for check in body["unresolved_checks"])
    assert all(check["href"] == "/" for check in body["unresolved_checks"])


def test_patch_sale_review_updates_ui_state_only(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    sales_before = _items(client.get("/api/sales").json())
    sell_id = sales_before[0]["id"]
    detail_before = client.get(f"/api/sales/{sell_id}").json()
    years_before = client.get("/api/years").json()

    patched = client.patch(
        f"/api/sales/{sell_id}/review",
        json={"review_status": "reviewed", "note": "Checked in UI"},
    )
    assert patched.status_code == 200
    body = patched.json()
    assert body["review_status"] == "reviewed"
    assert body["note"] == "Checked in UI"

    ui_state_path = project / ".ui_state.json"
    assert ui_state_path.exists()
    ui_state_text = ui_state_path.read_text(encoding="utf-8")
    assert sell_id in ui_state_text
    assert "Checked in UI" in ui_state_text

    detail_after = client.get(f"/api/sales/{sell_id}").json()
    years_after = client.get("/api/years").json()
    assert detail_before["total_gain_loss_czk"] == detail_after["total_gain_loss_czk"]
    assert detail_before["total_cost_basis_czk"] == detail_after["total_cost_basis_czk"]
    assert years_before == years_after


def test_sale_review_patch_survives_recalc_and_runtime_reload(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    sell_id = _items(client.get("/api/sales").json())[0]["id"]
    patched = client.patch(
        f"/api/sales/{sell_id}/review",
        json={"review_status": "flagged", "note": "Needs follow-up"},
    )
    assert patched.status_code == 200

    recalculated = client.app.state.runtime.calculate(write_workbook=False)
    sale_after_recalc = next(sale for sale in recalculated.sales.items if sale.id == sell_id)
    assert sale_after_recalc.review_status == "flagged"
    assert sale_after_recalc.note == "Needs follow-up"

    detail_after_recalc = client.get(f"/api/sales/{sell_id}").json()
    assert detail_after_recalc["review_status"] == "flagged"
    assert detail_after_recalc["note"] == "Needs follow-up"

    reloaded_client = TestClient(create_app(project_dir=project))
    detail_after_reload = reloaded_client.get(f"/api/sales/{sell_id}").json()
    assert detail_after_reload["review_status"] == "flagged"
    assert detail_after_reload["note"] == "Needs follow-up"


def test_ui_state_beats_conflicting_workbook_review_state(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    sell_id = _items(client.get("/api/sales").json())[0]["id"]
    raw_sell_id = _write_workbook_review_state(
        project,
        sell_id,
        review_status="flagged",
        note="Workbook conflict",
    )

    state = ui_state_module.UIState()
    state.set_review(sell_id, review_status="reviewed", note="UI is canonical")
    ui_state_module.save(project / "stock_tax_export.xlsx", state)

    reloaded_client = TestClient(create_app(project_dir=project))
    detail = reloaded_client.get(f"/api/sales/{sell_id}").json()
    assert detail["review_status"] == "reviewed"
    assert detail["note"] == "UI is canonical"

    workbook_result = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=project / "stock_tax_export.xlsx",
        fetch_missing_fx=False,
    )
    assert workbook_result.review_state[sell_id]["review_status"] == "reviewed"
    assert workbook_result.review_state[sell_id]["operator_note"] == "UI is canonical"

    ui_payload = json.loads((project / ".ui_state.json").read_text(encoding="utf-8"))
    assert sell_id in ui_payload["sells"]
    assert raw_sell_id not in ui_payload["sells"]


def test_api_outputs_reflect_project_state_instrument_map(tmp_path):
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_instrument_map_row(
        project,
        symbol,
        instrument_id="SHOP_WORKBOOK",
        isin="WB000SHOP",
        instrument_name="Workbook SHOP",
        notes="workbook map",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            instrument_map={
                symbol: {
                    "yahoo_symbol": symbol,
                    "instrument_id": "SHOP_STATE",
                    "isin": "STATE000SHOP",
                    "instrument_name": "State SHOP",
                    "notes": "state map",
                }
            }
        ),
    )

    client = TestClient(create_app(project_dir=project))

    sales = client.get("/api/sales")
    assert sales.status_code == 200
    shop_sale = next(row for row in _items(sales.json()) if row["ticker"] == symbol)
    assert shop_sale["instrument_id"] == "SHOP_STATE"

    positions = client.get("/api/open-positions")
    assert positions.status_code == 200


def test_workbook_export_reflects_backend_ui_state(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    sell_id = _items(client.get("/api/sales").json())[0]["id"]
    patched = client.patch(
        f"/api/sales/{sell_id}/review",
        json={"review_status": "reviewed", "note": "Export me"},
    )
    assert patched.status_code == 200

    result = client.app.state.runtime.calculate(write_workbook=True)
    sale_after_write = next(sale for sale in result.sales.items if sale.id == sell_id)
    assert sale_after_write.review_status == "reviewed"
    assert sale_after_write.note == "Export me"

    row = _find_workbook_review_state_row(project, sell_id)
    assert row is not None

    wb = load_workbook(project / "stock_tax_export.xlsx")
    ws = wb["Review_State"]
    assert ws.cell(row=row, column=2).value == "reviewed"
    assert ws.cell(row=row, column=3).value == "Export me"


def test_legacy_workbook_review_state_migrates_when_ui_state_missing(tmp_path):
    project = _copy_project_fixture(tmp_path)
    raw_existing_sell_id, sell_id = _first_workbook_review_sell_id(project)
    raw_sell_id = _write_workbook_review_state(
        project,
        sell_id,
        review_status="flagged",
        note="Migrated from workbook",
    )
    assert raw_sell_id == raw_existing_sell_id
    ui_state_path = project / ".ui_state.json"
    assert not ui_state_path.exists()

    reloaded_client = TestClient(create_app(project_dir=project))
    detail = reloaded_client.get(f"/api/sales/{sell_id}").json()
    assert detail["review_status"] == "flagged"
    assert detail["note"] == "Migrated from workbook"

    assert ui_state_path.exists()
    payload = json.loads(ui_state_path.read_text(encoding="utf-8"))
    assert payload["sells"][sell_id]["review_status"] == "flagged"
    assert payload["sells"][sell_id]["note"] == "Migrated from workbook"
    assert raw_sell_id not in payload["sells"]


def test_fx_resolver_missing_daily_rate_is_explicit_not_22():
    resolver = workbook_module.FXResolver(
        yearly={2020: 23.14},
        daily={},
        settings={2020: {"fx_method": "FX_DAILY_CNB"}},
    )

    rate, label = resolver.inspect_date(date(2020, 2, 5))
    assert rate is None
    assert label == "FX_DAILY_CNB_missing"

    with pytest.raises(ValueError, match="Missing FX_DAILY_CNB rate"):
        resolver.rate_for(date(2020, 2, 5))


def test_fx_resolver_complete_fx_still_works():
    resolver = workbook_module.FXResolver(
        yearly={2021: 21.68},
        daily={date(2020, 2, 4): 22.5},
        settings={
            2020: {"fx_method": "FX_DAILY_CNB"},
            2021: {"fx_method": "FX_UNIFIED_GFR"},
        },
    )

    assert resolver.rate_for(date(2020, 2, 5)) == (22.5, "FX_DAILY_CNB_back1d")
    assert resolver.rate_for(date(2021, 3, 1)) == (21.68, "FX_UNIFIED_GFR")


def test_api_status_exposes_missing_fx_and_blocks_calculation(tmp_path, monkeypatch):
    project = _copy_project_fixture(tmp_path)
    _set_year_fx_method(project, 2020, "FX_DAILY_CNB")
    _clear_fx_daily_rows(project)
    monkeypatch.setattr(
        workbook_module,
        "download_cnb_daily_rates_year",
        lambda year, timeout=15: {},
    )

    client = TestClient(create_app(project_dir=project))

    status = client.get("/api/status")
    assert status.status_code == 200
    body = status.json()
    assert body["global_status"] == "blocked"
    assert body["next_action"] is not None
    assert body["next_action"]["href"] in engine_core.FRONTEND_READY_HREFS
    assert any("FX_DAILY_CNB" in check["message"] for check in body["unresolved_checks"])
    assert all(check["href"] in engine_core.FRONTEND_READY_HREFS for check in body["unresolved_checks"])

    sales = client.get("/api/sales")
    assert sales.status_code == 200
    assert sales.json()["items"] == []
    assert sales.json()["truth"]["empty_meaning"] == "blocked"

    years = client.get("/api/years")
    assert years.status_code == 200
    assert years.json()["items"] == []
    assert years.json()["truth"]["empty_meaning"] == "blocked"

    fx = client.get("/api/fx")
    assert fx.status_code == 200
    fx_2020 = next(row for row in fx.json()["items"] if row["year"] == 2020)
    assert fx_2020["missing_dates"]


def test_blocked_collections_expose_truth_metadata_not_ambiguous_empty_success(tmp_path, monkeypatch):
    project = _copy_project_fixture(tmp_path)
    _set_year_fx_method(project, 2020, "FX_DAILY_CNB")
    _clear_fx_daily_rows(project)
    monkeypatch.setattr(
        workbook_module,
        "download_cnb_daily_rates_year",
        lambda year, timeout=15: {},
    )

    client = TestClient(create_app(project_dir=project))

    status_body = client.get("/api/status").json()
    assert status_body["global_status"] == "blocked"
    assert any(reason["message"] for reason in status_body["status_reasons"])

    sales_body = client.get("/api/sales").json()
    years_body = client.get("/api/years").json()
    positions_body = client.get("/api/open-positions").json()

    for payload in (sales_body, years_body, positions_body):
        assert payload["items"] == []
        assert payload["truth"]["status"] == "blocked"
        assert payload["truth"]["empty_meaning"] == "blocked"
        assert payload["truth"]["reasons"]


def test_sales_list_blocked_empty_has_no_financial_rows(tmp_path, monkeypatch):
    project = _copy_project_fixture(tmp_path)
    _set_year_fx_method(project, 2020, "FX_DAILY_CNB")
    _clear_fx_daily_rows(project)
    monkeypatch.setattr(
        workbook_module,
        "download_cnb_daily_rates_year",
        lambda year, timeout=15: {},
    )

    client = TestClient(create_app(project_dir=project))
    sales_body = client.get("/api/sales").json()

    assert sales_body["items"] == []
    assert sales_body["truth"]["status"] == "blocked"
    assert sales_body["truth"]["empty_meaning"] == "blocked"


def test_project_state_fx_can_unblock_strict_daily_fx(tmp_path, monkeypatch):
    project = _copy_project_fixture(tmp_path)
    _set_year_fx_method(project, 2020, "FX_DAILY_CNB")
    _clear_fx_daily_rows(project)
    monkeypatch.setattr(
        workbook_module,
        "download_cnb_daily_rates_year",
        lambda year, timeout=15: {},
    )

    blocked_client = TestClient(create_app(project_dir=project))
    blocked_before = blocked_client.get("/api/status").json()
    assert blocked_before["global_status"] == "blocked"

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=project / "stock_tax_export.xlsx",
        fetch_missing_fx=False,
    )
    required_dates = sorted({tx.trade_date.isoformat() for tx in calc.txs if tx.trade_date.year == 2020})
    project_store.save_project_state(
        project,
        ProjectState(
            fx_daily={
                day: {
                    "currency_pair": "USD/CZK",
                    "rate": 23.55,
                    "source_note": "state strict unblock",
                    "manual": True,
                }
                for day in required_dates
            }
        ),
    )

    client = TestClient(create_app(project_dir=project))
    status = client.get("/api/status")
    assert status.status_code == 200
    body = status.json()
    assert body["global_status"] != "blocked"
    assert not any("FX_DAILY_CNB" in check["message"] for check in body["unresolved_checks"])

    sales = client.get("/api/sales")
    assert sales.status_code == 200
    assert sales.json()["items"]

    fx = client.get("/api/fx")
    fx_2020 = next(row for row in fx.json()["items"] if row["year"] == 2020)
    assert fx_2020["missing_dates"] == []
    assert fx_2020["daily_cached"] >= len(required_dates)


def test_api_provenance_exposes_project_state_owned_domains(tmp_path, monkeypatch):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]
    _set_year_fx_method(project, 2020, "FX_DAILY_CNB")
    _clear_fx_daily_rows(project)
    monkeypatch.setattr(
        workbook_module,
        "download_cnb_daily_rates_year",
        lambda year, timeout=15: {},
    )

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=project / "stock_tax_export.xlsx",
        fetch_missing_fx=False,
    )
    required_dates = sorted(
        {tx.trade_date.isoformat() for tx in calc.txs if tx.trade_date.year == 2020}
    )
    project_store.save_project_state(
        project,
        ProjectState(
            year_settings={
                2025: {
                    "tax_rate": 0.2,
                    "fx_method": "FX_UNIFIED_GFR",
                    "apply_100k": False,
                    "notes": "state-owned",
                }
            },
            method_selection={first_sale.year: {"STATE_PROVENANCE": "MAX_GAIN"}},
            fx_daily={
                day: {
                    "currency_pair": "USD/CZK",
                    "rate": 23.55,
                    "source_note": "state strict unblock",
                    "manual": True,
                }
                for day in required_dates
            },
            instrument_map={
                first_sale.ticker: {
                    "yahoo_symbol": first_sale.ticker,
                    "instrument_id": "STATE_PROVENANCE",
                    "isin": "STATE123",
                    "instrument_name": "State Provenance Instrument",
                    "notes": "state-owned",
                }
            },
        ),
    )

    client = TestClient(create_app(project_dir=project))

    years_body = client.get("/api/years").json()
    year_2025 = next(year for year in years_body["items"] if year["year"] == 2025)
    year_first = next(year for year in years_body["items"] if year["year"] == first_sale.year)
    assert year_2025["tax_rate"] == 0.2
    assert year_2025["settings_source"] == "project_state"
    assert year_first["method_source"] == "project_state"

    sales_body = client.get("/api/sales").json()
    first_sale_row = next(row for row in sales_body["items"] if row["ticker"] == first_sale.ticker)
    assert first_sale_row["instrument_id"] == "STATE_PROVENANCE"
    assert first_sale_row["instrument_map_source"] == "project_state"

    fx_body = client.get("/api/fx").json()
    fx_2020 = next(row for row in fx_body["items"] if row["year"] == 2020)
    assert fx_2020["rate_source"] == "project_state"
    assert fx_2020["missing_dates"] == []


def test_missing_fx_still_blocks_after_project_state_merge_path(tmp_path, monkeypatch):
    project = _copy_project_fixture(tmp_path)
    _set_year_fx_method(project, 2020, "FX_DAILY_CNB")
    _clear_fx_daily_rows(project)
    monkeypatch.setattr(
        workbook_module,
        "download_cnb_daily_rates_year",
        lambda year, timeout=15: {},
    )

    client = TestClient(create_app(project_dir=project))
    status = client.get("/api/status")
    assert status.status_code == 200
    body = status.json()
    assert body["global_status"] == "blocked"
    assert any("FX_DAILY_CNB" in check["message"] for check in body["unresolved_checks"])

    years = client.get("/api/years")
    assert years.status_code == 200
    assert years.json()["items"] == []
    assert years.json()["truth"]["empty_meaning"] == "blocked"

    fx = client.get("/api/fx")
    assert fx.status_code == 200
    fx_2020 = next(row for row in fx.json()["items"] if row["year"] == 2020)
    assert fx_2020["unified_rate"] == 23.14
    assert fx_2020["missing_dates"]
    assert fx_2020["daily_cached"] == 0


def test_open_positions_unknown_rows_have_explicit_truth_reason(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.get("/api/open-positions")
    assert response.status_code == 200
    body = response.json()
    assert body["truth"]["status"] == "partial"

    unknown_rows = [row for row in body["items"] if row["status"] == "unknown"]
    assert unknown_rows
    assert all(row["truth_status"] == "unknown" for row in unknown_rows)
    assert all(row["status_reason_code"] for row in unknown_rows)
    assert all(row["status_reason"] for row in unknown_rows)


def test_open_positions_exact_match_is_ok_and_ready(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline_client = TestClient(create_app(project_dir=project))
    baseline_unknown = _first_unknown_open_position(baseline_client)

    source_row = _append_position_row_with_snapshot(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=float(baseline_unknown["calculated_qty"]),
        snapshot_date="2026/04/23",
    )

    client = TestClient(create_app(project_dir=project))
    response = client.get("/api/open-positions")
    assert response.status_code == 200
    body = response.json()

    row = next(item for item in body["items"] if item["instrument_id"] == baseline_unknown["instrument_id"])
    assert row["status"] == "ok"
    assert row["truth_status"] == "ready"
    assert row["reported_qty"] == pytest.approx(float(baseline_unknown["calculated_qty"]))
    assert row["difference"] == pytest.approx(0.0)
    assert row["status_reason_code"] == "reconciled_within_tolerance"
    assert row["tolerance"] == pytest.approx(1e-4)
    assert row["reported_position_source_file"] == "Trading212.csv"
    assert row["reported_position_source_row"] == source_row
    assert row["reported_position_broker"] == "Trading212"
    assert row["reported_position_account"] is None
    assert row["reported_position_snapshot_date"] == "2026-04-23"
    assert row["reported_position_source_status"] == "ready"
    assert row["reported_position_source_reason"] is None
    assert row["reported_position_source_count"] == 1
    assert body["truth"]["status"] in {"ready", "partial", "needs_review"}


def test_open_positions_warn_difference_creates_needs_review_and_status_check(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline_client = TestClient(create_app(project_dir=project))
    baseline_unknown = _first_unknown_open_position(baseline_client)
    calculated_qty = float(baseline_unknown["calculated_qty"])
    reported_qty = calculated_qty - 0.005

    _append_position_row_with_snapshot(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=reported_qty,
        snapshot_date="2026/04/23",
    )

    client = TestClient(create_app(project_dir=project))
    positions = client.get("/api/open-positions")
    assert positions.status_code == 200
    positions_body = positions.json()

    row = next(item for item in positions_body["items"] if item["instrument_id"] == baseline_unknown["instrument_id"])
    assert row["status"] == "warn"
    assert row["truth_status"] == "needs_review"
    assert row["status_reason_code"] == "difference_above_tolerance"
    assert row["status_reason"]

    status = client.get("/api/status")
    assert status.status_code == 200
    status_body = status.json()
    open_position_checks = [check for check in status_body["unresolved_checks"] if check["id"].startswith("open-position-")]
    assert open_position_checks
    assert any(check["level"] == "warn" for check in open_position_checks)
    assert any(baseline_unknown["instrument_id"] in check["message"] for check in open_position_checks)


def test_open_positions_material_difference_blocks_collection_and_surfaces_audit_reason(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline_client = TestClient(create_app(project_dir=project))
    baseline_unknown = _first_unknown_open_position(baseline_client)
    calculated_qty = float(baseline_unknown["calculated_qty"])
    reported_qty = calculated_qty - 0.5

    _append_position_row_with_snapshot(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=reported_qty,
        snapshot_date="2026/04/23",
    )

    client = TestClient(create_app(project_dir=project))
    positions = client.get("/api/open-positions")
    assert positions.status_code == 200
    positions_body = positions.json()
    row = next(item for item in positions_body["items"] if item["instrument_id"] == baseline_unknown["instrument_id"])

    assert row["status"] == "error"
    assert row["truth_status"] == "blocked"
    assert row["status_reason_code"] == "material_difference"
    assert positions_body["truth"]["status"] == "blocked"

    status = client.get("/api/status")
    assert status.status_code == 200
    status_body = status.json()
    assert status_body["global_status"] == "blocked"
    assert any(
        check["id"].startswith("open-position-") and check["level"] == "error"
        for check in status_body["unresolved_checks"]
    )

    audit = client.get("/api/audit")
    assert audit.status_code == 200
    audit_body = audit.json()
    assert any(reason["code"] == "open_positions_blocked" for reason in audit_body["status_reasons"])
    assert any(reason["code"].startswith("open_position_error_") for reason in audit_body["status_reasons"])


def test_open_positions_missing_reported_position_is_unknown_not_ok(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.get("/api/open-positions")
    assert response.status_code == 200
    body = response.json()

    unknown_rows = [row for row in body["items"] if row["status"] == "unknown"]
    assert unknown_rows
    assert all(row["truth_status"] == "unknown" for row in unknown_rows)
    assert all(row["status"] != "ok" for row in unknown_rows)
    assert all(row["reported_qty"] is None for row in unknown_rows)
    assert all(row["status_reason_code"] in {"unknown_missing_mapping", "unknown_missing_yahoo_position"} for row in unknown_rows)


def test_open_positions_tolerance_behavior_ok_vs_warn(tmp_path):
    ok_root = tmp_path / "ok_case"
    ok_root.mkdir(parents=True, exist_ok=True)
    project_ok = _copy_project_fixture(ok_root)
    ok_baseline_client = TestClient(create_app(project_dir=project_ok))
    ok_unknown = _first_unknown_open_position(ok_baseline_client)
    ok_calculated_qty = float(ok_unknown["calculated_qty"])
    _append_position_row_with_snapshot(
        project_ok,
        symbol=ok_unknown["ticker"],
        quantity=ok_calculated_qty - 0.00005,
        snapshot_date="2026/04/23",
    )

    ok_client = TestClient(create_app(project_dir=project_ok))
    ok_response = ok_client.get("/api/open-positions")
    assert ok_response.status_code == 200
    ok_body = ok_response.json()

    ok_row = next(item for item in ok_body["items"] if item["instrument_id"] == ok_unknown["instrument_id"])
    assert ok_row["status"] == "ok"
    assert abs(float(ok_row["difference"])) <= float(ok_row["tolerance"])

    warn_root = tmp_path / "warn_case"
    warn_root.mkdir(parents=True, exist_ok=True)
    project_warn = _copy_project_fixture(warn_root)
    warn_baseline_client = TestClient(create_app(project_dir=project_warn))
    warn_unknown = _first_unknown_open_position(warn_baseline_client)
    warn_calculated_qty = float(warn_unknown["calculated_qty"])
    _append_position_row_with_snapshot(
        project_warn,
        symbol=warn_unknown["ticker"],
        quantity=warn_calculated_qty - 0.005,
        snapshot_date="2026/04/23",
    )

    warn_client = TestClient(create_app(project_dir=project_warn))
    warn_response = warn_client.get("/api/open-positions")
    assert warn_response.status_code == 200
    warn_body = warn_response.json()
    warn_row = next(item for item in warn_body["items"] if item["instrument_id"] == warn_unknown["instrument_id"])

    assert warn_row["status"] == "warn"
    assert abs(float(warn_row["difference"])) > float(warn_row["tolerance"])


def test_open_positions_provenance_missing_snapshot_date_is_honest(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline_client = TestClient(create_app(project_dir=project))
    baseline_unknown = _first_unknown_open_position(baseline_client)

    source_row = _append_position_row(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=float(baseline_unknown["calculated_qty"]),
    )

    client = TestClient(create_app(project_dir=project))
    positions = client.get("/api/open-positions")
    assert positions.status_code == 200
    body = positions.json()

    row = next(item for item in body["items"] if item["instrument_id"] == baseline_unknown["instrument_id"])
    assert row["status"] == "ok"
    assert row["reported_position_source_file"] == "Trading212.csv"
    assert row["reported_position_source_row"] == source_row
    assert row["reported_position_snapshot_date"] is None
    assert row["reported_position_source_status"] == "partial"
    assert row["reported_position_source_reason"]
    assert "Snapshot date is unavailable" in row["reported_position_source_reason"]
    assert row["truth_status"] == "needs_review"
    assert row["status_reason_code"] == "reported_position_source_needs_review"


def test_open_positions_multiple_reported_rows_expose_ambiguity_and_source_count(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline_client = TestClient(create_app(project_dir=project))
    baseline_unknown = _first_unknown_open_position(baseline_client)
    calculated_qty = float(baseline_unknown["calculated_qty"])

    _append_position_row_with_snapshot(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=calculated_qty / 2,
        snapshot_date="2026/04/23",
    )
    _append_position_row_with_snapshot(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=calculated_qty / 2,
        snapshot_date="2026/04/23",
    )

    client = TestClient(create_app(project_dir=project))
    positions = client.get("/api/open-positions")
    assert positions.status_code == 200
    body = positions.json()

    row = next(item for item in body["items"] if item["instrument_id"] == baseline_unknown["instrument_id"])
    assert row["status"] == "ok"
    assert row["difference"] == pytest.approx(0.0)
    assert row["reported_position_source_count"] == 2
    assert row["reported_position_source_status"] == "partial"
    assert row["reported_position_source_reason"]
    assert "aggregated from 2 source rows" in row["reported_position_source_reason"]
    assert len(row["reported_position_sources"]) == 2


def test_status_and_audit_include_provenance_checks_for_quantity_match(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline_client = TestClient(create_app(project_dir=project))
    baseline_unknown = _first_unknown_open_position(baseline_client)

    _append_position_row(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=float(baseline_unknown["calculated_qty"]),
    )

    client = TestClient(create_app(project_dir=project))
    status = client.get("/api/status")
    assert status.status_code == 200
    status_body = status.json()
    assert any(
        check["id"].startswith("open-position-") and "provenance" in check["message"].lower()
        for check in status_body["unresolved_checks"]
    )

    audit = client.get("/api/audit")
    assert audit.status_code == 200
    audit_body = audit.json()
    assert any(reason["code"].startswith("open_position_provenance_") for reason in audit_body["status_reasons"])


def test_settings_domain_source_reports_project_state_corporate_actions(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    target = next(pos for pos in baseline.open_positions.items if float(pos.calculated_qty) > 0)
    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "ca-domain-source",
                    "action_type": "split",
                    "effective_date": "2100-01-01",
                    "instrument_id": target.instrument_id,
                    "ratio_numerator": 2.0,
                    "ratio_denominator": 1.0,
                    "source": "project_state",
                    "note": "source-check",
                    "enabled": True,
                }
            ]
        ),
    )

    client = TestClient(create_app(project_dir=project))
    settings = client.get("/api/settings")
    assert settings.status_code == 200
    body = settings.json()
    assert body["domain_sources"]["corporate_actions"] == "project_state"


def test_invalid_corporate_actions_surface_in_status_and_audit(tmp_path):
    project = _copy_project_fixture(tmp_path)
    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "dup-id",
                    "action_type": "split",
                    "effective_date": "bad-date",
                    "instrument_id": "UNKNOWN_INST",
                    "ratio_numerator": 0.0,
                    "ratio_denominator": 1.0,
                    "enabled": True,
                },
                {
                    "action_id": "unknown-type",
                    "action_type": "mystery_action",
                    "effective_date": "2100-01-01",
                    "instrument_id": "UNKNOWN_INST",
                    "enabled": True,
                },
                {
                    "action_id": "dup-id",
                    "action_type": "split",
                    "effective_date": "2100-01-01",
                    "instrument_id": "UNKNOWN_INST",
                    "enabled": True,
                },
                {
                    "action_id": "ticker-missing-target",
                    "action_type": "ticker_change",
                    "effective_date": "2100-01-01",
                    "instrument_id": "UNKNOWN_INST",
                    "enabled": True,
                },
            ]
        ),
    )

    client = TestClient(create_app(project_dir=project))
    status = client.get("/api/status")
    assert status.status_code == 200
    status_body = status.json()
    messages = [check["message"].lower() for check in status_body["unresolved_checks"]]

    assert any("invalid or missing effective date" in message for message in messages)
    assert any("duplicate action_id" in message for message in messages)
    assert any("unknown action_type" in message for message in messages)
    assert any("missing target" in message for message in messages)
    assert status_body["global_status"] in {"needs_review", "blocked"}

    audit = client.get("/api/audit")
    assert audit.status_code == 200
    audit_body = audit.json()
    assert any(reason["code"] == "corporate_action_checks" for reason in audit_body["status_reasons"])


def test_ticker_change_action_moves_open_inventory_identity(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    target = next(pos for pos in baseline.open_positions.items if float(pos.calculated_qty) > 0)
    renamed_id = f"{target.instrument_id}_RENAMED"
    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "ticker-change",
                    "action_type": "ticker_change",
                    "effective_date": "2100-01-01",
                    "instrument_id": target.instrument_id,
                    "target_instrument_id": renamed_id,
                    "ratio_numerator": 1.0,
                    "ratio_denominator": 1.0,
                    "enabled": True,
                }
            ]
        ),
    )

    client = TestClient(create_app(project_dir=project))
    positions = client.get("/api/open-positions")
    assert positions.status_code == 200
    body = positions.json()
    assert any(item["instrument_id"] == renamed_id for item in body["items"])
    assert not any(item["instrument_id"] == target.instrument_id for item in body["items"])


def test_open_positions_detect_mismatch_after_missing_split_action(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline_client = TestClient(create_app(project_dir=project))
    baseline_unknown = _first_unknown_open_position(baseline_client)
    baseline_qty = float(baseline_unknown["calculated_qty"])

    _append_position_row_with_snapshot(
        project,
        symbol=baseline_unknown["ticker"],
        quantity=baseline_qty,
        snapshot_date="2026/04/23",
    )

    ok_client = TestClient(create_app(project_dir=project))
    ok_positions = ok_client.get("/api/open-positions").json()
    ok_row = next(item for item in ok_positions["items"] if item["instrument_id"] == baseline_unknown["instrument_id"])
    assert ok_row["status"] == "ok"

    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "missing-split",
                    "action_type": "split",
                    "effective_date": "2100-01-01",
                    "instrument_id": baseline_unknown["instrument_id"],
                    "ratio_numerator": 2.0,
                    "ratio_denominator": 1.0,
                    "enabled": True,
                }
            ]
        ),
    )

    mismatch_client = TestClient(create_app(project_dir=project))
    mismatch_positions = mismatch_client.get("/api/open-positions")
    assert mismatch_positions.status_code == 200
    mismatch_body = mismatch_positions.json()
    mismatch_row = next(
        item for item in mismatch_body["items"] if item["instrument_id"] == baseline_unknown["instrument_id"]
    )
    assert mismatch_row["status"] in {"warn", "error"}

    status = mismatch_client.get("/api/status")
    assert status.status_code == 200
    status_body = status.json()
    assert any(check["id"].startswith("open-position-") for check in status_body["unresolved_checks"])


def test_settings_truth_discloses_display_only_and_domain_ownership(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.get("/api/settings")
    assert response.status_code == 200
    body = response.json()

    assert body["truth_status"] == "partial"
    assert body["status_reasons"]
    assert all(meta["editability"] != "editable" for meta in body["field_meta"].values())
    assert body["field_meta"]["default_tax_rate"]["editability"] == "display_only"
    assert body["field_meta"]["default_tax_rate"]["source"] == "static_config"
    assert body["domain_sources"]["year_settings"] == "project_state"
    assert body["domain_sources"]["corporate_actions"] == "workbook_fallback"


def test_audit_summary_truth_does_not_imply_final_export_readiness(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.get("/api/audit")
    assert response.status_code == 200
    body = response.json()

    assert body["summary_only"] is True
    assert body["truth_status"] == "partial"
    assert body["workbook_backed_domains"]
    assert any(reason["code"] == "audit_summary_only" for reason in body["status_reasons"])
    assert any(reason["code"] == "workbook_backed_domains" for reason in body["status_reasons"])


def test_workbook_export_reflects_project_state_fx(tmp_path):
    project = _copy_project_fixture(tmp_path)
    target_day = "2020-02-03"
    _set_fx_yearly_row(project, 2025, 19.75, "workbook yearly")
    _set_fx_daily_row(project, target_day, 19.25, "workbook daily")
    project_store.save_project_state(
        project,
        ProjectState(
            fx_yearly={
                2025: {
                    "currency_pair": "USD/CZK",
                    "rate": 24.01,
                    "source_note": "state export yearly",
                    "manual": True,
                }
            },
            fx_daily={
                target_day: {
                    "currency_pair": "USD/CZK",
                    "rate": 24.02,
                    "source_note": "state export daily",
                    "manual": True,
                }
            },
        ),
    )

    client = TestClient(create_app(project_dir=project))
    result = client.app.state.runtime.calculate(write_workbook=True)
    assert result.fx_years.items

    wb = load_workbook(project / "stock_tax_export.xlsx")
    yearly_ws = wb["FX_Yearly"]
    daily_ws = wb["FX_Daily"]

    yearly_rate = None
    yearly_note = None
    for row in range(4, yearly_ws.max_row + 1):
        if yearly_ws.cell(row=row, column=1).value == 2025:
            yearly_rate = yearly_ws.cell(row=row, column=2).value
            yearly_note = yearly_ws.cell(row=row, column=3).value
            break

    daily_rate = None
    daily_note = None
    for row in range(4, daily_ws.max_row + 1):
        value = daily_ws.cell(row=row, column=1).value
        current = value.date().isoformat() if hasattr(value, "date") else str(value or "")
        if current == target_day:
            daily_rate = daily_ws.cell(row=row, column=2).value
            daily_note = daily_ws.cell(row=row, column=3).value
            break

    assert yearly_rate == 24.01
    assert yearly_note == "state export yearly"
    assert daily_rate == 24.02
    assert daily_note == "state export daily"


def test_blocked_fx_run_skips_workbook_write_and_write_path_fails_cleanly(tmp_path, monkeypatch):
    project = _copy_project_fixture(tmp_path)
    _set_year_fx_method(project, 2020, "FX_DAILY_CNB")
    _clear_fx_daily_rows(project)
    monkeypatch.setattr(
        workbook_module,
        "download_cnb_daily_rates_year",
        lambda year, timeout=15: {},
    )

    workbook_path = project / "stock_tax_export.xlsx"
    before_mtime = workbook_path.stat().st_mtime_ns

    client = TestClient(create_app(project_dir=project))
    result = client.app.state.runtime.calculate(write_workbook=True)
    assert result.app_status.global_status == "blocked"
    assert workbook_path.stat().st_mtime_ns == before_mtime

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=workbook_path,
        fetch_missing_fx=False,
    )
    assert calc.calculation_blocked is True
    with pytest.raises(RuntimeError, match="required FX rates are missing"):
        workbook_module.write_calculation_result(calc)


def test_api_rejects_2024_method_change(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch("/api/years/2024", json={"method": "FIFO"})
    assert response.status_code == 409
    assert "2024 is locked" in response.json()["detail"]


def test_build_locked_years_allows_explicit_unlock_of_filed_year():
    locked_years = workbook_module.build_locked_years(
        {"Locked_Years": [{"Tax year": 2024, "Locked?": False}]},
        [2024],
    )

    assert locked_years[2024] is False
    assert policy.check_unlock(2024) is None


def test_locked_year_snapshot_rebuild_required_when_earlier_year_locked_under_later_snapshot(tmp_path):
    project = _copy_project_fixture(tmp_path)
    _set_locked_year(project, 2020, True)

    rows = _build_check_rows_for_project(project)
    rebuild_rows = [
        row for row in rows
        if row.get("Category") == "locked_year_snapshot_rebuild_required"
    ]

    assert rebuild_rows
    detail = rebuild_rows[0]["Detail"]
    assert "Year 2020" in detail
    assert "2024" in detail
    assert "stale" in detail.lower()
    assert "rebuild/recalculate frozen snapshots from 2020 onward is required" in detail.lower()


def test_stale_frozen_snapshot_manifest_persists_rebuild_required_check(tmp_path):
    project = _copy_project_fixture(tmp_path)
    _set_locked_year(project, 2020, True)

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=project / "stock_tax_export.xlsx",
        fetch_missing_fx=False,
    )
    workbook_module.write_workbook(
        calc.output_path,
        calc.raw_rows,
        calc.txs,
        calc.ignored,
        calc.problems,
        calc.instrument_map,
        calc.fx_yearly,
        calc.fx_daily,
        calc.fx_daily_sources,
        calc.corporate_actions,
        calc.method_selection,
        calc.locked_years,
        calc.settings,
        calc.frozen_inventory,
        calc.frozen_matching,
        calc.frozen_snapshots,
        calc.fx,
        calc.lots_final,
        calc.match_lines,
        calc.sim_warnings,
        calc.yearly_summary,
        calc.method_comparison,
        calc.split_warnings,
        calc.year_end_inventory,
        calc.import_log,
        calc.review_state,
        calc.filed_reconciliation,
        fx_yearly_sources=calc.fx_yearly_sources,
    )

    rows = _build_check_rows_for_project(project)
    rebuild_rows = [
        row for row in rows
        if row.get("Category") == "locked_year_snapshot_rebuild_required"
    ]

    assert rebuild_rows
    assert any("2024" in str(row.get("Detail") or "") for row in rebuild_rows)


def test_api_patch_year_updates_method_for_unlocked_year(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch("/api/years/2025", json={"method": "MAX_GAIN"})
    assert response.status_code == 200
    body = response.json()
    assert body["year"] == 2025
    assert body["method"] == "MAX_GAIN"
    assert body["method_source"] == "project_state"

    state = project_store.load_project_state(project)
    assert state.year_settings[2025]["method"] == "MAX_GAIN"
    assert state.method_selection.get(2025, {}) == {}


def test_api_patch_year_method_migrates_legacy_uniform_instrument_rows(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    instrument_ids = {
        sale.instrument_id for sale in baseline.sales.items if sale.instrument_id
    }
    instrument_ids.update(
        position.instrument_id for position in baseline.open_positions.items if position.instrument_id
    )
    project_store.save_project_state(
        project,
        ProjectState(
            method_selection={2025: {instrument_id: "MAX_GAIN" for instrument_id in instrument_ids}}
        ),
    )

    client = TestClient(create_app(project_dir=project))
    response = client.patch("/api/years/2025", json={"method": "LIFO"})
    assert response.status_code == 200
    body = response.json()
    assert body["method"] == "LIFO"

    state = project_store.load_project_state(project)
    assert state.year_settings[2025]["method"] == "LIFO"
    assert state.method_selection.get(2025, {}) == {}


def test_api_patch_year_method_supports_year_without_known_instruments(tmp_path):
    project = _copy_project_fixture(tmp_path)
    project_store.save_project_state(
        project,
        ProjectState(
            year_settings={
                2030: {
                    "tax_rate": 0.15,
                    "fx_method": "FX_UNIFIED_GFR",
                    "apply_100k": False,
                }
            }
        ),
    )

    client = TestClient(create_app(project_dir=project))
    response = client.patch("/api/years/2030", json={"method": "MAX_GAIN"})
    assert response.status_code == 200
    body = response.json()
    assert body["year"] == 2030
    assert body["method"] == "MAX_GAIN"
    assert body["method_source"] == "project_state"

    years = client.get("/api/years")
    assert years.status_code == 200
    year_2030 = next(year for year in years.json()["items"] if year["year"] == 2030)
    assert year_2030["method"] == "MAX_GAIN"
    assert year_2030["method_source"] == "project_state"

    state = project_store.load_project_state(project)
    assert state.year_settings[2030]["method"] == "MAX_GAIN"
    assert state.method_selection.get(2030, {}) == {}


def test_api_patch_year_updates_tax_rate_for_unlocked_year(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch("/api/years/2025", json={"tax_rate": 0.17})
    assert response.status_code == 200
    body = response.json()
    assert body["year"] == 2025
    assert body["tax_rate"] == pytest.approx(0.17)
    assert body["settings_source"] == "project_state"

    state = project_store.load_project_state(project)
    assert state.year_settings[2025]["tax_rate"] == pytest.approx(0.17)


def test_api_patch_year_updates_fx_method_for_unlocked_year(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch("/api/years/2025", json={"fx_method": "FX_DAILY_CNB"})
    assert response.status_code == 200
    body = response.json()
    assert body["year"] == 2025
    assert body["fx_method"] == "FX_DAILY_CNB"
    assert body["settings_source"] == "project_state"

    state = project_store.load_project_state(project)
    assert state.year_settings[2025]["fx_method"] == "FX_DAILY_CNB"


def test_api_patch_year_rejects_invalid_fx_method(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch("/api/years/2025", json={"fx_method": "NOT_A_FX_METHOD"})
    assert response.status_code == 422
    assert "Unsupported fx_method" in response.json()["detail"]


def test_api_patch_year_rejects_invalid_method(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch("/api/years/2025", json={"method": "NOT_A_METHOD"})
    assert response.status_code == 422
    assert "Unsupported method" in response.json()["detail"]


@pytest.mark.parametrize("tax_rate", ["0.2", -0.01, True, None])
def test_api_patch_year_rejects_invalid_tax_rate(tmp_path, tax_rate):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch("/api/years/2025", json={"tax_rate": tax_rate})
    if tax_rate is None:
        assert response.status_code == 400
        assert "No editable year fields" in response.json()["detail"]
    else:
        assert response.status_code == 422
        assert "tax_rate" in response.json()["detail"]


def test_year_settings_patch_survives_recalc_and_runtime_reload(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch(
        "/api/years/2025",
        json={
            "method": "MIN_GAIN",
            "fx_method": "FX_UNIFIED_GFR",
            "tax_rate": 0.19,
            "apply_100k_exemption": True,
        },
    )
    assert response.status_code == 200

    recalculated = client.app.state.runtime.calculate(write_workbook=False)
    year_after_recalc = next(year for year in recalculated.tax_years.items if year.year == 2025)
    assert year_after_recalc.method == "MIN_GAIN"
    assert year_after_recalc.fx_method == "FX_UNIFIED_GFR"
    assert year_after_recalc.tax_rate == pytest.approx(0.19)
    assert year_after_recalc.exemption_100k is True

    reloaded_client = TestClient(create_app(project_dir=project))
    reloaded_years = reloaded_client.get("/api/years")
    assert reloaded_years.status_code == 200
    year_after_reload = next(year for year in reloaded_years.json()["items"] if year["year"] == 2025)
    assert year_after_reload["method"] == "MIN_GAIN"
    assert year_after_reload["fx_method"] == "FX_UNIFIED_GFR"
    assert year_after_reload["tax_rate"] == pytest.approx(0.19)
    assert year_after_reload["exemption_100k"] is True


def test_get_years_reflects_project_state_values_and_provenance_after_patch(tmp_path):
    project = _copy_project_fixture(tmp_path)
    client = TestClient(create_app(project_dir=project))

    response = client.patch(
        "/api/years/2025",
        json={"method": "LIFO", "tax_rate": 0.23, "fx_method": "FX_UNIFIED_GFR"},
    )
    assert response.status_code == 200

    years = client.get("/api/years")
    assert years.status_code == 200
    year_2025 = next(year for year in years.json()["items"] if year["year"] == 2025)
    assert year_2025["method"] == "LIFO"
    assert year_2025["tax_rate"] == pytest.approx(0.23)
    assert year_2025["fx_method"] == "FX_UNIFIED_GFR"
    assert year_2025["settings_source"] == "project_state"
    assert year_2025["method_source"] == "project_state"


def test_locked_output_fails_without_alternate_workbook(tmp_path, monkeypatch):
    out_path = tmp_path / "stock_tax_system.xlsx"
    temp_path = tmp_path / ".stock_tax_system.tmp.xlsx"
    temp_path.write_text("temp", encoding="utf-8")

    def raise_permission_error(src, dst):
        raise PermissionError("locked")

    monkeypatch.setattr(workbook_module.os, "replace", raise_permission_error)

    with pytest.raises(RuntimeError) as excinfo:
        workbook_module._replace_output_or_fail(temp_path, out_path)

    assert str(excinfo.value) == (
        "Cannot write stock_tax_system.xlsx because it is open or locked. "
        "Close Excel and rerun."
    )
    assert not temp_path.exists()
    assert not any(tmp_path.glob("stock_tax_system_new*.xlsx"))
