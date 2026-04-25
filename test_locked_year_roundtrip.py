#!/usr/bin/env python3
"""End-to-end check for soft-lock semantics in a temp sandbox.

Locking an earlier year beneath an existing later frozen snapshot must
surface an explicit stale-snapshot rebuild requirement instead of
failing later with confusing unmatched sells.
"""
from __future__ import annotations

import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent


def make_sandbox() -> tuple[tempfile.TemporaryDirectory[str], Path, list[str]]:
    tmpdir = tempfile.TemporaryDirectory(prefix="locked_year_roundtrip_")
    sandbox = Path(tmpdir.name)
    shutil.copytree(ROOT / ".csv", sandbox / ".csv")
    workbook_path = sandbox / "stock_tax_system.xlsx"
    inputs = [
        str(sandbox / ".csv" / name)
        for name in ["XTB_CZK.csv", "XTB_USD.csv", "Lynx.csv", "Revolut.csv", "Trading212.csv"]
    ]
    return tmpdir, workbook_path, inputs


def run_build(workbook_path: Path, inputs: list[str], *, expect_success: bool = True) -> subprocess.CompletedProcess[str]:
    cmd = [sys.executable, str(ROOT / "build_stock_tax_workbook.py"),
           "--input", *inputs, "--output", str(workbook_path)]
    res = subprocess.run(cmd, capture_output=True, text=True)
    if expect_success and res.returncode != 0:
        print("BUILD FAILED:\n" + res.stdout + res.stderr)
        sys.exit(1)
    return res


def set_cell(path: str, sheet: str, row: int, col: int, value) -> None:
    wb = load_workbook(path)
    wb[sheet].cell(row=row, column=col, value=value)
    wb.save(path)


def read_cell(path: str, sheet: str, row: int, col: int):
    return load_workbook(path)[sheet].cell(row=row, column=col).value


def read_check_rows(path: str):
    ws = load_workbook(path)["Checks"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(value is None for value in row):
            continue
        out.append({
            "Severity": row[0],
            "Category": row[1],
            "Detail": row[2],
        })
    return out


def find_row(path: str, sheet: str, header_row: int, key_col: int, key):
    wb = load_workbook(path)
    ws = wb[sheet]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= header_row:
            continue
        if row[key_col - 1] == key:
            return i
    return None


def main() -> int:
    tmpdir, workbook_path, inputs = make_sandbox()
    with tmpdir:
        # Pass 1: baseline build with clean defaults.
        print("== Pass 1: initial build ==")
        run_build(workbook_path, inputs)
        summary_row_2020 = find_row(str(workbook_path), "Yearly_Tax_Summary", 1, 1, 2020)
        assert summary_row_2020, "2020 not in yearly summary after pass 1"
        tax_2020_pass1 = read_cell(str(workbook_path), "Yearly_Tax_Summary",
                                   summary_row_2020, 13)
        print(f"  Yearly_Tax_Summary 2020 tax = {tax_2020_pass1}")

        # Lock 2020 on Locked_Years sheet (header on row 3, 2020 first data row).
        ly_row = find_row(str(workbook_path), "Locked_Years", 3, 1, 2020)
        assert ly_row, "Locked_Years 2020 not found"
        set_cell(str(workbook_path), "Locked_Years", ly_row, 2, True)

        # Pass 2: rebuild with an earlier lock beneath the existing 2024 snapshot.
        print("\n== Pass 2: rebuild with 2020 locked under existing 2024 snapshot ==")
        res = run_build(workbook_path, inputs, expect_success=False)
        combined = res.stdout + res.stderr
        if res.returncode == 0:
            print("FAIL: expected controlled rebuild-required failure, but build succeeded.")
            return 1
        if "locked_year_snapshot_rebuild_required" not in combined:
            print("FAIL: missing explicit locked_year_snapshot_rebuild_required failure output.")
            return 1
        if "Year 2020" not in combined:
            print("FAIL: rebuild-required failure did not mention changed year 2020.")
            return 1
        if "2024" not in combined:
            print("FAIL: rebuild-required failure did not mention later snapshot 2024.")
            return 1
        print("  Controlled failure surfaced explicit stale snapshot guidance.")

        # Pass 3: explicit unlock remains available as an operator action.
        print("\n== Pass 3: unlock 2020 explicitly and rebuild ==")
        set_cell(str(workbook_path), "Locked_Years", ly_row, 2, False)
        run_build(workbook_path, inputs)
        print("PASS: soft-lock roundtrip stayed inside sandbox and surfaced explicit rebuild guidance.")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
