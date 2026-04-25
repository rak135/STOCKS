"""Workbook writing helpers.

Behavior-preserving extraction from build_stock_tax_workbook.py.
Contains all openpyxl sheet-writing logic, write_workbook, and
write_calculation_result.

Callable utilities (safe_float, parse_trade_date) are injected via
write_workbook() to avoid importing from build_stock_tax_workbook.py.
Monolith-defined types (Lot, Transaction, MatchLine, RawRow,
CalculationResult) are accepted as Any.
"""

from __future__ import annotations

import datetime as _dt
import os
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from stock_tax_app.engine import policy, ui_state
from stock_tax_app.engine.checks import build_check_rows as _engine_build_check_rows
from stock_tax_app.engine.fx import FXResolver, SUPPORTED_FX_METHODS, DEFAULT_FX_YEARLY, GFR_OFFICIAL_RATES
from stock_tax_app.engine.tax_summary import (
    DEFAULT_TAX_RATE,
    DEFAULT_APPLY_100K,
    DEFAULT_100K_THRESHOLD,
    DEFAULT_FX_METHOD,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

YAHOO_COLUMNS = [
    "Symbol", "Current Price", "Date", "Time", "Change", "Open", "High",
    "Low", "Volume", "Trade Date", "Purchase Price", "Quantity",
    "Commission", "High Limit", "Low Limit", "Comment", "Transaction Type",
]

CA_TYPES = ("SPLIT", "REVERSE_SPLIT", "TICKER_CHANGE")
FILED_YEARS = policy.FILED_YEARS


def _to_float(value: Any, default: float) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default

# ---------------------------------------------------------------------------
# Styling constants (mirrors those in build_stock_tax_workbook.py)
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF2F5496")
SUBHEADER_FILL = PatternFill("solid", fgColor="FFDEEBF7")
EDITABLE_FILL = PatternFill("solid", fgColor="FFFFF2CC")
LOCKED_FILL = PatternFill("solid", fgColor="FFD9D9D9")
ERROR_FILL = PatternFill("solid", fgColor="FFF8CBAD")
WARNING_FILL = PatternFill("solid", fgColor="FFFFE699")
OK_FILL = PatternFill("solid", fgColor="FFC6EFCE")

THIN = Side(style="thin", color="FFB4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Sheet utility helpers
# ---------------------------------------------------------------------------

def autosize_columns(ws, min_width: int = 10, max_width: int = 42) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def write_header(ws, headers: List[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def add_table(ws, name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False,
                                        showLastColumn=False,
                                        showRowStripes=True,
                                        showColumnStripes=False)
    ws.add_table(tab)


def build_open_position_rows(
    raw_rows: List[Any],
    instrument_map: Dict[str, Dict[str, str]],
    lots: List[Any],
    *,
    safe_float: Callable,
    parse_trade_date: Callable,
    ok_tolerance: float = 1e-4,
    warn_tolerance: float = 1e-2,
) -> List[Dict[str, Any]]:
    yahoo, position_provenance = extract_position_rows_with_provenance(
        raw_rows, instrument_map,
        safe_float=safe_float, parse_trade_date=parse_trade_date)
    calc: Dict[str, float] = defaultdict(float)
    for lot in lots:
        if lot.quantity_remaining > 1e-9:
            calc[lot.instrument_id] += lot.quantity_remaining
    instruments = sorted(set(yahoo.keys()) | set(calc.keys()))
    rows: List[Dict[str, Any]] = []
    for inst in instruments:
        yq = yahoo.get(inst)
        cq = calc.get(inst, 0.0)
        sources = position_provenance.get(inst, [])
        source_count = len(sources)
        primary_source = sources[0] if sources else {}

        if yq is None:
            source_status = "unknown"
            source_reason = "No broker/Yahoo position row is available for this instrument."
        else:
            source_status = "ready"
            source_reason_parts: List[str] = []
            missing_snapshot_dates = [src for src in sources if src.get("snapshot_date") is None]
            if missing_snapshot_dates:
                source_status = "partial"
                source_reason_parts.append(
                    "Snapshot date is unavailable in one or more source rows."
                )
            if source_count > 1:
                source_status = "partial"
                source_reason_parts.append(
                    f"Reported quantity is aggregated from {source_count} source rows."
                )
            source_reason = " ".join(source_reason_parts) if source_reason_parts else None

        if yq is None:
            diff = None
            status = "UNKNOWN"
        else:
            diff = cq - yq
            status = "OK" if abs(diff) <= ok_tolerance else (
                "WARN" if abs(diff) <= warn_tolerance else "ERROR"
            )
        rows.append({
            "Instrument_ID": inst,
            "Reported qty": yq,
            "Yahoo qty": yq,
            "Calculated qty": cq,
            "Difference": diff,
            "Status": status,
            "Reported position source file": primary_source.get("source_file"),
            "Reported position source row": primary_source.get("source_row"),
            "Reported position broker": primary_source.get("broker"),
            "Reported position account": primary_source.get("account"),
            "Reported position snapshot date": primary_source.get("snapshot_date"),
            "Reported position source type": (
                primary_source.get("source_type") if sources else "csv_position_row"
            ),
            "Reported position source status": source_status,
            "Reported position source reason": source_reason,
            "Reported position source count": source_count,
            "Reported position sources": sources,
        })
    return rows



def write_calculation_result(
    result: Any,
    *,
    backup_existing: bool = False,
    safe_float: Callable,
    parse_trade_date: Callable,
    supported_methods: Optional[Tuple[str, ...]] = None,
) -> Path:
    if result.calculation_blocked:
        raise RuntimeError(
            "Cannot write workbook because required FX rates are missing. "
            "Fix the FX data and rerun; no fallback yearly or 22.0 rate was used."
        )
    out_path = result.output_path
    temp_out_path = _tmp_output_path(out_path)
    if temp_out_path.exists():
        temp_out_path.unlink(missing_ok=True)

    write_workbook(
        out_path=temp_out_path,
        raw_rows=result.raw_rows,
        txs=result.txs,
        ignored=result.ignored,
        problems=result.problems,
        instrument_map=result.instrument_map,
        fx_yearly=result.fx_yearly,
        fx_yearly_sources=result.fx_yearly_sources,
        fx_daily=result.fx_daily,
        fx_daily_sources=result.fx_daily_sources,
        corporate_actions=result.corporate_actions,
        method_selection=result.method_selection,
        locked_years=result.locked_years,
        settings=result.settings,
        frozen_inventory=result.frozen_inventory,
        frozen_matching=result.frozen_matching,
        frozen_snapshots=result.frozen_snapshots,
        fx=result.fx,
        lots_final=result.lots_final,
        match_lines=result.match_lines,
        sim_warnings=result.sim_warnings,
        yearly_summary=result.yearly_summary,
        method_comparison=result.method_comparison,
        split_warnings=result.split_warnings,
        year_end_inventory=result.year_end_inventory,
        import_log=result.import_log,
        review_state=result.review_state,
        filed_reconciliation=result.filed_reconciliation,
        safe_float=safe_float,
        parse_trade_date=parse_trade_date,
        supported_methods=supported_methods,
    )

    import verify_workbook as _verify_workbook

    verify_rc = _verify_workbook.main(str(temp_out_path))
    if verify_rc != 0:
        temp_out_path.unlink(missing_ok=True)
        raise RuntimeError(
            "Validation failed for temporary workbook; requested output was not replaced."
        )

    if backup_existing and out_path.exists():
        _backup_existing_output(out_path)

    _replace_output_or_fail(temp_out_path, out_path)
    return out_path


# -----------------------------------------------------------------------
# Workbook writing
# -----------------------------------------------------------------------

def write_workbook(
    out_path: Path,
    raw_rows: List[Any],
    txs: List[Any],
    ignored: List[Dict[str, Any]],
    problems: List[Dict[str, Any]],
    instrument_map: Dict[str, Dict[str, str]],
    fx_yearly: Dict[int, float],
    fx_daily: Dict[date, float],
    fx_daily_sources: Dict[date, str],
    corporate_actions: List[Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    locked_years: Dict[int, bool],
    settings: Dict[int, Dict[str, Any]],
    frozen_inventory: Dict[int, List[Dict[str, Any]]],
    frozen_matching: Dict[int, List[Dict[str, Any]]],
    frozen_snapshots: Dict[int, Dict[str, Any]],
    fx: FXResolver,
    lots_final: List[Any],
    match_lines: List[Any],
    sim_warnings: List[Dict[str, Any]],
    yearly_summary: List[Dict[str, Any]],
    method_comparison: List[Dict[str, Any]],
    split_warnings: List[Dict[str, Any]],
    year_end_inventory: Dict[int, List[Any]],
    import_log: List[Dict[str, Any]],
    review_state: Dict[str, Dict[str, Any]],
    filed_reconciliation: Dict[int, Dict[str, Any]],
    fx_yearly_sources: Optional[Dict[int, str]] = None,
    *,
    safe_float: Callable,
    parse_trade_date: Callable,
    supported_methods: Optional[Tuple[str, ...]] = None,
) -> None:
    if supported_methods is None:
        supported_methods = policy.SUPPORTED_METHODS
    wb = Workbook()
    # Remove default sheet — we'll add ours in order.
    wb.remove(wb.active)

    _write_readme(wb)
    _write_dashboard(wb, yearly_summary, method_comparison, sim_warnings,
                     problems, import_log)
    _write_settings(wb, settings, sorted(settings.keys()))
    _write_import_log(wb, import_log)
    _write_raw_yahoo(wb, raw_rows)
    _write_ignored(wb, ignored)
    _write_transactions(wb, txs)
    _write_instrument_map(wb, instrument_map)
    _write_fx_daily(wb, fx_daily, fx_daily_sources)
    _write_fx_yearly(wb, fx_yearly, sorted(settings.keys()), fx_yearly_sources)
    _write_corporate_actions(wb, corporate_actions)
    _write_split_audit(wb, split_warnings)
    _write_method_selection(wb, method_selection, supported_methods)
    instrument_ids = sorted({t.instrument_id for t in txs})
    _write_method_plan(
        wb,
        sorted(settings.keys()),
        instrument_ids,
        method_selection,
        settings,
        method_comparison,
        match_lines,
    )
    _write_filed_year_reconciliation(
        wb,
        yearly_summary,
        method_selection,
        instrument_ids,
        filed_reconciliation,
    )
    _write_locked_years(wb, locked_years, sorted(settings.keys()))
    _write_frozen_inventory(wb, frozen_inventory, year_end_inventory,
                            locked_years)
    _write_frozen_matching(wb, frozen_matching, match_lines, locked_years)
    _write_frozen_snapshots(wb, frozen_snapshots, year_end_inventory,
                            locked_years, match_lines, sim_warnings)
    _write_review_state(wb, review_state, match_lines)
    _write_lots(wb, lots_final)
    _write_lot_matching(wb, match_lines)
    _write_sell_review(wb, match_lines, review_state)
    _write_open_lots_review(wb, lots_final, txs, fx)
    _write_open_position_check(wb, raw_rows, instrument_map, lots_final,
                               safe_float=safe_float,
                               parse_trade_date=parse_trade_date)
    _write_yearly_summary(wb, yearly_summary)
    _write_method_comparison(wb, method_comparison)
    _write_checks(wb, sim_warnings, problems, fx_yearly, fx_daily,
                  settings, locked_years, frozen_inventory, split_warnings,
                  method_selection, yearly_summary, match_lines, lots_final,
                  year_end_inventory, frozen_snapshots, fx, supported_methods)
    _write_audit_report(wb, raw_rows, txs, ignored, match_lines,
                        yearly_summary)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _tmp_output_path(out_path: Path) -> Path:
    """Temporary workbook path in the same directory for atomic replace."""
    return out_path.parent / f".{out_path.stem}.tmp{out_path.suffix}"


def _backup_existing_output(out_path: Path) -> Path:
    backup_dir = out_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = backup_dir / f"{out_path.stem}_{ts}{out_path.suffix}"
    shutil.copy2(out_path, backup_path)
    return backup_path


def _replace_output_or_fail(temp_path: Path, out_path: Path) -> None:
    try:
        os.replace(temp_path, out_path)
    except PermissionError as exc:
        temp_path.unlink(missing_ok=True)
        raise RuntimeError(
            f"Cannot write {out_path.name} because it is open or locked. "
            "Close Excel and rerun."
        ) from exc


# ------------------- sheet writers -------------------

def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    lines = [
        ("Czech Personal Income Tax — Stock Trade Workbook",
         Font(bold=True, size=14)),
        ("", None),
        ("Purpose", Font(bold=True, size=12)),
        ("Compute Czech personal income tax from Yahoo Portfolio CSV "
         "exports.", None),
        ("This is a CALCULATION workbook. It is NOT official tax advice.",
         Font(italic=True, color="FFC00000")),
        ("", None),
        ("Key assumptions", Font(bold=True, size=12)),
        ("* Tax currency: CZK. All trade data is treated as USD-denominated.",
         None),
        ("* Trade Date (column from Yahoo CSV) is the operative date.", None),
        ("* FX fees are IGNORED. Only the Commission column is used as a "
         "broker fee.", None),
        ("* Buy commission increases acquisition cost; sell commission "
         "reduces proceeds.", None),
        ("* Matching happens GLOBALLY across brokers/accounts per "
         "Instrument_ID.", None),
        ("* Each matched lot is evaluated independently against the Czech "
         "3-year time test.", None),
        ("* Time test: a lot is exempt if sell_date > buy_date + 3 "
         "calendar years (same calendar day, 3 years later + 1 day).", None),
        ("* Within a year, taxable losses offset taxable gains; base floored "
         "at 0. No carryforward.", None),
        ("* 100,000 CZK annual gross proceeds exemption is OPTIONAL per "
         "year (Settings.Apply 100k exemption).", None),
        ("* Locked years are not recalculated. Frozen_Inventory captures "
         "open lots at year-end.", None),
        ("* Corporate actions (splits, reverse splits, ticker changes) are "
         "ENTERED MANUALLY.", None),
        ("* Default tax rate is 15 %; editable per year in Settings.", None),
        ("* Default FX method is FX_UNIFIED_GFR (GFŘ official rate). Switch per year in Settings; "
         "do not mix within one year.", None),
        ("", None),
        ("How to use", Font(bold=True, size=12)),
        ("1. Fill FX_Yearly (and FX_Daily if using FX_DAILY_CNB) against "
         "Czech National Bank tables.", None),
        ("2. Review Instrument_Map and add ISIN / stable Instrument_IDs "
         "where needed.", None),
        ("3. Add any splits / reverse splits / ticker changes to "
         "Corporate_Actions.", None),
        ("4. Pick a matching method per (year, instrument) in "
         "Method_Selection (default FIFO).", None),
        ("5. Open Settings and confirm tax rate, FX method and "
         "100k toggle per year.", None),
        ("6. Re-run build_stock_tax_workbook.py — it preserves locked-year "
         "snapshots.", None),
        ("7. For filed years: set Locked? = TRUE on the Locked_Years "
         "sheet. The workbook will freeze that year's inventory and "
         "matching on next regeneration.", None),
        ("", None),
        ("Traceability", Font(bold=True, size=12)),
        ("Every final tax number in Yearly_Tax_Summary is traceable "
         "through Lot_Matching to the original transaction row "
         "(Source file + Source row).", None),
        ("", None),
        ("Sheet index", Font(bold=True, size=12)),
    ]
    for title in [
        "README", "Operator_Dashboard", "Settings", "Import_Log",
        "Raw_Yahoo", "Ignored_Rows", "Transactions", "Instrument_Map",
        "FX_Daily", "FX_Yearly", "Corporate_Actions", "Split_Audit",
        "Method_Selection", "Locked_Years", "Frozen_Inventory",
        "Frozen_Lot_Matching", "Frozen_Snapshots", "Lots", "Lot_Matching",
        "Yearly_Tax_Summary", "Method_Comparison", "Checks",
        "Audit_Report",
    ]:
        lines.append((f"  - {title}", None))

    for i, (text, fnt) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if fnt is not None:
            cell.font = fnt
        cell.alignment = WRAP_LEFT
    ws.column_dimensions["A"].width = 96
    ws.freeze_panes = "A2"


def _write_dashboard(
    wb: Workbook, yearly_summary: List[Dict[str, Any]],
    method_comparison: List[Dict[str, Any]],
    sim_warnings: List[Dict[str, Any]],
    problems: List[Dict[str, Any]],
    import_log: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Operator_Dashboard")
    ws["A1"] = "Operator Dashboard"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Totals by tax year"
    ws["A3"].font = Font(bold=True, size=12)
    headers = ["Tax year", "Gross proceeds CZK", "Tax base CZK",
               "Tax due CZK", "Locked?", "FX method", "Apply 100k?"]
    write_header(ws, headers, row=4)
    for i, row in enumerate(yearly_summary, start=5):
        ws.cell(row=i, column=1, value=row["Tax year"])
        ws.cell(row=i, column=2,
                value=row["Gross proceeds CZK (all sells)"])
        ws.cell(row=i, column=3, value=row["Final tax base CZK"])
        ws.cell(row=i, column=4, value=row["Tax due CZK"])
        ws.cell(row=i, column=5, value=bool(row["Locked?"]))
        ws.cell(row=i, column=6, value=row["FX method"])
        ws.cell(row=i, column=7, value=bool(row["Apply 100k exemption?"]))

    base = len(yearly_summary) + 7
    ws.cell(row=base, column=1, value="Method comparison (tax CZK)")
    ws.cell(row=base, column=1).font = Font(bold=True, size=12)
    headers = ["Tax year", "FIFO", "LIFO", "MIN_GAIN", "MAX_GAIN",
               "Selected", "Best method", "Delta selected-best"]
    write_header(ws, headers, row=base + 1)
    for i, row in enumerate(method_comparison, start=base + 2):
        ws.cell(row=i, column=1, value=row["Tax year"])
        ws.cell(row=i, column=2, value=row["FIFO tax CZK"])
        ws.cell(row=i, column=3, value=row["LIFO tax CZK"])
        ws.cell(row=i, column=4, value=row["MIN_GAIN tax CZK"])
        ws.cell(row=i, column=5, value=row["MAX_GAIN tax CZK"])
        ws.cell(row=i, column=6, value=row["Selected method tax CZK"])
        ws.cell(row=i, column=7, value=row["Best method"])
        ws.cell(row=i, column=8, value=row["Delta selected vs best CZK"])

    checks_base = base + len(method_comparison) + 4
    ws.cell(row=checks_base, column=1, value="Validation summary")
    ws.cell(row=checks_base, column=1).font = Font(bold=True, size=12)

    err = sum(1 for p in sim_warnings + problems if p.get("severity") == "ERROR")
    warn = sum(1 for p in sim_warnings + problems if p.get("severity") == "WARN")
    ws.cell(row=checks_base + 1, column=1, value="Errors")
    ws.cell(row=checks_base + 1, column=2, value=err)
    ws.cell(row=checks_base + 2, column=1, value="Warnings")
    ws.cell(row=checks_base + 2, column=2, value=warn)
    if err > 0:
        ws.cell(row=checks_base + 1, column=2).fill = ERROR_FILL
    if warn > 0:
        ws.cell(row=checks_base + 2, column=2).fill = WARNING_FILL

    ws.cell(row=checks_base + 4, column=1,
            value="Check the Checks sheet and Audit_Report sheet for "
                  "details. This dashboard is a high-level overview only.")
    ws.cell(row=checks_base + 4, column=1).font = Font(italic=True)

    import_base = checks_base + 6
    ws.cell(row=import_base, column=1, value="Last import run")
    ws.cell(row=import_base, column=1).font = Font(bold=True, size=12)
    headers = ["Source file", "Raw rows", "Transactions", "Ignored", "Broker"]
    write_header(ws, headers, row=import_base + 1)
    for i, row in enumerate(import_log, start=import_base + 2):
        ws.cell(row=i, column=1, value=row["Source file"])
        ws.cell(row=i, column=2, value=row["Raw rows"])
        ws.cell(row=i, column=3, value=row["Transactions"])
        ws.cell(row=i, column=4, value=row["Ignored"])
        ws.cell(row=i, column=5, value=row["Broker"])

    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_settings(
    wb: Workbook, settings: Dict[int, Dict[str, Any]], years: List[int],
) -> None:
    ws = wb.create_sheet("Settings")
    headers = ["Tax year", "Tax rate", "FX method",
               "Apply 100k exemption?", "Locked year?", "Notes"]
    write_header(ws, headers)
    for i, y in enumerate(years, start=2):
        s = settings[y]
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=s["tax_rate"]).number_format = "0.00%"
        ws.cell(row=i, column=3, value=s["fx_method"])
        ws.cell(row=i, column=4, value=bool(s["apply_100k"]))
        ws.cell(row=i, column=5, value=bool(s["locked"]))
        ws.cell(row=i, column=6, value=s.get("notes", ""))
        for col in (2, 3, 4, 5, 6):
            ws.cell(row=i, column=col).fill = EDITABLE_FILL
    n = max(2, 1 + len(years))
    add_table(ws, "tbl_Settings", f"A1:F{n}")
    fx_dv = DataValidation(type="list",
                           formula1=f'"{",".join(SUPPORTED_FX_METHODS)}"',
                           allow_blank=False, showDropDown=False)
    fx_dv.add(f"C2:C{n}")
    ws.add_data_validation(fx_dv)
    bool_dv1 = DataValidation(type="list", formula1='"TRUE,FALSE"',
                              allow_blank=False, showDropDown=False)
    bool_dv1.add(f"D2:D{n}")
    ws.add_data_validation(bool_dv1)
    bool_dv2 = DataValidation(type="list", formula1='"TRUE,FALSE"',
                              allow_blank=False, showDropDown=False)
    bool_dv2.add(f"E2:E{n}")
    ws.add_data_validation(bool_dv2)
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_import_log(wb: Workbook, import_log: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Import_Log")
    headers = ["Source file", "Broker", "Account", "Raw rows",
               "Transactions", "Ignored", "Min Trade Date",
               "Max Trade Date", "Unique symbols", "Import timestamp"]
    write_header(ws, headers)
    for i, row in enumerate(import_log, start=2):
        ws.cell(row=i, column=1, value=row["Source file"])
        ws.cell(row=i, column=2, value=row["Broker"])
        ws.cell(row=i, column=3, value=row["Account"])
        ws.cell(row=i, column=4, value=row["Raw rows"])
        ws.cell(row=i, column=5, value=row["Transactions"])
        ws.cell(row=i, column=6, value=row["Ignored"])
        ws.cell(row=i, column=7, value=row["Min Trade Date"])
        ws.cell(row=i, column=8, value=row["Max Trade Date"])
        ws.cell(row=i, column=9, value=row["Unique symbols"])
        ws.cell(row=i, column=10, value=row["Import timestamp"])
    if import_log:
        add_table(ws, "tbl_ImportLog", f"A1:J{1+len(import_log)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_raw_yahoo(wb: Workbook, raw_rows: List[Any]) -> None:
    ws = wb.create_sheet("Raw_Yahoo")
    headers = ["Source file", "Source row", "Source broker",
               "Source account"] + YAHOO_COLUMNS
    write_header(ws, headers)
    for i, raw in enumerate(raw_rows, start=2):
        ws.cell(row=i, column=1, value=raw.source_file)
        ws.cell(row=i, column=2, value=raw.source_row)
        ws.cell(row=i, column=3, value=raw.source_broker)
        ws.cell(row=i, column=4, value=raw.source_account)
        for j, col in enumerate(YAHOO_COLUMNS, start=5):
            ws.cell(row=i, column=j, value=raw.data.get(col, ""))
    if raw_rows:
        add_table(ws, "tbl_Raw",
                  f"A1:{get_column_letter(4 + len(YAHOO_COLUMNS))}"
                  f"{1+len(raw_rows)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_ignored(wb: Workbook, ignored: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Ignored_Rows")
    headers = ["Source file", "Source row", "Symbol", "Reason", "Detail"]
    write_header(ws, headers)
    for i, r in enumerate(ignored, start=2):
        ws.cell(row=i, column=1, value=r["source_file"])
        ws.cell(row=i, column=2, value=r["source_row"])
        ws.cell(row=i, column=3, value=r["symbol"])
        ws.cell(row=i, column=4, value=r["reason"])
        ws.cell(row=i, column=5, value=r["detail"])
    if ignored:
        add_table(ws, "tbl_Ignored", f"A1:E{1+len(ignored)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_transactions(wb: Workbook, txs: List[Any]) -> None:
    ws = wb.create_sheet("Transactions")
    headers = ["Tx_ID", "Source file", "Source row", "Source broker",
               "Source account", "Symbol", "Instrument_ID", "Trade Date",
               "Side", "Quantity", "Price USD", "Commission USD", "Comment"]
    write_header(ws, headers)
    for i, tx in enumerate(txs, start=2):
        ws.cell(row=i, column=1, value=tx.tx_id)
        ws.cell(row=i, column=2, value=tx.source_file)
        ws.cell(row=i, column=3, value=tx.source_row)
        ws.cell(row=i, column=4, value=tx.source_broker)
        ws.cell(row=i, column=5, value=tx.source_account)
        ws.cell(row=i, column=6, value=tx.symbol)
        ws.cell(row=i, column=7, value=tx.instrument_id)
        ws.cell(row=i, column=8, value=tx.trade_date)
        ws.cell(row=i, column=8).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=9, value=tx.side)
        ws.cell(row=i, column=10, value=tx.quantity)
        ws.cell(row=i, column=11, value=tx.price_usd)
        ws.cell(row=i, column=12, value=tx.commission_usd)
        ws.cell(row=i, column=13, value=tx.comment)
    if txs:
        add_table(ws, "tbl_Transactions", f"A1:M{1+len(txs)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_instrument_map(
    wb: Workbook, instrument_map: Dict[str, Dict[str, str]],
) -> None:
    ws = wb.create_sheet("Instrument_Map")
    headers = ["Yahoo Symbol", "Instrument_ID", "ISIN",
               "Instrument name", "Notes"]
    write_header(ws, headers)
    for i, sym in enumerate(sorted(instrument_map.keys()), start=2):
        info = instrument_map[sym]
        for j, col in enumerate(headers, start=1):
            c = ws.cell(row=i, column=j, value=info.get(col, ""))
            if col in ("Instrument_ID", "ISIN", "Instrument name", "Notes"):
                c.fill = EDITABLE_FILL
    n = max(2, 1 + len(instrument_map))
    add_table(ws, "tbl_Instrument_Map", f"A1:E{n}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_fx_daily(
    wb: Workbook,
    fx_daily: Dict[date, float],
    fx_daily_sources: Optional[Dict[date, str]] = None,
) -> None:
    ws = wb.create_sheet("FX_Daily")
    ws["A1"] = ("CNB daily USD/CZK rates. Used when a tax year's FX method "
                "is FX_DAILY_CNB. Rates can be downloaded automatically by "
                "the builder, or pasted manually. Missing FX_DAILY_CNB rates "
                "on required trade dates are reported as ERROR in Checks.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:C1")
    headers = ["Date", "USD_CZK", "Source / note"]
    write_header(ws, headers, row=3)
    row = 4
    src = fx_daily_sources or {}
    for d in sorted(fx_daily.keys()):
        ws.cell(row=row, column=1, value=d)
        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2, value=fx_daily[d])
        ws.cell(row=row, column=2).number_format = "0.0000"
        ws.cell(row=row, column=3, value=src.get(d, ""))
        row += 1
    add_table(ws, "tbl_FX_Daily", f"A3:C{max(4, row-1)}")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A4"


def _write_fx_yearly(wb: Workbook, fx_yearly: Dict[int, float],
                     years: List[int],
                     fx_yearly_sources: Optional[Dict[int, str]] = None) -> None:
    ws = wb.create_sheet("FX_Yearly")
    ws["A1"] = ("Official GF\u0158 annual USD/CZK rates (FX_UNIFIED_GFR method). "
                "GF\u0159-D-65 (2024), GF\u0159-D-75 (2025) are authoritative and "
                "pre-filled. Verify other years against CNB \u2018Kurzy dev. trhu \u2013 "
                "pr\u016fm\u011br roku\u2019. Editable per year.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:C1")
    headers = ["Tax year", "USD_CZK", "Source / note"]
    write_header(ws, headers, row=3)
    src = fx_yearly_sources or {}
    all_years = sorted(set(list(fx_yearly.keys()) + years))
    for i, y in enumerate(all_years, start=4):
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=fx_yearly.get(y, DEFAULT_FX_YEARLY.get(y)))
        ws.cell(row=i, column=2).number_format = "0.0000"
        ws.cell(row=i, column=2).fill = EDITABLE_FILL
        source_label = src.get(y, "")
        if not source_label:
            source_label = "GF\u0158 official" if y in GFR_OFFICIAL_RATES else "DEFAULT \u2014 verify"
        ws.cell(row=i, column=3, value=source_label)
    add_table(ws, "tbl_FX_Yearly", f"A3:C{3+len(all_years)}")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A4"


def _write_corporate_actions(
    wb: Workbook, actions: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Corporate_Actions")
    ws["A1"] = ("Enter corporate actions manually. For a SPLIT 3-for-1, "
                "set Ratio old=1, Ratio new=3. For a REVERSE_SPLIT 1-for-10, "
                "set Ratio old=10, Ratio new=1. Splits adjust open-lot "
                "quantity and price-per-share but NOT total cost basis and "
                "NOT holding period. This workbook does NOT detect splits "
                "automatically — see Split_Audit for hints.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:I1")
    headers = ["Date", "Instrument_ID", "Action type", "Ratio old",
               "Ratio new", "Cash in lieu", "Notes", "Applied?",
               "Audit status"]
    write_header(ws, headers, row=3)
    max_row = 3
    for i, ca in enumerate(actions, start=4):
        ws.cell(row=i, column=1, value=ca["Date"])
        ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=ca["Instrument_ID"])
        ws.cell(row=i, column=3, value=ca["Action type"])
        ws.cell(row=i, column=4, value=ca["Ratio old"])
        ws.cell(row=i, column=5, value=ca["Ratio new"])
        ws.cell(row=i, column=6, value=ca.get("Cash in lieu", 0.0))
        ws.cell(row=i, column=7, value=ca.get("Notes", ""))
        ws.cell(row=i, column=8, value=bool(ca.get("Applied", True)))
        ws.cell(row=i, column=9, value="applied" if ca.get("Applied") else
                "not applied")
        max_row = i
    # Empty template rows to encourage user entry.
    start_empty = max(max_row + 1, 4)
    for i in range(start_empty, start_empty + 8):
        for col in (1, 2, 3, 4, 5, 6, 7, 8):
            ws.cell(row=i, column=col).fill = EDITABLE_FILL
        max_row = i
    add_table(ws, "tbl_Corporate_Actions", f"A3:I{max_row}")
    ca_dv = DataValidation(type="list",
                           formula1=f'"{",".join(CA_TYPES)}"',
                           allow_blank=True, showDropDown=False)
    ca_dv.add(f"C4:C{max_row}")
    ws.add_data_validation(ca_dv)
    bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"',
                             allow_blank=True, showDropDown=False)
    bool_dv.add(f"H4:H{max_row}")
    ws.add_data_validation(bool_dv)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_split_audit(wb: Workbook, warns: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Split_Audit")
    ws["A1"] = ("Heuristic only. Identifies cases where consecutive trade "
                "prices for the same instrument differ by a factor of "
                "~3x or more. If Yahoo data is ALREADY split-adjusted, DO "
                "NOT also enter the split in Corporate_Actions (double "
                "application will mis-state cost basis). If Yahoo data is "
                "raw, add the split to Corporate_Actions.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:G1")
    headers = ["Instrument_ID", "From date", "To date", "Prev price USD",
               "Next price USD", "Ratio (next/prev)", "Hint"]
    write_header(ws, headers, row=3)
    for i, w in enumerate(warns, start=4):
        ws.cell(row=i, column=1, value=w["Instrument_ID"])
        ws.cell(row=i, column=2, value=w["From date"])
        ws.cell(row=i, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=3, value=w["To date"])
        ws.cell(row=i, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=4, value=w["Prev price USD"])
        ws.cell(row=i, column=5, value=w["Next price USD"])
        ws.cell(row=i, column=6, value=w["Ratio (next/prev)"])
        ws.cell(row=i, column=7, value=w["Hint"])
    if warns:
        add_table(ws, "tbl_Split_Audit", f"A3:G{3+len(warns)}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_method_selection(
    wb: Workbook, method_selection: Dict[Tuple[int, str], str],
    supported_methods: Tuple[str, ...],
) -> None:
    ws = wb.create_sheet("Method_Selection")
    ws["A1"] = ("Pick a matching method per (Tax year, Instrument_ID). "
                "Changes here influence Lot_Matching and Yearly_Tax_Summary. "
                "Use Method_Comparison to see the tax impact.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:D1")
    headers = ["Tax year", "Instrument_ID", "Method", "Notes"]
    write_header(ws, headers, row=3)
    keys = sorted(method_selection.keys(), key=lambda k: (k[0], k[1]))
    for i, key in enumerate(keys, start=4):
        y, inst = key
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=inst)
        ws.cell(row=i, column=3, value=method_selection[key])
        ws.cell(row=i, column=3).fill = EDITABLE_FILL
    n = max(4, 3 + len(keys))
    add_table(ws, "tbl_Method_Selection", f"A3:D{n}")
    dv = DataValidation(type="list",
                        formula1=f'"{",".join(supported_methods)}"',
                        allow_blank=False, showDropDown=False)
    dv.add(f"C4:C{n}")
    ws.add_data_validation(dv)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_locked_years(
    wb: Workbook, locked_years: Dict[int, bool], years: List[int],
) -> None:
    ws = wb.create_sheet("Locked_Years")
    ws["A1"] = ("Locked years are soft locks. Use Locked? = TRUE to protect a year "
                "from accidental mutation, and explicitly unlock it when a rebuild "
                "or correction is needed.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:C1")
    headers = ["Tax year", "Locked?", "Notes"]
    write_header(ws, headers, row=3)
    for i, y in enumerate(years, start=4):
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=bool(locked_years.get(y, False)))
        ws.cell(row=i, column=2).fill = EDITABLE_FILL
        ws.cell(row=i, column=3, value="").fill = EDITABLE_FILL
    n = max(4, 3 + len(years))
    add_table(ws, "tbl_Locked_Years", f"A3:C{n}")
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"',
                        allow_blank=False, showDropDown=False)
    dv.add(f"B4:B{n}")
    ws.add_data_validation(dv)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_frozen_inventory(
    wb: Workbook,
    existing: Dict[int, List[Dict[str, Any]]],
    year_end_inventory: Dict[int, List[Any]],
    locked_years: Dict[int, bool],
) -> None:
    ws = wb.create_sheet("Frozen_Inventory")
    ws["A1"] = ("Per-lot snapshot of OPEN lots at year-end for locked years. "
                "Future calculations may seed from the latest valid frozen snapshot, "
                "but stale snapshots must be rebuilt first.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:L1")
    headers = ["Snapshot year", "Lot_ID", "Tx_ID", "Instrument_ID",
               "Source broker", "Source account", "Source file",
               "Source row", "Buy date", "Quantity original",
               "Quantity remaining", "Price per share USD",
               "Buy commission USD"]
    write_header(ws, headers, row=3)

    # Compose final set: preserve existing locked-year snapshots;
    # add new snapshot for any year flipped to Locked? in this run.
    emitted: Dict[int, List[Dict[str, Any]]] = {}
    for y, rows in existing.items():
        if locked_years.get(y):
            emitted[y] = list(rows)
    for y, inv in year_end_inventory.items():
        if not locked_years.get(y):
            continue
        if y in emitted:
            continue
        rows_for_y: List[Dict[str, Any]] = []
        for lot in inv:
            if lot.quantity_remaining <= 1e-9:
                continue
            rows_for_y.append({
                "Snapshot year": y,
                "Lot_ID": lot.lot_id,
                "Tx_ID": lot.tx_id,
                "Instrument_ID": lot.instrument_id,
                "Source broker": lot.source_broker,
                "Source account": lot.source_account,
                "Source file": lot.source_file,
                "Source row": lot.source_row,
                "Buy date": lot.buy_date,
                "Quantity original": lot.quantity_original,
                "Quantity remaining": lot.quantity_remaining,
                "Price per share USD": lot.price_per_share_usd,
                "Buy commission USD": lot.buy_commission_total_usd,
            })
        emitted[y] = rows_for_y

    i = 3
    for y in sorted(emitted.keys()):
        for rec in emitted[y]:
            i += 1
            ws.cell(row=i, column=1, value=rec["Snapshot year"])
            ws.cell(row=i, column=2, value=rec["Lot_ID"])
            ws.cell(row=i, column=3, value=rec["Tx_ID"])
            ws.cell(row=i, column=4, value=rec["Instrument_ID"])
            ws.cell(row=i, column=5, value=rec["Source broker"])
            ws.cell(row=i, column=6, value=rec["Source account"])
            ws.cell(row=i, column=7, value=rec["Source file"])
            ws.cell(row=i, column=8, value=rec["Source row"])
            ws.cell(row=i, column=9, value=rec["Buy date"])
            ws.cell(row=i, column=9).number_format = "yyyy-mm-dd"
            ws.cell(row=i, column=10, value=rec["Quantity original"])
            ws.cell(row=i, column=11, value=rec["Quantity remaining"])
            ws.cell(row=i, column=12, value=rec["Price per share USD"])
            ws.cell(row=i, column=13, value=rec["Buy commission USD"])
    last = max(4, i)
    add_table(ws, "tbl_Frozen_Inventory", f"A3:M{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_frozen_matching(
    wb: Workbook,
    existing: Dict[int, List[Dict[str, Any]]],
    current_match_lines: List[Any],
    locked_years: Dict[int, bool],
) -> None:
    ws = wb.create_sheet("Frozen_Lot_Matching")
    ws["A1"] = ("Snapshot of matched lot lines for locked years. Later runs may reuse "
                "only valid snapshots; stale snapshots must be rebuilt, not silently trusted.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:S1")

    headers = ["Tax year", "Match_ID", "Sell_ID", "Sell date",
               "Sell source broker", "Sell source file", "Sell source row",
               "Instrument_ID", "Buy Lot_ID", "Buy Tx_ID", "Buy date",
               "Buy source broker", "Buy source file", "Buy source row",
               "Quantity", "Buy price USD", "Sell price USD",
               "Allocated buy commission USD",
               "Allocated sell commission USD", "FX rate buy",
               "FX rate sell", "Cost basis CZK", "Proceeds CZK",
               "Holding days", "Time-test exempt?", "Taxable?",
               "Taxable gain CZK", "Method"]
    write_header(ws, headers, row=3)

    # Compose emitted rows: existing for locked years + snapshot current lines
    # for years newly flipped locked.
    emitted: List[Dict[str, Any]] = []
    covered_years = set()
    for y, rows in existing.items():
        if locked_years.get(y):
            covered_years.add(y)
            for r in rows:
                r = dict(r)
                r.setdefault("Tax year", y)
                emitted.append(r)
    for m in current_match_lines:
        if not locked_years.get(m.tax_year):
            continue
        if m.tax_year in covered_years:
            continue
        emitted.append({
            "Tax year": m.tax_year, "Match_ID": m.match_id,
            "Sell_ID": m.sell_tx_id, "Sell date": m.sell_date,
            "Sell source broker": m.sell_source_broker,
            "Sell source file": m.sell_source_file,
            "Sell source row": m.sell_source_row,
            "Instrument_ID": m.instrument_id,
            "Buy Lot_ID": m.buy_lot_id, "Buy Tx_ID": m.buy_tx_id,
            "Buy date": m.buy_date,
            "Buy source broker": m.buy_source_broker,
            "Buy source file": m.buy_source_file,
            "Buy source row": m.buy_source_row,
            "Quantity": m.quantity,
            "Buy price USD": m.buy_price_per_share_usd,
            "Sell price USD": m.sell_price_per_share_usd,
            "Allocated buy commission USD": m.allocated_buy_commission_usd,
            "Allocated sell commission USD": m.allocated_sell_commission_usd,
            "FX rate buy": m.fx_rate_buy, "FX rate sell": m.fx_rate_sell,
            "Cost basis CZK": m.cost_basis_czk,
            "Proceeds CZK": m.proceeds_czk,
            "Holding days": m.holding_days,
            "Time-test exempt?": m.time_test_exempt,
            "Taxable?": m.taxable,
            "Taxable gain CZK": m.taxable_gain_czk,
            "Method": m.method,
        })

    for i, r in enumerate(emitted, start=4):
        for j, h in enumerate(headers, start=1):
            v = r.get(h, "")
            ws.cell(row=i, column=j, value=v)
            if h in ("Sell date", "Buy date") and isinstance(v, date):
                ws.cell(row=i, column=j).number_format = "yyyy-mm-dd"
    last = max(4, 3 + len(emitted))
    add_table(ws, "tbl_Frozen_Lot_Matching",
              f"A3:{get_column_letter(len(headers))}{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_lots(wb: Workbook, lots: List[Any]) -> None:
    ws = wb.create_sheet("Lots")
    headers = ["Lot_ID", "Tx_ID", "Instrument_ID", "Source broker",
               "Source account", "Source file", "Source row", "Buy date",
               "Quantity original", "Quantity remaining",
               "Price per share USD", "Buy commission USD",
               "Adjustments"]
    write_header(ws, headers)
    for i, lot in enumerate(lots, start=2):
        ws.cell(row=i, column=1, value=lot.lot_id)
        ws.cell(row=i, column=2, value=lot.tx_id)
        ws.cell(row=i, column=3, value=lot.instrument_id)
        ws.cell(row=i, column=4, value=lot.source_broker)
        ws.cell(row=i, column=5, value=lot.source_account)
        ws.cell(row=i, column=6, value=lot.source_file)
        ws.cell(row=i, column=7, value=lot.source_row)
        ws.cell(row=i, column=8, value=lot.buy_date)
        ws.cell(row=i, column=8).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=9, value=lot.quantity_original)
        ws.cell(row=i, column=10, value=lot.quantity_remaining)
        ws.cell(row=i, column=11, value=lot.price_per_share_usd)
        ws.cell(row=i, column=12, value=lot.buy_commission_total_usd)
        ws.cell(row=i, column=13,
                value="; ".join(lot.adjustments) if lot.adjustments else "")
    if lots:
        add_table(ws, "tbl_Lots", f"A1:M{1+len(lots)}")
    # Conditional formatting for remaining quantity < 0 (sanity).
    if lots:
        rng = f"J2:J{1+len(lots)}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"],
                            fill=ERROR_FILL))
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_lot_matching(wb: Workbook, matches: List[Any]) -> None:
    ws = wb.create_sheet("Lot_Matching")
    headers = ["Match_ID", "Tax year", "Sell_ID", "Sell date",
               "Sell source broker", "Sell source file", "Sell source row",
               "Instrument_ID", "Buy Lot_ID", "Buy Tx_ID", "Buy date",
               "Buy source broker", "Buy source file", "Buy source row",
               "Quantity", "Buy price USD", "Sell price USD",
               "Allocated buy commission USD",
               "Allocated sell commission USD", "FX rate buy",
               "FX rate sell", "Cost basis CZK", "Proceeds CZK",
               "Holding days", "Time-test exempt?", "Taxable?",
               "Taxable gain CZK", "Method"]
    write_header(ws, headers)
    for i, m in enumerate(matches, start=2):
        values = [
            m.match_id, m.tax_year, m.sell_tx_id, m.sell_date,
            m.sell_source_broker, m.sell_source_file, m.sell_source_row,
            m.instrument_id, m.buy_lot_id, m.buy_tx_id, m.buy_date,
            m.buy_source_broker, m.buy_source_file, m.buy_source_row,
            m.quantity, m.buy_price_per_share_usd,
            m.sell_price_per_share_usd, m.allocated_buy_commission_usd,
            m.allocated_sell_commission_usd, m.fx_rate_buy, m.fx_rate_sell,
            m.cost_basis_czk, m.proceeds_czk, m.holding_days,
            m.time_test_exempt, m.taxable, m.taxable_gain_czk, m.method,
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=j, value=v)
            if j in (4, 11) and isinstance(v, date):
                c.number_format = "yyyy-mm-dd"
    if matches:
        add_table(ws, "tbl_Lot_Matching",
                  f"A1:{get_column_letter(len(headers))}{1+len(matches)}")
        # Highlight exempt rows green, taxable losses red.
        last = 1 + len(matches)
        ws.conditional_formatting.add(
            f"Y2:Y{last}",
            CellIsRule(operator="equal", formula=['"TRUE"'], fill=OK_FILL),
        )
        ws.conditional_formatting.add(
            f"AA2:AA{last}",
            CellIsRule(operator="lessThan", formula=["0"], fill=WARNING_FILL),
        )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_frozen_snapshots(
    wb: Workbook,
    existing: Dict[int, Dict[str, Any]],
    year_end_inventory: Dict[int, List[Any]],
    locked_years: Dict[int, bool],
    current_match_lines: List[Any],
    sim_warnings: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Frozen_Snapshots")
    ws["A1"] = ("Snapshot manifest for locked years. One row per locked year "
                 "confirms that the frozen state has been captured even when "
                 "open inventory is zero, and records when a later snapshot is stale.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:H1")

    headers = [
        "Snapshot year",
        "Snapshot captured?",
        "Open lots captured",
        "Match lines captured",
        "Captured at UTC",
        "Snapshot status",
        "Stale after year",
        "Status detail",
    ]
    write_header(ws, headers, row=3)

    existing_rows: Dict[int, Dict[str, Any]] = {
        int(y): dict(r) for y, r in existing.items()
    }
    stale_rows: Dict[int, Dict[str, Any]] = {}
    for year, row in existing_rows.items():
        status = str(row.get("Snapshot status") or "").strip().upper()
        if status not in {"STALE", "NEEDS_REBUILD"}:
            continue
        stale_rows[year] = {
            "changed_year": row.get("Stale after year"),
            "detail": str(row.get("Status detail") or "").strip(),
        }
    for warning in sim_warnings:
        if warning.get("check") != "locked_year_snapshot_rebuild_required":
            continue
        changed_year = warning.get("changed_year")
        detail = str(warning.get("detail") or "").strip()
        for raw_year in warning.get("stale_snapshot_years") or []:
            try:
                year = int(raw_year)
            except (TypeError, ValueError):
                continue
            stale_rows[year] = {
                "changed_year": changed_year,
                "detail": detail,
            }
    match_count_by_year: Dict[int, int] = defaultdict(int)
    for m in current_match_lines:
        match_count_by_year[m.tax_year] += 1

    emitted: List[Dict[str, Any]] = []
    now_utc = datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")
    for y in sorted(locked_years.keys()):
        if not locked_years.get(y):
            continue
        prior = existing_rows.get(y, {})
        has_new_snapshot = y in year_end_inventory
        emitted.append({
            "Snapshot year": y,
            "Snapshot captured?": True,
            "Open lots captured": (
                len([l for l in year_end_inventory.get(y, [])
                     if l.quantity_remaining > 1e-9])
                if has_new_snapshot else int(prior.get("Open lots captured") or 0)
            ),
            "Match lines captured": (
                match_count_by_year.get(y, 0)
                if has_new_snapshot else int(prior.get("Match lines captured") or 0)
            ),
            "Captured at UTC": (
                now_utc if has_new_snapshot else (prior.get("Captured at UTC") or "")
            ),
            "Snapshot status": (
                "STALE"
                if y in stale_rows
                else str(prior.get("Snapshot status") or "READY")
            ),
            "Stale after year": (
                stale_rows[y].get("changed_year", "") if y in stale_rows else ""
            ),
            "Status detail": (
                stale_rows[y].get("detail", "") if y in stale_rows else ""
            ),
        })

    for i, row in enumerate(emitted, start=4):
        for j, h in enumerate(headers, start=1):
            ws.cell(row=i, column=j, value=row.get(h, ""))
    last = max(4, 3 + len(emitted))
    add_table(ws, "tbl_Frozen_Snapshots", f"A3:H{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _symbol_by_instrument(txs: List[Any]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for tx in txs:
        out.setdefault(tx.instrument_id, tx.symbol)
    return out


def extract_position_rows(
    raw_rows: List[Any],
    instrument_map: Dict[str, Dict[str, str]],
    *,
    safe_float: Callable,
    parse_trade_date: Callable,
) -> Dict[str, float]:
    """Extract Yahoo position rows (no Trade Date/Tx Type, has Quantity)."""
    out, _ = extract_position_rows_with_provenance(
        raw_rows, instrument_map,
        safe_float=safe_float, parse_trade_date=parse_trade_date)
    return out


def extract_position_rows_with_provenance(
    raw_rows: List[Any],
    instrument_map: Dict[str, Dict[str, str]],
    *,
    safe_float: Callable,
    parse_trade_date: Callable,
) -> Tuple[Dict[str, float], Dict[str, List[Dict[str, Any]]]]:
    """Extract Yahoo position rows plus source provenance by instrument."""
    out: Dict[str, float] = defaultdict(float)
    provenance: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for raw in raw_rows:
        tt = (raw.data.get("Transaction Type") or "").strip()
        td = (raw.data.get("Trade Date") or "").strip()
        if tt or td:
            continue
        qty_raw = raw.data.get("Quantity")
        qty, ok = safe_float(qty_raw or "", default=None)
        if not ok or qty is None:
            continue
        symbol = (raw.data.get("Symbol") or "").strip()
        inst = (instrument_map.get(symbol, {}).get("Instrument_ID") or symbol)
        if not inst:
            continue
        out[inst] += qty
        snapshot_date = parse_trade_date((raw.data.get("Date") or "").strip())
        provenance[inst].append(
            {
                "source_file": raw.source_file,
                "source_row": raw.source_row,
                "broker": raw.source_broker or None,
                "account": raw.source_account or None,
                "snapshot_date": snapshot_date,
                "source_type": "csv_position_row",
            }
        )
    return dict(out), dict(provenance)


def _write_sell_review(
    wb: Workbook,
    match_lines: List[Any],
    review_state: Dict[str, Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Sell_Review")
    headers = [
        "Row type", "Tax year", "Instrument_ID", "Sell_ID", "Sell date",
        "Sell broker", "Sell source file", "Sell source row", "Sell quantity",
        "Matched quantity", "Difference", "Method used", "Time-test mixed?",
        "Total proceeds CZK", "Total cost CZK", "Taxable gain/loss CZK",
        "Buy Lot_ID", "Buy date", "Buy broker", "Buy source file",
        "Buy source row", "Quantity matched", "Buy price USD", "Sell price USD",
        "FX buy", "FX sell", "Cost basis CZK", "Proceeds CZK",
        "Holding days", "Time-test exempt?", "Review status", "Operator note",
    ]
    write_header(ws, headers)
    by_sell: Dict[str, List[Any]] = defaultdict(list)
    for m in match_lines:
        by_sell[m.sell_tx_id].append(m)

    row = 2
    for sell_id in sorted(by_sell.keys()):
        group = by_sell[sell_id]
        g0 = group[0]
        matched_qty = sum(m.quantity for m in group)
        proceeds = sum(m.proceeds_czk for m in group)
        cost = sum(m.cost_basis_czk for m in group)
        taxable_gain = sum(m.taxable_gain_czk for m in group if m.taxable)
        methods = sorted(set(m.method for m in group))
        mixed = len({bool(m.time_test_exempt) for m in group}) > 1
        state = review_state.get(ui_state.canonical_sell_id(sell_id), {})
        header_vals = [
            "SELL", g0.tax_year, g0.instrument_id, sell_id, g0.sell_date,
            g0.sell_source_broker, g0.sell_source_file, g0.sell_source_row,
            matched_qty, matched_qty, 0.0,
            ",".join(methods), mixed, round(proceeds, 2), round(cost, 2),
            round(taxable_gain, 2),
            "", "", "", "", "", "", "", "", "", "", "", "", "", "",
            state.get("review_status", ""),
            state.get("operator_note", ""),
        ]
        for c, v in enumerate(header_vals, start=1):
            ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=c).fill = SUBHEADER_FILL
        ws.cell(row=row, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=31).fill = EDITABLE_FILL
        ws.cell(row=row, column=32).fill = EDITABLE_FILL
        row += 1
        for m in group:
            vals = [
                "LOT", m.tax_year, m.instrument_id, m.sell_tx_id, m.sell_date,
                m.sell_source_broker, m.sell_source_file, m.sell_source_row,
                "", "", "", m.method, "",
                "", "", m.taxable_gain_czk,
                m.buy_lot_id, m.buy_date, m.buy_source_broker, m.buy_source_file,
                m.buy_source_row, m.quantity, m.buy_price_per_share_usd,
                m.sell_price_per_share_usd, m.fx_rate_buy, m.fx_rate_sell,
                m.cost_basis_czk, m.proceeds_czk, m.holding_days,
                m.time_test_exempt, "", "",
            ]
            for c, v in enumerate(vals, start=1):
                ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=5).number_format = "yyyy-mm-dd"
            ws.cell(row=row, column=18).number_format = "yyyy-mm-dd"
            row += 1

    last = max(2, row - 1)
    add_table(ws, "tbl_Sell_Review", f"A1:{get_column_letter(len(headers))}{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_open_lots_review(
    wb: Workbook,
    lots: List[Any],
    txs: List[Any],
    fx: FXResolver,
) -> None:
    ws = wb.create_sheet("Open_Lots_Review")
    headers = [
        "Instrument_ID", "Symbol", "Broker/source", "Lot_ID", "Buy date",
        "Orig qty", "Qty remaining", "Price per share USD",
        "Remaining cost basis USD", "Est. CZK cost basis", "Source file",
        "Source row",
    ]
    write_header(ws, headers)
    sym_map = _symbol_by_instrument(txs)
    open_lots = [l for l in lots if l.quantity_remaining > 1e-9]
    open_lots.sort(key=lambda l: (l.instrument_id, l.buy_date, l.source_file, l.source_row))
    for i, lot in enumerate(open_lots, start=2):
        remaining_cost_usd = (
            lot.quantity_remaining * lot.price_per_share_usd
            + lot.buy_commission_total_usd * (lot.quantity_remaining / lot.quantity_original)
            if lot.quantity_original > 0 else 0.0
        )
        fx_buy, _ = fx.rate_for(lot.buy_date)
        vals = [
            lot.instrument_id,
            sym_map.get(lot.instrument_id, ""),
            f"{lot.source_broker}/{lot.source_account}",
            lot.lot_id,
            lot.buy_date,
            lot.quantity_original,
            lot.quantity_remaining,
            lot.price_per_share_usd,
            remaining_cost_usd,
            remaining_cost_usd * fx_buy,
            lot.source_file,
            lot.source_row,
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=i, column=c, value=v)
        ws.cell(row=i, column=5).number_format = "yyyy-mm-dd"
    last = max(2, 1 + len(open_lots))
    add_table(ws, "tbl_Open_Lots_Review", f"A1:{get_column_letter(len(headers))}{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_open_position_check(
    wb: Workbook,
    raw_rows: List[Any],
    instrument_map: Dict[str, Dict[str, str]],
    lots: List[Any],
    *,
    safe_float: Callable,
    parse_trade_date: Callable,
) -> None:
    ws = wb.create_sheet("Open_Position_Check")
    headers = ["Instrument_ID", "Yahoo qty", "Calculated qty", "Difference", "Status"]
    write_header(ws, headers)
    rows = build_open_position_rows(
        raw_rows, instrument_map, lots,
        safe_float=safe_float, parse_trade_date=parse_trade_date)
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row["Instrument_ID"])
        ws.cell(row=i, column=2, value=row["Yahoo qty"])
        ws.cell(row=i, column=3, value=row["Calculated qty"])
        ws.cell(row=i, column=4, value=row["Difference"])
        ws.cell(row=i, column=5, value=row["Status"])
    last = max(2, 1 + len(rows))
    add_table(ws, "tbl_Open_Position_Check", f"A1:E{last}")
    ws.conditional_formatting.add(
        f"E2:E{last}",
        CellIsRule(operator="equal", formula=['"ERROR"'], fill=ERROR_FILL),
    )
    ws.conditional_formatting.add(
        f"E2:E{last}",
        CellIsRule(operator="equal", formula=['"WARN"'], fill=WARNING_FILL),
    )
    ws.conditional_formatting.add(
        f"E2:E{last}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL),
    )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_method_plan(
    wb: Workbook,
    years: List[int],
    instrument_ids: List[str],
    method_selection: Dict[Tuple[int, str], str],
    settings: Dict[int, Dict[str, Any]],
    method_comparison: List[Dict[str, Any]],
    match_lines: List[Any],
) -> None:
    ws = wb.create_sheet("Method_Plan")
    headers = [
        "Tax year", "Instrument_ID", "Selected method", "Filed?", "Locked?",
        "Best simulated method", "Selected tax CZK", "Best simulated tax CZK",
        "Delta CZK", "Action required",
    ]
    write_header(ws, headers)
    cmp_by_year = {int(r.get("Tax year")): r for r in method_comparison}
    by_yi_gain: Dict[Tuple[int, str], float] = defaultdict(float)
    for m in match_lines:
        if m.taxable:
            by_yi_gain[(m.tax_year, m.instrument_id)] += m.taxable_gain_czk

    row = 2
    for y in years:
        rcmp = cmp_by_year.get(y, {})
        best_method = rcmp.get("Best method", "")
        best_tax_year = float(rcmp.get("Best method tax CZK") or 0.0)
        for inst in instrument_ids:
            sel = method_selection.get((y, inst), policy.default_method_for(y))
            filed = policy.is_filed(y)
            locked = bool(settings.get(y, {}).get("locked", False))
            tax_rate = float(settings.get(y, {}).get("tax_rate", DEFAULT_TAX_RATE))
            yi_tax = max(0.0, by_yi_gain.get((y, inst), 0.0)) * tax_rate
            delta = yi_tax - best_tax_year
            if locked:
                action = "LOCKED - do not optimize"
            elif filed:
                action = "FILED - keep filed method"
            elif sel != best_method and best_method:
                action = "Review possible optimization"
            else:
                action = "OK"
            vals = [
                y, inst, sel, filed, locked, best_method,
                round(yi_tax, 2), round(best_tax_year, 2), round(delta, 2), action,
            ]
            for c, v in enumerate(vals, start=1):
                ws.cell(row=row, column=c, value=v)
            row += 1
    last = max(2, row - 1)
    add_table(ws, "tbl_Method_Plan", f"A1:J{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_filed_year_reconciliation(
    wb: Workbook,
    yearly_summary: List[Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    instrument_ids: List[str],
    existing: Dict[int, Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Filed_Year_Reconciliation")
    headers = [
        "Tax year", "Expected filed method", "Workbook method",
        "Workbook tax due CZK", "Filed tax due CZK", "Difference",
        "Status", "Note",
    ]
    write_header(ws, headers)
    summary_by_year = {int(r.get("Tax year")): r for r in yearly_summary}
    row = 2
    for y in sorted(FILED_YEARS.keys()):
        expected_method = policy.filed_method(y) or policy.default_method_for(y)
        methods = {
            method_selection.get((y, inst), policy.default_method_for(y))
            for inst in instrument_ids
        }
        workbook_method = methods.pop() if len(methods) == 1 else "MIXED"
        workbook_tax = float(summary_by_year.get(y, {}).get("Tax due CZK") or 0.0)
        filed_tax = _to_float(existing.get(y, {}).get("filed_tax_due"), workbook_tax)
        diff = workbook_tax - filed_tax
        status = "OK" if abs(diff) <= 0.5 and workbook_method == expected_method else "ERROR"
        note = ""
        if workbook_method != expected_method:
            note = f"Expected method {expected_method}, got {workbook_method}."
        vals = [y, expected_method, workbook_method, workbook_tax, filed_tax, diff, status, note]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        ws.cell(row=row, column=5).fill = EDITABLE_FILL
        row += 1
    last = max(2, row - 1)
    add_table(ws, "tbl_Filed_Year_Reconciliation", f"A1:H{last}")
    ws.conditional_formatting.add(
        f"G2:G{last}",
        CellIsRule(operator="equal", formula=['"ERROR"'], fill=ERROR_FILL),
    )
    ws.conditional_formatting.add(
        f"G2:G{last}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL),
    )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_review_state(
    wb: Workbook,
    review_state: Dict[str, Dict[str, Any]],
    match_lines: List[Any],
) -> None:
    ws = wb.create_sheet("Review_State")
    headers = ["Sell_ID", "Review status", "Operator note"]
    write_header(ws, headers)
    sell_id_map = {
        ui_state.canonical_sell_id(m.sell_tx_id): m.sell_tx_id
        for m in match_lines
    }
    sell_ids = sorted(
        set(sell_id_map.values()) | {sell_id_map.get(sid, sid) for sid in review_state.keys()}
    )
    for i, sid in enumerate(sell_ids, start=2):
        state = review_state.get(ui_state.canonical_sell_id(sid), {})
        ws.cell(row=i, column=1, value=sid)
        ws.cell(row=i, column=2, value=state.get("review_status", ""))
        ws.cell(row=i, column=3, value=state.get("operator_note", ""))
        ws.cell(row=i, column=2).fill = EDITABLE_FILL
        ws.cell(row=i, column=3).fill = EDITABLE_FILL
    last = max(2, 1 + len(sell_ids))
    add_table(ws, "tbl_Review_State", f"A1:C{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_yearly_summary(
    wb: Workbook, rows: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Yearly_Tax_Summary")
    headers = ["Tax year", "Gross proceeds CZK (all sells)",
               "Time-test exempt proceeds CZK",
               "Non-exempt cost basis CZK", "Non-exempt proceeds CZK",
               "Taxable gains CZK", "Taxable losses CZK",
               "Pre-exemption tax base CZK",
               "100k threshold met?", "Apply 100k exemption?",
               "Final tax base CZK", "Tax rate", "Tax due CZK",
               "FX method", "Locked?", "Match line count", "Note"]
    write_header(ws, headers)
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            v = row.get(h, "")
            c = ws.cell(row=i, column=j, value=v)
            if h == "Tax rate":
                c.number_format = "0.00%"
            elif "CZK" in h:
                c.number_format = "#,##0.00"
    if rows:
        add_table(ws, "tbl_Yearly_Summary",
                  f"A1:{get_column_letter(len(headers))}{1+len(rows)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_method_comparison(
    wb: Workbook, rows: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Method_Comparison")
    headers = ["Tax year",
               "FIFO tax base CZK", "FIFO tax CZK",
               "LIFO tax base CZK", "LIFO tax CZK",
               "MIN_GAIN tax base CZK", "MIN_GAIN tax CZK",
               "MAX_GAIN tax base CZK", "MAX_GAIN tax CZK",
               "Selected method tax base CZK",
               "Selected method tax CZK",
               "Best method", "Best method tax CZK",
               "Delta selected vs best CZK"]
    write_header(ws, headers)
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=i, column=j, value=row.get(h, ""))
            if "CZK" in h:
                c.number_format = "#,##0.00"
    if rows:
        add_table(ws, "tbl_Method_Comparison",
                  f"A1:{get_column_letter(len(headers))}{1+len(rows)}")
        last = 1 + len(rows)
        ws.conditional_formatting.add(
            f"N2:N{last}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=WARNING_FILL),
        )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_checks(
    wb: Workbook,
    sim_warnings: List[Dict[str, Any]],
    problems: List[Dict[str, Any]],
    fx_yearly: Dict[int, float], fx_daily: Dict[date, float],
    settings: Dict[int, Dict[str, Any]],
    locked_years: Dict[int, bool],
    frozen_inventory: Dict[int, List[Dict[str, Any]]],
    split_warnings: List[Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    yearly_summary: List[Dict[str, Any]],
    match_lines: List[Any],
    lots_final: List[Any],
    year_end_inventory: Dict[int, List[Any]],
    frozen_snapshots: Dict[int, Dict[str, Any]],
    fx: FXResolver,
    supported_methods: Tuple[str, ...],
) -> None:
    ws = wb.create_sheet("Checks")
    headers = ["Severity", "Category", "Detail", "Source file",
               "Source row"]
    write_header(ws, headers)
    rows = _engine_build_check_rows(
        sim_warnings=sim_warnings,
        problems=problems,
        fx_yearly=fx_yearly,
        fx_daily=fx_daily,
        settings=settings,
        locked_years=locked_years,
        frozen_inventory=frozen_inventory,
        split_warnings=split_warnings,
        method_selection=method_selection,
        yearly_summary=yearly_summary,
        match_lines=match_lines,
        lots_final=lots_final,
        year_end_inventory=year_end_inventory,
        frozen_snapshots=frozen_snapshots,
        fx=fx,
        supported_methods=supported_methods,
    )

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row["Severity"])
        ws.cell(row=i, column=2, value=row["Category"])
        ws.cell(row=i, column=3, value=row["Detail"])
        ws.cell(row=i, column=4, value=row["Source file"])
        ws.cell(row=i, column=5, value=row["Source row"])
    add_table(ws, "tbl_Checks", f"A1:E{1+len(rows)}")
    last = 1 + len(rows)
    ws.conditional_formatting.add(
        f"A2:A{last}",
        FormulaRule(formula=['EXACT($A2,"ERROR")'], fill=ERROR_FILL))
    ws.conditional_formatting.add(
        f"A2:A{last}",
        FormulaRule(formula=['EXACT($A2,"WARN")'], fill=WARNING_FILL))
    ws.conditional_formatting.add(
        f"A2:A{last}",
        FormulaRule(formula=['EXACT($A2,"INFO")'], fill=SUBHEADER_FILL))
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_audit_report(
    wb: Workbook,
    raw_rows: List[Any], txs: List[Any],
    ignored: List[Dict[str, Any]], match_lines: List[Any],
    yearly_summary: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Audit_Report")
    ws["A1"] = "Audit Report — end-to-end traceability summary"
    ws["A1"].font = Font(bold=True, size=14)

    # Per source file totals.
    per_file: Dict[str, Dict[str, int]] = defaultdict(
        lambda: {"raw": 0, "tx": 0, "ignored": 0})
    for r in raw_rows:
        per_file[r.source_file]["raw"] += 1
    for t in txs:
        per_file[t.source_file]["tx"] += 1
    for ig in ignored:
        per_file[ig["source_file"]]["ignored"] += 1

    ws["A3"] = "Per source file"
    ws["A3"].font = Font(bold=True, size=12)
    headers = ["Source file", "Raw rows", "Transactions", "Ignored"]
    write_header(ws, headers, row=4)
    i = 5
    for src in sorted(per_file.keys()):
        d = per_file[src]
        ws.cell(row=i, column=1, value=src)
        ws.cell(row=i, column=2, value=d["raw"])
        ws.cell(row=i, column=3, value=d["tx"])
        ws.cell(row=i, column=4, value=d["ignored"])
        i += 1
    ws.cell(row=i, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=i, column=2,
            value=sum(d["raw"] for d in per_file.values()))
    ws.cell(row=i, column=3,
            value=sum(d["tx"] for d in per_file.values()))
    ws.cell(row=i, column=4,
            value=sum(d["ignored"] for d in per_file.values()))

    # Per-year trace counts
    base = i + 3
    ws.cell(row=base, column=1, value="Per tax year reconciliation")
    ws.cell(row=base, column=1).font = Font(bold=True, size=12)
    headers = ["Tax year", "Match lines",
               "Gross proceeds CZK", "Tax base CZK", "Tax due CZK",
               "Link"]
    write_header(ws, headers, row=base + 1)
    for k, row in enumerate(yearly_summary, start=base + 2):
        ws.cell(row=k, column=1, value=row["Tax year"])
        ws.cell(row=k, column=2, value=row["Match line count"])
        ws.cell(row=k, column=3,
                value=row["Gross proceeds CZK (all sells)"])
        ws.cell(row=k, column=4, value=row["Final tax base CZK"])
        ws.cell(row=k, column=5, value=row["Tax due CZK"])
        ws.cell(row=k, column=6,
                value="See Lot_Matching filtered by Tax year for details")

    autosize_columns(ws)
    ws.freeze_panes = "A2"


# -----------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------

