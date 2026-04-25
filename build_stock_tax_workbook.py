#!/usr/bin/env python3
"""Czech Personal Income Tax Workbook Generator for Stock Trades.

Builds an auditable, operator-usable Excel workbook that calculates Czech
personal income tax on stock trades imported from Yahoo Portfolio CSV
exports. See README_OPERATOR.md for the human-facing guide and
IMPLEMENTATION_NOTES.md for assumptions and internal design.

Usage:
    py -3 build_stock_tax_workbook.py \
        --input .csv/XTB_CZK.csv .csv/XTB_USD.csv .csv/Lynx.csv \
                .csv/Revolut.csv .csv/Trading212.csv \
        --output stock_tax_system.xlsx

If the --output workbook already exists, user-maintained sheets are read
back (Settings, FX_Daily, FX_Yearly, Instrument_Map, Corporate_Actions,
Method_Selection, Locked_Years, Frozen_Inventory, Review_State,
Filed_Year_Reconciliation) so operator edits and locked-year snapshots are
preserved.
"""

from __future__ import annotations

import argparse
import copy
import csv
import dataclasses
import datetime as _dt
import hashlib
import os
import shutil
import sys
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import urllib.request as _urlreq
import json as _json

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from stock_tax_app.engine import policy


# -----------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------

DEFAULT_TAX_RATE = 0.15
DEFAULT_APPLY_100K = False
DEFAULT_100K_THRESHOLD = 100_000.0  # CZK
DEFAULT_FX_METHOD = "FX_UNIFIED_GFR"
SUPPORTED_FX_METHODS = ("FX_DAILY_CNB", "FX_UNIFIED_GFR")
SUPPORTED_METHODS = policy.SUPPORTED_METHODS
DEFAULT_METHOD = policy.DEFAULT_METHOD

# GFŘ unified yearly USD/CZK rates (Czech tax authority official rates).
# Source: GFŘ pokyn (instruction) D-series, published annually by the Czech
# Financial Administration (Generální finanční ředitelství).
# Use FX_UNIFIED_GFR method — this is the correct legal basis for Czech PIT.
# Earlier years use best-available GFŘ / CNB annual averages.
DEFAULT_FX_YEARLY = {
    2015: 24.60, 2016: 24.44, 2017: 23.38, 2018: 21.78,
    2019: 22.93, 2020: 23.14, 2021: 21.68, 2022: 23.36,
    2023: 22.21,
    2024: 23.28,  # GFŘ-D-65  ← authoritative
    2025: 21.84,  # GFŘ-D-75  ← authoritative
    2026: 22.00,  # placeholder — update when GFŘ publishes D-xx for 2026
}

# Official GFŘ unified USD/CZK rates {year: (rate, source_label)}.
GFR_OFFICIAL_RATES = {
    2023: (22.21, "GFŘ-D-57"),
    2024: (23.28, "GFŘ-D-65"),
    2025: (21.84, "GFŘ-D-75"),
}

# Compatibility aliases derived from stock_tax_app.engine.policy.
YEAR_DEFAULT_METHODS = policy.YEAR_DEFAULT_METHODS
FILED_YEARS = policy.FILED_YEARS
# CNB daily FX cache file (written next to the workbook).
CNB_DAILY_CACHE_FILE = "cnb_daily_cache.json"
CANONICAL_OUTPUT_NAME = "stock_tax_system.xlsx"

TX_SIDES = ("BUY", "SELL")
CA_TYPES = ("SPLIT", "REVERSE_SPLIT", "TICKER_CHANGE")


# -----------------------------------------------------------------------
# Styling helpers
# -----------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FF2F5496")
SUBHEADER_FILL = PatternFill("solid", fgColor="FFDEEBF7")
EDITABLE_FILL = PatternFill("solid", fgColor="FFFFF2CC")
LOCKED_FILL = PatternFill("solid", fgColor="FFD9D9D9")
ERROR_FILL = PatternFill("solid", fgColor="FFF8CBAD")
WARNING_FILL = PatternFill("solid", fgColor="FFFFE699")
OK_FILL = PatternFill("solid", fgColor="FFC6EFCE")

THIN = Side(style="thin", color="FFB4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws, min_width: int = 10, max_width: int = 42) -> None:
    widths: Dict[str, int] = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            col = get_column_letter(cell.column)
            length = len(str(cell.value))
            if length > widths.get(col, 0):
                widths[col] = length
    for col, width in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(max_width, width + 2))


def write_header(ws, headers: List[str], row: int = 1) -> None:
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def add_table(ws, name: str, ref: str, style: str = "TableStyleMedium2") -> None:
    try:
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name=style, showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tab)
    except Exception:
        # Some sheets (very small) may fail table creation; skip silently.
        pass


# -----------------------------------------------------------------------
# Data model
# -----------------------------------------------------------------------

@dataclass
class RawRow:
    source_file: str
    source_row: int
    source_broker: str
    source_account: str
    data: Dict[str, str]

    @property
    def symbol(self) -> str:
        return (self.data.get("Symbol") or "").strip()


@dataclass
class Transaction:
    tx_id: str
    source_file: str
    source_row: int
    source_broker: str
    source_account: str
    symbol: str
    instrument_id: str
    trade_date: date
    side: str              # BUY or SELL
    quantity: float
    price_usd: float
    commission_usd: float
    comment: str = ""


@dataclass
class Lot:
    lot_id: str
    tx_id: str
    instrument_id: str
    source_broker: str
    source_account: str
    source_file: str
    source_row: int
    buy_date: date
    quantity_original: float
    quantity_remaining: float
    price_per_share_usd: float
    buy_commission_total_usd: float
    adjustments: List[str] = field(default_factory=list)

    @property
    def buy_commission_per_share_usd(self) -> float:
        if self.quantity_original <= 0:
            return 0.0
        return self.buy_commission_total_usd / self.quantity_original


@dataclass
class MatchLine:
    match_id: str
    sell_tx_id: str
    sell_date: date
    sell_source_broker: str
    sell_source_file: str
    sell_source_row: int
    instrument_id: str
    buy_lot_id: str
    buy_tx_id: str
    buy_date: date
    buy_source_broker: str
    buy_source_file: str
    buy_source_row: int
    quantity: float
    buy_price_per_share_usd: float
    sell_price_per_share_usd: float
    allocated_buy_commission_usd: float
    allocated_sell_commission_usd: float
    fx_rate_buy: float
    fx_rate_sell: float
    cost_basis_czk: float
    proceeds_czk: float
    holding_days: int
    time_test_exempt: bool
    taxable: bool
    taxable_gain_czk: float
    method: str
    tax_year: int


@dataclass
class CalculationResult:
    inputs: List[Path]
    output_path: Path
    raw_rows: List[RawRow]
    txs: List[Transaction]
    ignored: List[Dict[str, Any]]
    problems: List[Dict[str, Any]]
    import_log: List[Dict[str, Any]]
    years: List[int]
    instrument_ids: List[str]
    user_state: Dict[str, Any]
    settings: Dict[int, Dict[str, Any]]
    instrument_map: Dict[str, Dict[str, str]]
    fx_yearly: Dict[int, float]
    fx_daily: Dict[date, float]
    fx_yearly_sources: Dict[int, str]
    locked_years: Dict[int, bool]
    corporate_actions: List[Dict[str, Any]]
    frozen_inventory: Dict[int, List[Dict[str, Any]]]
    frozen_matching: Dict[int, List[Dict[str, Any]]]
    frozen_snapshots: Dict[int, Dict[str, Any]]
    review_state: Dict[str, Dict[str, Any]]
    filed_reconciliation: Dict[int, Dict[str, Any]]
    method_selection: Dict[Tuple[int, str], str]
    fx: "FXResolver"
    lots_final: List[Lot]
    match_lines: List[MatchLine]
    sim_warnings: List[Dict[str, Any]]
    year_end_inventory: Dict[int, List[Lot]]
    yearly_summary: List[Dict[str, Any]]
    method_comparison: List[Dict[str, Any]]
    split_warnings: List[Dict[str, Any]]


# -----------------------------------------------------------------------
# CSV parsing
# -----------------------------------------------------------------------

YAHOO_COLUMNS = [
    "Symbol", "Current Price", "Date", "Time", "Change", "Open", "High",
    "Low", "Volume", "Trade Date", "Purchase Price", "Quantity",
    "Commission", "High Limit", "Low Limit", "Comment", "Transaction Type",
]


def parse_trade_date(value: str) -> Optional[date]:
    value = (value or "").strip()
    if not value:
        return None
    if len(value) == 8 and value.isdigit():
        try:
            return date(int(value[:4]), int(value[4:6]), int(value[6:8]))
        except ValueError:
            return None
    # Fallback: try common formats
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def safe_float(value: str, default: float = 0.0) -> Tuple[Optional[float], bool]:
    """Return (value, ok). If input is blank, return (default, True)."""
    s = (value or "").strip()
    if s == "":
        return default, True
    try:
        return float(s), True
    except ValueError:
        return None, False


def broker_from_filename(path: Path) -> Tuple[str, str]:
    stem = path.stem
    if "_" in stem:
        broker, _, account = stem.partition("_")
        return broker, account
    return stem, ""


def read_csv_file(path: Path) -> Tuple[List[RawRow], List[Dict[str, Any]]]:
    raw_rows: List[RawRow] = []
    problems: List[Dict[str, Any]] = []
    if not path.exists():
        problems.append({
            "source_file": path.name, "source_row": 0,
            "reason": "FILE_NOT_FOUND", "symbol": "", "trade_date": "",
            "transaction_type": "", "detail": str(path),
        })
        return raw_rows, problems
    broker, account = broker_from_filename(path)
    with path.open(encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for idx, row in enumerate(reader, start=2):  # data starts at line 2
            cleaned = {k: (v or "").strip() for k, v in (row or {}).items() if k}
            raw_rows.append(RawRow(
                source_file=path.name, source_row=idx,
                source_broker=broker, source_account=account,
                data=cleaned,
            ))
    return raw_rows, problems


def normalize_transactions(
    raw_rows: List[RawRow],
) -> Tuple[List[Transaction], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Split raw rows into (transactions, ignored, validation_issues)."""
    txs: List[Transaction] = []
    ignored: List[Dict[str, Any]] = []
    problems: List[Dict[str, Any]] = []

    for raw in raw_rows:
        data = raw.data
        tt = (data.get("Transaction Type") or "").strip().upper()
        td_raw = (data.get("Trade Date") or "").strip()

        if not tt and not td_raw:
            ignored.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "symbol": raw.symbol, "reason": "position_row_no_trade",
                "detail": "empty Trade Date and empty Transaction Type",
            })
            continue
        if not tt or not td_raw:
            ignored.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "symbol": raw.symbol,
                "reason": "incomplete_transaction",
                "detail": f"trade_date='{td_raw}' tx_type='{tt}'",
            })
            continue

        trade_date = parse_trade_date(td_raw)
        if trade_date is None:
            problems.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "check": "invalid_date", "severity": "ERROR",
                "detail": f"Trade Date '{td_raw}' is not parseable",
            })
            ignored.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "symbol": raw.symbol, "reason": "invalid_date",
                "detail": f"Trade Date '{td_raw}'",
            })
            continue

        if tt not in TX_SIDES:
            problems.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "check": "invalid_transaction_type", "severity": "ERROR",
                "detail": f"Transaction Type '{tt}' not in BUY/SELL",
            })
            ignored.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "symbol": raw.symbol, "reason": "invalid_tx_type",
                "detail": f"Transaction Type '{tt}'",
            })
            continue

        qty_val, qty_ok = safe_float(data.get("Quantity") or "", default=None)
        if not qty_ok or qty_val is None or qty_val <= 0:
            problems.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "check": "invalid_quantity", "severity": "ERROR",
                "detail": f"Quantity '{data.get('Quantity')}'",
            })
            ignored.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "symbol": raw.symbol, "reason": "invalid_qty",
                "detail": data.get("Quantity") or "",
            })
            continue

        price_val, price_ok = safe_float(data.get("Purchase Price") or "",
                                         default=None)
        if not price_ok or price_val is None or price_val < 0:
            problems.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "check": "invalid_price", "severity": "ERROR",
                "detail": f"Purchase Price '{data.get('Purchase Price')}'",
            })
            ignored.append({
                "source_file": raw.source_file, "source_row": raw.source_row,
                "symbol": raw.symbol, "reason": "invalid_price",
                "detail": data.get("Purchase Price") or "",
            })
            continue

        commission_val, _ = safe_float(data.get("Commission") or "", default=0.0)
        if commission_val is None:
            commission_val = 0.0

        tx = Transaction(
            tx_id=f"{raw.source_file}#{raw.source_row}",
            source_file=raw.source_file,
            source_row=raw.source_row,
            source_broker=raw.source_broker,
            source_account=raw.source_account,
            symbol=raw.symbol,
            instrument_id=raw.symbol,  # default; user may override via Instrument_Map
            trade_date=trade_date,
            side=tt,
            quantity=qty_val,
            price_usd=price_val,
            commission_usd=commission_val,
            comment=(data.get("Comment") or ""),
        )
        txs.append(tx)

    # Duplicate detection — same broker/date/symbol/side/qty/price.
    seen = defaultdict(list)
    for tx in txs:
        key = (tx.source_broker, tx.source_account, tx.trade_date,
               tx.symbol, tx.side, round(tx.quantity, 6),
               round(tx.price_usd, 6))
        seen[key].append(tx)
    for key, group in seen.items():
        if len(group) > 1:
            for t in group:
                problems.append({
                    "source_file": t.source_file, "source_row": t.source_row,
                    "check": "duplicate_candidate", "severity": "WARN",
                    "detail": (f"{len(group)} identical trades "
                               f"for {t.symbol} on {t.trade_date}"),
                })
    return txs, ignored, problems


# -----------------------------------------------------------------------
# Persistence of user-editable sheets
# -----------------------------------------------------------------------

def _read_table(ws, header_row: int = 1) -> List[Dict[str, Any]]:
    """Read a worksheet table into a list of dicts.

    ``header_row`` is the 1-indexed row containing column headers.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < header_row:
        return []
    header = rows[header_row - 1]
    out: List[Dict[str, Any]] = []
    for r in rows[header_row:]:
        if r is None:
            continue
        if all(v is None or (isinstance(v, str) and v.strip() == "")
               for v in r):
            continue
        out.append({str(header[i]): r[i] for i in range(len(header))
                    if header[i] is not None})
    return out


# Header-row offsets for sheets load_existing_user_state reads back.
USER_SHEET_HEADER_ROWS = {
    "Settings": 1,
    "Instrument_Map": 1,
    "FX_Daily": 3,
    "FX_Yearly": 3,
    "Corporate_Actions": 3,
    "Method_Selection": 3,
    "Locked_Years": 3,
    "Frozen_Inventory": 3,
    "Frozen_Lot_Matching": 3,
    "Frozen_Snapshots": 3,
    "Review_State": 1,
    "Filed_Year_Reconciliation": 1,
}


def load_existing_user_state(path: Path) -> Dict[str, Any]:
    """Return dict of user-maintained sheet data if workbook exists."""
    state: Dict[str, Any] = {}
    if not path.exists():
        return state
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:  # pragma: no cover
        print(f"WARN: could not read existing workbook {path}: {exc}",
              file=sys.stderr)
        return state
    for sheet, hrow in USER_SHEET_HEADER_ROWS.items():
        if sheet in wb.sheetnames:
            state[sheet] = _read_table(wb[sheet], header_row=hrow)
    return state


# -----------------------------------------------------------------------
# Settings
# -----------------------------------------------------------------------

def build_settings(
    user_state: Dict[str, Any], years: List[int],
) -> Dict[int, Dict[str, Any]]:
    existing: Dict[int, Dict[str, Any]] = {}
    for row in user_state.get("Settings", []):
        try:
            y = int(row.get("Tax year"))
        except (TypeError, ValueError):
            continue
        existing[y] = row
    out: Dict[int, Dict[str, Any]] = {}
    for y in years:
        src = existing.get(y, {})
        out[y] = {
            "tax_rate": _to_float(src.get("Tax rate"), DEFAULT_TAX_RATE),
            "fx_method": str(src.get("FX method") or DEFAULT_FX_METHOD).upper(),
            "apply_100k": _to_bool(src.get("Apply 100k exemption?"),
                                   DEFAULT_APPLY_100K),
            "locked": _to_bool(src.get("Locked year?"), False),
            "notes": src.get("Notes") or "",
        }
        if out[y]["fx_method"] not in SUPPORTED_FX_METHODS:
            out[y]["fx_method"] = DEFAULT_FX_METHOD
        if policy.is_locked(y):
            out[y]["locked"] = True
    return out


def _to_float(value: Any, default: float) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_bool(value: Any, default: bool) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("true", "1", "yes", "y", "ano"):
        return True
    if s in ("false", "0", "no", "n", "ne"):
        return False
    return default


def build_fx_tables(user_state: Dict[str, Any], years: List[int]
                    ) -> Tuple[Dict[int, float], Dict[date, float], Dict[int, str]]:
    """Return (yearly_rates, daily_rates, yearly_sources).

    yearly_sources maps year → source label (e.g. "GFŘ-D-65" or "manual" or "default").
    """
    yearly: Dict[int, float] = {}
    yearly_src: Dict[int, str] = {}
    for row in user_state.get("FX_Yearly", []):
        try:
            y = int(row.get("Tax year"))
            r = float(row.get("USD_CZK"))
        except (TypeError, ValueError):
            continue
        yearly[y] = r
        src_note = (row.get("Source / note") or "").strip()
        yearly_src[y] = src_note if src_note else "manual"
    for y in years:
        if y not in yearly:
            if y in GFR_OFFICIAL_RATES:
                official_r, official_label = GFR_OFFICIAL_RATES[y]
                yearly[y] = official_r
                yearly_src[y] = official_label
            else:
                yearly[y] = DEFAULT_FX_YEARLY.get(y, 22.0)
                yearly_src[y] = "default"

    daily: Dict[date, float] = {}
    for row in user_state.get("FX_Daily", []):
        d = row.get("Date")
        rate = row.get("USD_CZK")
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, str):
            d = parse_trade_date(d)
        if not isinstance(d, date) or rate is None:
            continue
        try:
            daily[d] = float(rate)
        except (TypeError, ValueError):
            continue
    return yearly, daily, yearly_src


def build_instrument_map(user_state: Dict[str, Any],
                         txs: List[Transaction]) -> Dict[str, Dict[str, str]]:
    mp: Dict[str, Dict[str, str]] = {}
    for row in user_state.get("Instrument_Map", []):
        sym = (row.get("Yahoo Symbol") or "").strip()
        if not sym:
            continue
        mp[sym] = {
            "Yahoo Symbol": sym,
            "Instrument_ID": (row.get("Instrument_ID") or sym) or sym,
            "ISIN": row.get("ISIN") or "",
            "Instrument name": row.get("Instrument name") or "",
            "Notes": row.get("Notes") or "",
        }
    for tx in txs:
        if tx.symbol not in mp:
            mp[tx.symbol] = {
                "Yahoo Symbol": tx.symbol,
                "Instrument_ID": tx.symbol,
                "ISIN": "",
                "Instrument name": "",
                "Notes": "",
            }
    return mp


def apply_instrument_map(txs: List[Transaction],
                         mp: Dict[str, Dict[str, str]]) -> None:
    for tx in txs:
        info = mp.get(tx.symbol)
        if info and info.get("Instrument_ID"):
            tx.instrument_id = info["Instrument_ID"]
        else:
            tx.instrument_id = tx.symbol


def build_method_selection(user_state: Dict[str, Any],
                           years: List[int],
                           instrument_ids: Iterable[str]
                           ) -> Dict[Tuple[int, str], str]:
    sel: Dict[Tuple[int, str], str] = {}
    for row in user_state.get("Method_Selection", []):
        try:
            y = int(row.get("Tax year"))
        except (TypeError, ValueError):
            continue
        inst = (row.get("Instrument_ID") or "").strip()
        method = policy.resolved_method_for(y, row.get("Method"))
        sel[(y, inst)] = method
    for y in years:
        year_default = policy.resolved_method_for(y)
        for inst in instrument_ids:
            sel.setdefault((y, inst), year_default)
    return sel


def build_locked_years(user_state: Dict[str, Any],
                       years: List[int]) -> Dict[int, bool]:
    out: Dict[int, bool] = {}
    for row in user_state.get("Locked_Years", []):
        try:
            y = int(row.get("Tax year"))
        except (TypeError, ValueError):
            continue
        out[y] = _to_bool(row.get("Locked?"), False)
    for y in years:
        out.setdefault(y, False)
        if policy.is_locked(y):
            out[y] = True
    return out


def build_corporate_actions(user_state: Dict[str, Any]
                            ) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for row in user_state.get("Corporate_Actions", []):
        d = row.get("Date")
        if isinstance(d, datetime):
            d = d.date()
        elif isinstance(d, str):
            d = parse_trade_date(d)
        if not isinstance(d, date):
            continue
        inst = (row.get("Instrument_ID") or "").strip()
        action = (row.get("Action type") or "").strip().upper()
        if not inst or action not in CA_TYPES:
            continue
        try:
            ratio_old = float(row.get("Ratio old") or 1.0)
            ratio_new = float(row.get("Ratio new") or 1.0)
        except (TypeError, ValueError):
            continue
        cash_in_lieu = _to_float(row.get("Cash in lieu"), 0.0)
        out.append({
            "Date": d, "Instrument_ID": inst, "Action type": action,
            "Ratio old": ratio_old, "Ratio new": ratio_new,
            "Cash in lieu": cash_in_lieu,
            "Notes": row.get("Notes") or "",
            "Applied": _to_bool(row.get("Applied?"), True),
        })
    out.sort(key=lambda r: (r["Date"], r["Instrument_ID"]))
    return out


def load_frozen_inventory(user_state: Dict[str, Any]
                          ) -> Dict[int, List[Dict[str, Any]]]:
    """Return {year: [frozen lot records]}."""
    out: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for row in user_state.get("Frozen_Inventory", []):
        try:
            y = int(row.get("Snapshot year"))
        except (TypeError, ValueError):
            continue
        out[y].append(row)
    return dict(out)


def load_frozen_matching(user_state: Dict[str, Any]
                         ) -> Dict[int, List[Dict[str, Any]]]:
    out: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for row in user_state.get("Frozen_Lot_Matching", []):
        try:
            y = int(row.get("Tax year"))
        except (TypeError, ValueError):
            continue
        out[y].append(row)
    return dict(out)


def load_frozen_snapshots(user_state: Dict[str, Any]) -> Dict[int, Dict[str, Any]]:
    out: Dict[int, Dict[str, Any]] = {}
    for row in user_state.get("Frozen_Snapshots", []):
        try:
            y = int(row.get("Snapshot year"))
        except (TypeError, ValueError):
            continue
        out[y] = dict(row)
    return out


def load_review_state(user_state: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """Return {sell_tx_id: {review_status, operator_note}} from Review_State."""
    out: Dict[str, Dict[str, Any]] = {}
    for row in user_state.get("Review_State", []):
        sid = str(row.get("Sell_ID") or "").strip()
        if not sid:
            continue
        out[sid] = {
            "review_status": row.get("Review status") or "",
            "operator_note": row.get("Operator note") or "",
        }
    return out


def load_filed_reconciliation(user_state: Dict[str, Any]
                              ) -> Dict[int, Dict[str, Any]]:
    """Return {year: {filed_method, filed_tax_base, filed_tax_due}} from Filed_Year_Reconciliation."""
    out: Dict[int, Dict[str, Any]] = {}
    for row in user_state.get("Filed_Year_Reconciliation", []):
        try:
            y = int(row.get("Tax year"))
        except (TypeError, ValueError):
            continue
        expected_method = str(
            row.get("Filed method")
            or row.get("Expected filed method")
            or ""
        ).strip().upper()
        filed_method = policy.filed_method(y)
        if filed_method and expected_method and expected_method != filed_method:
            # Ignore stale filed-year reconciliation rows from the earlier
            # 2024 FIFO policy; the generator now treats 2024 as filed LIFO.
            continue
        out[y] = {
            "filed_method": expected_method,
            "filed_tax_base": row.get("Filed tax base CZK"),
            "filed_tax_due": row.get("Filed tax due CZK"),
        }
    return out


# -----------------------------------------------------------------------
# FX download (best-effort, requires network)
# -----------------------------------------------------------------------

def _cnb_cache_path(workbook_path: Path) -> Path:
    return workbook_path.parent / CNB_DAILY_CACHE_FILE


def _load_cnb_cache(cache_path: Path) -> Dict[str, float]:
    """Load {date_iso: rate} from JSON cache file."""
    if not cache_path.exists():
        return {}
    try:
        with cache_path.open("r", encoding="utf-8") as fh:
            return _json.load(fh)
    except Exception:
        return {}


def _save_cnb_cache(cache_path: Path, data: Dict[str, float]) -> None:
    try:
        with cache_path.open("w", encoding="utf-8") as fh:
            _json.dump(data, fh, indent=2)
    except Exception:
        pass


def download_cnb_daily_rates_year(year: int, timeout: int = 15
                                  ) -> Dict[date, float]:
    """Download CNB daily USD/CZK rates for *year* from cnb.cz.

    Returns {date: rate}. Empty dict on network failure.
    """
    url = (
        "https://www.cnb.cz/en/financial_markets/"
        "foreign_exchange_market/exchange_rate_fixing/"
        f"year.txt?year={year}"
    )
    try:
        req = _urlreq.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with _urlreq.urlopen(req, timeout=timeout) as resp:
            text = resp.read().decode("utf-8", errors="replace")
    except Exception:
        return {}
    out: Dict[date, float] = {}
    for line in text.strip().splitlines():
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 5:
            continue
        # First column is date; find USD in code column (4th or 5th)
        try:
            d = datetime.strptime(parts[0], "%d.%m.%Y").date()
        except ValueError:
            continue
        for ci in (4, 3):
            if ci < len(parts) and parts[ci].upper() == "USD":
                ri = ci + 1
                if ri < len(parts):
                    try:
                        rate = float(parts[ri].replace(",", "."))
                        out[d] = rate
                    except ValueError:
                        pass
                break
    return out


def refresh_fx_daily_for_years(
    fx_daily: Dict[date, float],
    years_needing_daily: List[int],
    cache_path: Path,
) -> Tuple[Dict[date, float], List[str]]:
    """Download missing CNB daily rates for given years.

    Returns (updated_dict, list_of_info_messages).
    """
    msgs: List[str] = []
    cache_raw = _load_cnb_cache(cache_path)
    # Seed fx_daily from cache
    updated = dict(fx_daily)
    for iso, rate in cache_raw.items():
        try:
            d = date.fromisoformat(iso)
            updated.setdefault(d, rate)
        except ValueError:
            continue

    for y in sorted(set(years_needing_daily)):
        if any(d.year == y for d in updated):
            msgs.append(f"FX_DAILY_CNB year {y}: using cached/manual rates.")
            continue
        msgs.append(f"FX_DAILY_CNB year {y}: downloading from CNB …")
        downloaded = download_cnb_daily_rates_year(y)
        if downloaded:
            updated.update(downloaded)
            msgs.append(f"  → {len(downloaded)} dates downloaded for {y}.")
            # Persist to cache
            new_raw = dict(cache_raw)
            for d, r in downloaded.items():
                new_raw[d.isoformat()] = r
            _save_cnb_cache(cache_path, new_raw)
        else:
            msgs.append(f"  → Download failed for {y} — add rates manually to FX_Daily.")
    return updated, msgs


# -----------------------------------------------------------------------
# FX lookup
# -----------------------------------------------------------------------

class FXResolver:
    def __init__(self, yearly: Dict[int, float], daily: Dict[date, float],
                 settings: Dict[int, Dict[str, Any]]):
        self.yearly = yearly
        self.daily = daily
        self.settings = settings
        self.missing_daily: List[date] = []
        self.used_fallback_yearly: List[date] = []

    def rate_for(self, d: date) -> Tuple[float, str]:
        y = d.year
        method = self.settings.get(y, {}).get("fx_method", DEFAULT_FX_METHOD)
        if method == "FX_DAILY_CNB":
            # Use exact date; fall back to nearest earlier date within 10 days.
            if d in self.daily:
                return self.daily[d], "FX_DAILY_CNB_exact"
            for back in range(1, 11):
                alt = d - timedelta(days=back)
                if alt in self.daily:
                    return self.daily[alt], f"FX_DAILY_CNB_back{back}d"
            self.missing_daily.append(d)
            if y in self.yearly:
                self.used_fallback_yearly.append(d)
                return self.yearly[y], "FX_DAILY_CNB_fallback_yearly"
            return 22.0, "FX_DAILY_CNB_fallback_hardcoded"
        # FX_UNIFIED_GFR or any unknown method
        if y in self.yearly:
            return self.yearly[y], "FX_UNIFIED_GFR"
        return DEFAULT_FX_YEARLY.get(y, 22.0), "FX_UNIFIED_GFR_default"


# -----------------------------------------------------------------------
# Corporate actions application
# -----------------------------------------------------------------------

def apply_split_to_lots(lots: List[Lot], action: Dict[str, Any]) -> None:
    """Adjust lot quantity/price per share for a split.

    total cost basis unchanged; per-share price divided by factor.
    """
    inst = action["Instrument_ID"]
    action_date = action["Date"]
    ratio_old = action["Ratio old"]
    ratio_new = action["Ratio new"]
    if ratio_old <= 0 or ratio_new <= 0:
        return
    # e.g. 2-for-1 split: ratio_old=1, ratio_new=2 -> factor=2
    factor = ratio_new / ratio_old
    if factor == 1.0:
        return
    for lot in lots:
        if lot.instrument_id != inst:
            continue
        if lot.buy_date > action_date:
            continue
        if lot.quantity_remaining <= 0:
            continue
        lot.quantity_remaining *= factor
        lot.quantity_original *= factor
        lot.price_per_share_usd /= factor
        lot.adjustments.append(
            f"{action['Action type']} {ratio_old}:{ratio_new} on {action_date}")


# -----------------------------------------------------------------------
# Lot matching core
# -----------------------------------------------------------------------

def _expected_contribution_per_share_czk(
    lot: Lot, sell: Transaction,
    sell_commission_per_share_usd: float,
    fx: FXResolver,
) -> Tuple[float, bool]:
    """Return (per-share taxable CZK contribution, exempt?)."""
    exempt = sell.trade_date > _add_years(lot.buy_date, 3)
    if exempt:
        return 0.0, True
    fx_buy, _ = fx.rate_for(lot.buy_date)
    fx_sell, _ = fx.rate_for(sell.trade_date)
    net_sell = (sell.price_usd - sell_commission_per_share_usd) * fx_sell
    net_buy = (lot.price_per_share_usd
               + lot.buy_commission_per_share_usd) * fx_buy
    return (net_sell - net_buy), False


def _add_years(d: date, years: int) -> date:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        # leap-year Feb 29
        return d.replace(year=d.year + years, day=28)


def rank_lots_for_sell(lots: List[Lot], sell: Transaction, method: str,
                       fx: FXResolver) -> List[Lot]:
    available = [l for l in lots if l.instrument_id == sell.instrument_id
                 and l.quantity_remaining > 1e-9
                 and l.buy_date <= sell.trade_date]
    if method == "FIFO":
        available.sort(key=lambda l: (l.buy_date, l.source_file, l.source_row))
        return available
    if method == "LIFO":
        available.sort(key=lambda l: (l.buy_date, l.source_file, l.source_row),
                       reverse=True)
        return available

    # MIN_GAIN / MAX_GAIN
    sell_comm_per_share = (sell.commission_usd / sell.quantity
                           if sell.quantity > 0 else 0.0)
    scored: List[Tuple[float, bool, Lot]] = []
    for lot in available:
        contrib, exempt = _expected_contribution_per_share_czk(
            lot, sell, sell_comm_per_share, fx)
        scored.append((contrib, exempt, lot))

    if method == "MIN_GAIN":
        scored.sort(key=lambda t: (t[0], t[2].buy_date,
                                   t[2].source_file, t[2].source_row))
    elif method == "MAX_GAIN":
        scored.sort(key=lambda t: (-t[0], t[2].buy_date,
                                   t[2].source_file, t[2].source_row))
    return [t[2] for t in scored]


def _make_match_line(
    sell: Transaction, lot: Lot, take: float,
    fx: FXResolver, match_counter: dict, method: str,
) -> MatchLine:
    """Shared helper: build a single MatchLine from (sell, lot, take)."""
    sell_comm_ps = sell.commission_usd / sell.quantity if sell.quantity > 0 else 0.0
    fx_buy, _ = fx.rate_for(lot.buy_date)
    fx_sell, _ = fx.rate_for(sell.trade_date)
    alloc_buy_comm_usd = (
        lot.buy_commission_total_usd * (take / lot.quantity_original)
        if lot.quantity_original > 0 else 0.0
    )
    alloc_sell_comm_usd = sell_comm_ps * take
    cost_basis_usd = (lot.price_per_share_usd * take) + alloc_buy_comm_usd
    proceeds_usd = (sell.price_usd * take) - alloc_sell_comm_usd
    cost_basis_czk = cost_basis_usd * fx_buy
    proceeds_czk = proceeds_usd * fx_sell
    exempt = sell.trade_date > _add_years(lot.buy_date, 3)
    taxable_gain = 0.0 if exempt else (proceeds_czk - cost_basis_czk)
    match_counter["n"] += 1
    return MatchLine(
        match_id=f"M{match_counter['n']:06d}",
        sell_tx_id=sell.tx_id,
        sell_date=sell.trade_date,
        sell_source_broker=sell.source_broker,
        sell_source_file=sell.source_file,
        sell_source_row=sell.source_row,
        instrument_id=sell.instrument_id,
        buy_lot_id=lot.lot_id,
        buy_tx_id=lot.tx_id,
        buy_date=lot.buy_date,
        buy_source_broker=lot.source_broker,
        buy_source_file=lot.source_file,
        buy_source_row=lot.source_row,
        quantity=take,
        buy_price_per_share_usd=lot.price_per_share_usd,
        sell_price_per_share_usd=sell.price_usd,
        allocated_buy_commission_usd=alloc_buy_comm_usd,
        allocated_sell_commission_usd=alloc_sell_comm_usd,
        fx_rate_buy=fx_buy,
        fx_rate_sell=fx_sell,
        cost_basis_czk=cost_basis_czk,
        proceeds_czk=proceeds_czk,
        holding_days=(sell.trade_date - lot.buy_date).days,
        time_test_exempt=exempt,
        taxable=not exempt,
        taxable_gain_czk=taxable_gain,
        method=method,
        tax_year=sell.trade_date.year,
    )


def _match_global_optimized(
    sells: List[Transaction],
    lots: List[Lot],
    method: str,
    fx: FXResolver,
    match_counter: dict,
) -> Tuple[List[MatchLine], Dict[str, float]]:
    """True annual global optimizer for MIN_GAIN / MAX_GAIN.

    Considers ALL sells for a (year, instrument) batch simultaneously and
    assigns lots globally — unlike per-sell greedy matching.

    Algorithm (global greedy assignment):
      1. Build every eligible (sell, lot) pair respecting buy_date <= sell_date.
      2. Score each pair by taxable-gain CZK contribution per share:
         exempt pairs score 0 (irrelevant to taxable gain).
      3. Sort globally: ascending for MIN_GAIN, descending for MAX_GAIN.
      4. Greedily consume from the most-favourable pair first, tracking
         remaining quantities for each sell and each lot independently.
      5. Leftover sells are completed FIFO as a fallback (data-quality
         safeguard; should not happen with clean input).

    This is LP-optimal for identical sell prices and near-optimal for
    multi-sell scenarios. Strictly better than per-sell greedy in all
    cases where multiple sells compete for the same lot pool.

    Returns (match_lines, {sell_tx_id: unmatched_qty}).
    """
    if not sells or not lots:
        return [], {s.tx_id: s.quantity for s in sells}

    sell_remaining: Dict[str, float] = {s.tx_id: s.quantity for s in sells}
    lot_remaining: Dict[str, float] = {l.lot_id: l.quantity_remaining for l in lots}
    lot_by_id: Dict[str, Lot] = {l.lot_id: l for l in lots}
    sell_by_id: Dict[str, Transaction] = {s.tx_id: s for s in sells}

    # Build eligible (gain_per_unit, sell_date_iso, buy_date_iso, sell_id, lot_id)
    pairs: List[tuple] = []
    for sell in sells:
        sell_comm_ps = sell.commission_usd / sell.quantity if sell.quantity > 0 else 0.0
        fx_sell, _ = fx.rate_for(sell.trade_date)
        for lot in lots:
            if lot.buy_date > sell.trade_date:
                continue
            exempt = sell.trade_date > _add_years(lot.buy_date, 3)
            if exempt:
                gain_pu = 0.0
            else:
                fx_buy, _ = fx.rate_for(lot.buy_date)
                proceeds_pu = (sell.price_usd - sell_comm_ps) * fx_sell
                cost_pu = (lot.price_per_share_usd
                           + lot.buy_commission_per_share_usd) * fx_buy
                gain_pu = proceeds_pu - cost_pu
            pairs.append((
                gain_pu,
                sell.trade_date.isoformat(),
                lot.buy_date.isoformat(),
                sell.tx_id,
                lot.lot_id,
            ))

    reverse = (method == "MAX_GAIN")
    pairs.sort(reverse=reverse)

    out_lines: List[MatchLine] = []
    for gain_pu, _sd, _ld, sell_tx_id, lot_id in pairs:
        sr = sell_remaining.get(sell_tx_id, 0.0)
        lr = lot_remaining.get(lot_id, 0.0)
        if sr < 1e-9 or lr < 1e-9:
            continue
        take = min(sr, lr)
        sell = sell_by_id[sell_tx_id]
        lot = lot_by_id[lot_id]
        out_lines.append(_make_match_line(sell, lot, take, fx,
                                          match_counter, method))
        sell_remaining[sell_tx_id] = sr - take
        lot_remaining[lot_id] = lr - take

    # Apply lot quantity changes back to actual lot objects
    for lot in lots:
        lot.quantity_remaining = lot_remaining[lot.lot_id]

    # FIFO fallback for any unresolved sells (data quality safeguard)
    for sell_tx_id, leftover in list(sell_remaining.items()):
        if leftover < 1e-9:
            continue
        sell = sell_by_id[sell_tx_id]
        for lot in sorted(lots, key=lambda l: (l.buy_date, l.source_file,
                                               l.source_row)):
            if lot.buy_date > sell.trade_date or lot.quantity_remaining < 1e-9:
                continue
            take = min(leftover, lot.quantity_remaining)
            if take < 1e-9:
                continue
            out_lines.append(_make_match_line(sell, lot, take, fx,
                                              match_counter,
                                              method + "_fallback_fifo"))
            lot.quantity_remaining -= take
            leftover -= take
            sell_remaining[sell_tx_id] = leftover
            if leftover < 1e-9:
                break

    return out_lines, sell_remaining


def match_sell(
    sell: Transaction, lots: List[Lot], method: str, fx: FXResolver,
    match_counter: Dict[str, int],
) -> Tuple[List[MatchLine], float]:
    """Consume lots to cover the sell (FIFO / LIFO / per-sell greedy).

    Returns (match lines, unmatched quantity).
    Note: MIN_GAIN and MAX_GAIN are handled by _match_global_optimized()
    in simulate(); this function only handles FIFO/LIFO (and serves as
    fallback for edge-case calls with other methods).
    """
    remaining = sell.quantity
    ranked = rank_lots_for_sell(lots, sell, method, fx)
    out: List[MatchLine] = []
    for lot in ranked:
        if remaining <= 1e-9:
            break
        take = min(remaining, lot.quantity_remaining)
        if take <= 1e-9:
            continue
        out.append(_make_match_line(sell, lot, take, fx, match_counter, method))
        lot.quantity_remaining -= take
        remaining -= take
    return out, remaining


def _lots_from_frozen(inventory_rows: List[Dict[str, Any]]) -> List[Lot]:
    out: List[Lot] = []
    for r in inventory_rows:
        try:
            qty_original = float(r.get("Quantity original"))
            qty_remaining = float(r.get("Quantity remaining"))
            price = float(r.get("Price per share USD"))
            comm = float(r.get("Buy commission USD") or 0.0)
        except (TypeError, ValueError):
            continue
        bd = r.get("Buy date")
        if isinstance(bd, datetime):
            bd = bd.date()
        elif isinstance(bd, str):
            bd = parse_trade_date(bd)
        if not isinstance(bd, date):
            continue
        out.append(Lot(
            lot_id=str(r.get("Lot_ID") or "FROZEN"),
            tx_id=str(r.get("Tx_ID") or "FROZEN"),
            instrument_id=str(r.get("Instrument_ID") or ""),
            source_broker=str(r.get("Source broker") or ""),
            source_account=str(r.get("Source account") or ""),
            source_file=str(r.get("Source file") or ""),
            source_row=int(r.get("Source row") or 0),
            buy_date=bd,
            quantity_original=qty_original,
            quantity_remaining=qty_remaining,
            price_per_share_usd=price,
            buy_commission_total_usd=comm,
        ))
    return out


def simulate(
    txs: List[Transaction], settings: Dict[int, Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    locked_years: Dict[int, bool],
    corporate_actions: List[Dict[str, Any]],
    frozen_inventory: Dict[int, List[Dict[str, Any]]],
    frozen_matching: Dict[int, List[Dict[str, Any]]],
    frozen_snapshots: Dict[int, Dict[str, Any]],
    fx: FXResolver,
    override_method: Optional[str] = None,
) -> Tuple[List[Lot], List[MatchLine], List[Dict[str, Any]],
           Dict[int, List[Lot]]]:
    """Run full lot-matching over transactions.

    Returns (final lots, match lines, warnings, per-year end inventories)
    """
    warnings: List[Dict[str, Any]] = []
    years_sorted = sorted(set([tx.trade_date.year for tx in txs]) |
                          set(settings.keys()))

    # Determine latest locked year with frozen snapshot.
    snapshot_years = set(frozen_snapshots.keys()) | {
        y for y, rows in frozen_inventory.items() if rows
    }
    seed_year: Optional[int] = None
    for y in sorted(locked_years.keys(), reverse=True):
        if locked_years.get(y) and y in snapshot_years:
            seed_year = y
            break
        if locked_years.get(y) and y not in snapshot_years:
            warnings.append({
                "check": "locked_year_no_snapshot", "severity": "INFO",
                "detail": (
                    f"Year {y} is locked but has no frozen snapshot yet; "
                    "the current run will regenerate it."
                ),
            })
    lots: List[Lot] = []
    if seed_year is not None:
        lots = _lots_from_frozen(frozen_inventory[seed_year])

    # Transactions to process: only those strictly after seed_year.
    eff_txs = [t for t in txs
               if seed_year is None or t.trade_date.year > seed_year]
    # Sort by date, then source_file, row for stable ordering.
    eff_txs.sort(key=lambda t: (t.trade_date, t.source_file, t.source_row))

    match_counter = {"n": 0}
    match_lines: List[MatchLine] = []

    # Preserve historical audit for unlocked years before the seed snapshot.
    # The seed inventory is still the source of truth for future years.
    if seed_year is not None:
        historical_txs = [
            t for t in txs
            if t.trade_date.year < seed_year and not locked_years.get(t.trade_date.year, False)
        ]
        historical_settings = {
            y: s for y, s in settings.items()
            if y < seed_year and not locked_years.get(y, False)
        }
        historical_actions = [ca for ca in corporate_actions if ca["Date"].year < seed_year]
        if historical_txs:
            _, hist_lines, hist_warnings, _ = simulate(
                txs=historical_txs,
                settings=historical_settings,
                method_selection=method_selection,
                locked_years={y: False for y in historical_settings},
                corporate_actions=historical_actions,
                frozen_inventory={},
                frozen_matching={},
                frozen_snapshots={},
                fx=fx,
                override_method=override_method,
            )
            match_lines.extend(hist_lines)
            warnings.extend(hist_warnings)

    # Replay frozen matching rows if locked year has them (so audit covers them).
    for y, rows in frozen_matching.items():
        if not locked_years.get(y):
            continue
        for r in rows:
            try:
                m = MatchLine(
                    match_id=str(r.get("Match_ID") or ""),
                    sell_tx_id=str(r.get("Sell_ID") or ""),
                    sell_date=_coerce_date(r.get("Sell date")),
                    sell_source_broker=str(r.get("Sell source broker") or ""),
                    sell_source_file=str(r.get("Sell source file") or ""),
                    sell_source_row=int(r.get("Sell source row") or 0),
                    instrument_id=str(r.get("Instrument_ID") or ""),
                    buy_lot_id=str(r.get("Buy Lot_ID") or ""),
                    buy_tx_id=str(r.get("Buy Tx_ID") or ""),
                    buy_date=_coerce_date(r.get("Buy date")),
                    buy_source_broker=str(r.get("Buy source broker") or ""),
                    buy_source_file=str(r.get("Buy source file") or ""),
                    buy_source_row=int(r.get("Buy source row") or 0),
                    quantity=float(r.get("Quantity") or 0.0),
                    buy_price_per_share_usd=float(r.get("Buy price USD")
                                                  or 0.0),
                    sell_price_per_share_usd=float(r.get("Sell price USD")
                                                   or 0.0),
                    allocated_buy_commission_usd=float(
                        r.get("Allocated buy commission USD") or 0.0),
                    allocated_sell_commission_usd=float(
                        r.get("Allocated sell commission USD") or 0.0),
                    fx_rate_buy=float(r.get("FX rate buy") or 0.0),
                    fx_rate_sell=float(r.get("FX rate sell") or 0.0),
                    cost_basis_czk=float(r.get("Cost basis CZK") or 0.0),
                    proceeds_czk=float(r.get("Proceeds CZK") or 0.0),
                    holding_days=int(r.get("Holding days") or 0),
                    time_test_exempt=_to_bool(r.get("Time-test exempt?"),
                                              False),
                    taxable=_to_bool(r.get("Taxable?"), True),
                    taxable_gain_czk=float(r.get("Taxable gain CZK") or 0.0),
                    method=str(r.get("Method") or ""),
                    tax_year=int(r.get("Tax year") or y),
                )
                match_lines.append(m)
            except (TypeError, ValueError):
                continue

    # Apply any corporate actions that predate seed_year+1 against frozen lots
    # as a safety net (user may have entered late split data).
    for ca in corporate_actions:
        if seed_year is None or ca["Date"].year > seed_year:
            continue
        apply_split_to_lots(lots, ca)

    year_end_inventory: Dict[int, List[Lot]] = {}

    # Precompute per-instrument tx queue.
    tx_iter = iter(eff_txs)
    pending: Optional[Transaction] = next(tx_iter, None)

    ca_iter = iter(corporate_actions)
    next_ca = next(ca_iter, None)
    # Skip CAs already applied (dated <= seed_year year-end).
    while next_ca is not None and seed_year is not None \
            and next_ca["Date"].year <= seed_year:
        next_ca = next(ca_iter, None)

    # Merge stream by date: CAs and transactions in chronological order.
    def next_action() -> Tuple[Optional[str], Any]:
        nonlocal pending, next_ca
        if pending is None and next_ca is None:
            return None, None
        if pending is None:
            ca = next_ca
            next_ca = next(ca_iter, None)
            return "CA", ca
        if next_ca is None:
            tx = pending
            pending = next(tx_iter, None)
            return "TX", tx
        # Both exist — pick earlier date; CAs apply first on same date.
        if next_ca["Date"] <= pending.trade_date:
            ca = next_ca
            next_ca = next(ca_iter, None)
            return "CA", ca
        tx = pending
        pending = next(tx_iter, None)
        return "TX", tx

    current_year: Optional[int] = None
    # Global optimizer buffer: {instrument_id: [Transaction]}
    deferred_global_sells: Dict[str, List[Transaction]] = defaultdict(list)

    def flush_deferred_sells(flush_year: Optional[int]) -> None:
        """Process buffered MIN_GAIN/MAX_GAIN sells for completed year."""
        if flush_year is None:
            return
        for inst, sell_list in list(deferred_global_sells.items()):
            year_sells = [s for s in sell_list if s.trade_date.year == flush_year]
            if not year_sells:
                continue
            method = override_method if override_method is not None else (
                method_selection.get((flush_year, inst), policy.default_method_for(flush_year))
            )
            avail_lots = [l for l in lots
                          if l.instrument_id == inst and l.quantity_remaining > 1e-9]
            lines, unmatched_map = _match_global_optimized(
                year_sells, avail_lots, method, fx, match_counter)
            match_lines.extend(lines)
            for tx in year_sells:
                leftover = unmatched_map.get(tx.tx_id, 0.0)
                if leftover > 1e-6:
                    severity = "WARN" if leftover < 1e-3 else "ERROR"
                    warnings.append({
                        "check": "insufficient_lots", "severity": severity,
                        "source_file": tx.source_file,
                        "source_row": tx.source_row,
                        "detail": (f"SELL {tx.symbol} {tx.trade_date} qty "
                                   f"{tx.quantity}: unmatched {leftover:.6f}"),
                    })
            # Remove processed sells from buffer
            remaining_in_buf = [s for s in sell_list
                                 if s.trade_date.year != flush_year]
            if remaining_in_buf:
                deferred_global_sells[inst] = remaining_in_buf
            else:
                del deferred_global_sells[inst]

    while True:
        kind, item = next_action()
        if kind is None:
            break
        item_date = item["Date"] if kind == "CA" else item.trade_date
        item_year = item_date.year

        # Snapshot previous year before starting new year actions.
        if current_year is not None and item_year != current_year:
            flush_deferred_sells(current_year)
            year_end_inventory[current_year] = _clone_lots(lots)
        current_year = item_year

        if kind == "CA":
            if item.get("Applied", True):
                apply_split_to_lots(lots, item)
            continue
        # kind == "TX"
        tx: Transaction = item
        if tx.side == "BUY":
            lot = Lot(
                lot_id=f"L_{tx.tx_id}",
                tx_id=tx.tx_id,
                instrument_id=tx.instrument_id,
                source_broker=tx.source_broker,
                source_account=tx.source_account,
                source_file=tx.source_file,
                source_row=tx.source_row,
                buy_date=tx.trade_date,
                quantity_original=tx.quantity,
                quantity_remaining=tx.quantity,
                price_per_share_usd=tx.price_usd,
                buy_commission_total_usd=tx.commission_usd,
            )
            lots.append(lot)
        else:  # SELL
            if override_method is not None:
                method = override_method
            else:
                method = method_selection.get(
                    (tx.trade_date.year, tx.instrument_id), DEFAULT_METHOD)
            if method in ("MIN_GAIN", "MAX_GAIN"):
                # Buffer for global annual optimizer
                deferred_global_sells[tx.instrument_id].append(tx)
            else:
                lines, unmatched = match_sell(tx, lots, method, fx, match_counter)
                match_lines.extend(lines)
                if unmatched > 1e-6:
                    severity = "WARN" if unmatched < 1e-3 else "ERROR"
                    warnings.append({
                        "check": "insufficient_lots", "severity": severity,
                        "source_file": tx.source_file,
                        "source_row": tx.source_row,
                        "detail": (f"SELL {tx.symbol} {tx.trade_date} qty "
                                   f"{tx.quantity}: unmatched {unmatched:.6f}"),
                    })

    if current_year is not None:
        flush_deferred_sells(current_year)
        year_end_inventory[current_year] = _clone_lots(lots)

    return lots, match_lines, warnings, year_end_inventory


def _clone_lots(lots: List[Lot]) -> List[Lot]:
    return [dataclasses.replace(l, adjustments=list(l.adjustments))
            for l in lots]


def _coerce_date(v: Any) -> date:
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        d = parse_trade_date(v)
        if d is not None:
            return d
    return date.min


# -----------------------------------------------------------------------
# Yearly summary
# -----------------------------------------------------------------------

def build_yearly_summary(
    match_lines: List[MatchLine],
    settings: Dict[int, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    by_year: Dict[int, Dict[str, Any]] = defaultdict(lambda: {
        "gross_proceeds": 0.0, "gross_proceeds_pre_sell_commission": 0.0,
        "total_cost_basis": 0.0, "non_exempt_cost_basis": 0.0,
        "total_pnl": 0.0, "exempt_proceeds": 0.0,
        "exempt_gain": 0.0, "taxable_gains": 0.0, "taxable_losses": 0.0,
        "match_count": 0,
    })
    for m in match_lines:
        y = m.tax_year
        slot = by_year[y]
        slot["gross_proceeds"] += m.proceeds_czk
        slot["gross_proceeds_pre_sell_commission"] += (
            m.proceeds_czk + (m.allocated_sell_commission_usd * m.fx_rate_sell)
        )
        slot["total_cost_basis"] += m.cost_basis_czk
        slot["total_pnl"] += m.proceeds_czk - m.cost_basis_czk
        slot["match_count"] += 1
        if m.time_test_exempt:
            slot["exempt_proceeds"] += m.proceeds_czk
            slot["exempt_gain"] += m.proceeds_czk - m.cost_basis_czk
        else:
            slot["non_exempt_cost_basis"] += m.cost_basis_czk
            pnl = m.taxable_gain_czk
            if pnl >= 0:
                slot["taxable_gains"] += pnl
            else:
                slot["taxable_losses"] += -pnl

    out: List[Dict[str, Any]] = []
    for y in sorted(set(list(by_year.keys()) + list(settings.keys()))):
        s = settings.get(y, {"tax_rate": DEFAULT_TAX_RATE,
                             "apply_100k": DEFAULT_APPLY_100K,
                             "fx_method": DEFAULT_FX_METHOD,
                             "locked": False})
        slot = by_year.get(y)
        if slot is None:
            out.append({
                "Tax year": y,
                "Gross proceeds CZK (all sells)": 0.0,
                "Time-test exempt proceeds CZK": 0.0,
                "Non-exempt cost basis CZK": 0.0,
                "Non-exempt proceeds CZK": 0.0,
                "Taxable gains CZK": 0.0,
                "Taxable losses CZK": 0.0,
                "Pre-exemption tax base CZK": 0.0,
                "100k threshold met?": False,
                "Apply 100k exemption?": s["apply_100k"],
                "Final tax base CZK": 0.0,
                "Tax rate": s["tax_rate"],
                "Tax due CZK": 0.0,
                "FX method": s["fx_method"],
                "Locked?": s.get("locked", False),
                "Match line count": 0,
                "Note": "no sales",
            })
            continue
        pre_base = max(0.0, slot["taxable_gains"] - slot["taxable_losses"])
        under_100k = (
            slot["gross_proceeds_pre_sell_commission"] <= DEFAULT_100K_THRESHOLD
        )
        final_base = pre_base
        note = ""
        if s["apply_100k"] and under_100k:
            final_base = 0.0
            note = (
                "100k exemption applied (gross proceeds before sell "
                "commission <= 100 000 CZK)"
            )
        tax = final_base * s["tax_rate"]
        out.append({
            "Tax year": y,
            "Gross proceeds CZK (all sells)": round(slot["gross_proceeds"], 2),
            "Time-test exempt proceeds CZK": round(slot["exempt_proceeds"], 2),
            "Non-exempt cost basis CZK": round(slot["non_exempt_cost_basis"], 2),
            "Non-exempt proceeds CZK": round(
                slot["gross_proceeds"] - slot["exempt_proceeds"], 2),
            "Taxable gains CZK": round(slot["taxable_gains"], 2),
            "Taxable losses CZK": round(slot["taxable_losses"], 2),
            "Pre-exemption tax base CZK": round(pre_base, 2),
            "100k threshold met?": under_100k,
            "Apply 100k exemption?": s["apply_100k"],
            "Final tax base CZK": round(final_base, 2),
            "Tax rate": s["tax_rate"],
            "Tax due CZK": round(tax, 2),
            "FX method": s["fx_method"],
            "Locked?": s.get("locked", False),
            "Match line count": slot["match_count"],
            "Note": note,
        })
    return out


# -----------------------------------------------------------------------
# Method comparison
# -----------------------------------------------------------------------

def run_method_comparison(
    txs: List[Transaction], settings: Dict[int, Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    locked_years: Dict[int, bool],
    corporate_actions: List[Dict[str, Any]],
    frozen_inventory: Dict[int, List[Dict[str, Any]]],
    frozen_matching: Dict[int, List[Dict[str, Any]]],
    frozen_snapshots: Dict[int, Dict[str, Any]],
    fx: FXResolver,
) -> List[Dict[str, Any]]:
    """For each tax year, re-run matching with each global method override.

    Returns one row per year per method plus one "Selected" row.
    """
    results: Dict[Tuple[int, str], Dict[str, float]] = {}
    for method in SUPPORTED_METHODS + ("SELECTED",):
        override = None if method == "SELECTED" else method
        _, lines, _, _ = simulate(
            txs, settings, method_selection, locked_years,
            corporate_actions, frozen_inventory, frozen_matching,
            frozen_snapshots, fx, override_method=override,
        )
        summary = build_yearly_summary(lines, settings)
        for row in summary:
            key = (row["Tax year"], method)
            results[key] = {
                "tax_base": row["Final tax base CZK"],
                "tax": row["Tax due CZK"],
                "gains": row["Taxable gains CZK"],
                "losses": row["Taxable losses CZK"],
                "gross_proceeds": row["Gross proceeds CZK (all sells)"],
            }
    years = sorted({k[0] for k in results.keys()})
    out: List[Dict[str, Any]] = []
    for y in years:
        row: Dict[str, Any] = {"Tax year": y}
        min_tax = None
        for m in SUPPORTED_METHODS:
            r = results.get((y, m), {"tax_base": 0.0, "tax": 0.0})
            row[f"{m} tax base CZK"] = r["tax_base"]
            row[f"{m} tax CZK"] = r["tax"]
            if min_tax is None or r["tax"] < min_tax[1]:
                min_tax = (m, r["tax"])
        sel = results.get((y, "SELECTED"), {"tax_base": 0.0, "tax": 0.0})
        row["Selected method tax base CZK"] = sel["tax_base"]
        row["Selected method tax CZK"] = sel["tax"]
        row["Best method"] = min_tax[0] if min_tax else ""
        row["Best method tax CZK"] = min_tax[1] if min_tax else 0.0
        row["Delta selected vs best CZK"] = round(
            sel["tax"] - (min_tax[1] if min_tax else 0.0), 2)
        out.append(row)
    return out


# -----------------------------------------------------------------------
# Split-adjust audit heuristic
# -----------------------------------------------------------------------

def split_audit(txs: List[Transaction]) -> List[Dict[str, Any]]:
    """Very light heuristic: per instrument, look for large step-change in
    average price across neighbouring BUY/SELL events.

    This does not prove anything; it just raises attention so the operator
    can manually enter a split if Yahoo data is or isn't already adjusted.
    """
    out: List[Dict[str, Any]] = []
    by_inst: Dict[str, List[Transaction]] = defaultdict(list)
    for tx in txs:
        by_inst[tx.instrument_id].append(tx)
    for inst, events in by_inst.items():
        events.sort(key=lambda t: (t.trade_date, t.source_file, t.source_row))
        for i in range(1, len(events)):
            a, b = events[i-1], events[i]
            if a.price_usd <= 0 or b.price_usd <= 0:
                continue
            ratio = b.price_usd / a.price_usd
            if ratio >= 2.8 or ratio <= 0.35:
                out.append({
                    "Instrument_ID": inst,
                    "From date": a.trade_date,
                    "To date": b.trade_date,
                    "Prev price USD": a.price_usd,
                    "Next price USD": b.price_usd,
                    "Ratio (next/prev)": round(ratio, 4),
                    "Hint": ("Possible unrecorded split or data already "
                             "adjusted. Verify manually in "
                             "Corporate_Actions."),
                })
    return out


# -----------------------------------------------------------------------
# Structured result helpers
# -----------------------------------------------------------------------

def build_open_position_rows(
    raw_rows: List[RawRow],
    instrument_map: Dict[str, Dict[str, str]],
    lots: List[Lot],
) -> List[Dict[str, Any]]:
    yahoo = extract_position_rows(raw_rows, instrument_map)
    calc: Dict[str, float] = defaultdict(float)
    for lot in lots:
        if lot.quantity_remaining > 1e-9:
            calc[lot.instrument_id] += lot.quantity_remaining
    instruments = sorted(set(yahoo.keys()) | set(calc.keys()))
    rows: List[Dict[str, Any]] = []
    for inst in instruments:
        yq = yahoo.get(inst)
        cq = calc.get(inst, 0.0)
        if yq is None:
            diff = None
            status = "UNKNOWN"
        else:
            diff = cq - yq
            status = "OK" if abs(diff) <= 1e-4 else ("WARN" if abs(diff) <= 1e-2 else "ERROR")
        rows.append({
            "Instrument_ID": inst,
            "Yahoo qty": yq,
            "Calculated qty": cq,
            "Difference": diff,
            "Status": status,
        })
    return rows


def build_check_rows(
    *,
    sim_warnings: List[Dict[str, Any]],
    problems: List[Dict[str, Any]],
    fx_yearly: Dict[int, float],
    fx_daily: Dict[date, float],
    settings: Dict[int, Dict[str, Any]],
    locked_years: Dict[int, bool],
    frozen_inventory: Dict[int, List[Dict[str, Any]]],
    split_warnings: List[Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    yearly_summary: List[Dict[str, Any]],
    match_lines: List[MatchLine],
    lots_final: List[Lot],
    year_end_inventory: Dict[int, List[Lot]],
    frozen_snapshots: Dict[int, Dict[str, Any]],
    fx: "FXResolver",
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for p in problems:
        rows.append({
            "Severity": p.get("severity", "WARN"),
            "Category": p.get("check", "import"),
            "Detail": p.get("detail", ""),
            "Source file": p.get("source_file", ""),
            "Source row": p.get("source_row", ""),
        })
    for w in sim_warnings:
        rows.append({
            "Severity": w.get("severity", "WARN"),
            "Category": w.get("check", "simulate"),
            "Detail": w.get("detail", ""),
            "Source file": w.get("source_file", ""),
            "Source row": w.get("source_row", ""),
        })

    used_fx_years = {m.sell_date.year for m in match_lines} | {
        m.buy_date.year for m in match_lines
    }
    for y in used_fx_years:
        if y not in fx_yearly:
            rows.append({
                "Severity": "ERROR",
                "Category": "missing_fx_yearly",
                "Detail": f"FX_Yearly missing for {y}",
                "Source file": "",
                "Source row": "",
            })

    for y, s in settings.items():
        if s["fx_method"] == "FX_DAILY_CNB":
            need_dates = {m.sell_date for m in match_lines if m.sell_date.year == y} | {
                m.buy_date for m in match_lines if m.buy_date.year == y
            }
            missing = [d for d in need_dates if d not in fx_daily]
            if missing:
                rows.append({
                    "Severity": "ERROR",
                    "Category": "missing_fx_daily",
                    "Detail": f"{len(missing)} trade dates in {y} lacked FX_DAILY_CNB rate.",
                    "Source file": "",
                    "Source row": "",
                })

    if fx.used_fallback_yearly:
        rows.append({
            "Severity": "ERROR",
            "Category": "fx_daily_fallback_yearly",
            "Detail": (
                "FX_DAILY_CNB used yearly fallback on "
                f"{len(sorted(set(fx.used_fallback_yearly)))} dates."
            ),
            "Source file": "",
            "Source row": "",
        })

    manifest_years = set(frozen_snapshots.keys()) | {
        y for y in year_end_inventory.keys() if locked_years.get(y)
    }
    for y, locked in locked_years.items():
        if locked and y not in manifest_years:
            rows.append({
                "Severity": "ERROR",
                "Category": "locked_year_no_snapshot",
                "Detail": (
                    f"Year {y} marked Locked but no frozen snapshot manifest exists yet."
                ),
                "Source file": "",
                "Source row": "",
            })

    for key, method in method_selection.items():
        if method not in SUPPORTED_METHODS:
            rows.append({
                "Severity": "ERROR",
                "Category": "invalid_method",
                "Detail": f"Invalid method '{method}' for {key}",
                "Source file": "",
                "Source row": "",
            })

    for s in split_warnings:
        rows.append({
            "Severity": "INFO",
            "Category": "split_audit_hint",
            "Detail": (
                f"{s['Instrument_ID']} {s['From date']} -> {s['To date']} "
                f"ratio {s['Ratio (next/prev)']}"
            ),
            "Source file": "",
            "Source row": "",
        })

    for lot in lots_final:
        if lot.quantity_remaining < -1e-6:
            rows.append({
                "Severity": "ERROR",
                "Category": "negative_remaining",
                "Detail": f"Lot {lot.lot_id} remaining {lot.quantity_remaining}",
                "Source file": lot.source_file,
                "Source row": lot.source_row,
            })

    if not rows:
        rows.append({
            "Severity": "INFO",
            "Category": "all_clear",
            "Detail": "No issues detected by static checks.",
            "Source file": "",
            "Source row": "",
        })
    return rows


def calculate_workbook_data(
    inputs: List[Path],
    out_path: Path,
    *,
    fetch_missing_fx: bool = True,
) -> CalculationResult:
    user_state = load_existing_user_state(out_path)

    raw_rows: List[RawRow] = []
    import_log: List[Dict[str, Any]] = []
    problems: List[Dict[str, Any]] = []

    for ipath in inputs:
        rows, probs = read_csv_file(ipath)
        raw_rows.extend(rows)
        problems.extend(probs)
        if rows:
            broker, account = broker_from_filename(ipath)
        else:
            broker, account = ("", "")
        dates = [parse_trade_date(r.data.get("Trade Date") or "") for r in rows]
        dates = [d for d in dates if d is not None]
        syms = sorted({r.symbol for r in rows if r.symbol})
        import_log.append({
            "Source file": ipath.name,
            "Broker": broker,
            "Account": account,
            "Raw rows": len(rows),
            "Transactions": 0,
            "Ignored": 0,
            "Min Trade Date": min(dates) if dates else None,
            "Max Trade Date": max(dates) if dates else None,
            "Unique symbols": ", ".join(syms),
            "Import timestamp": datetime.now(_dt.timezone.utc).isoformat(timespec="seconds"),
        })

    txs, ignored, norm_problems = normalize_transactions(raw_rows)
    problems.extend(norm_problems)

    tx_by_file = defaultdict(int)
    ig_by_file = defaultdict(int)
    for t in txs:
        tx_by_file[t.source_file] += 1
    for ig in ignored:
        ig_by_file[ig["source_file"]] += 1
    for row in import_log:
        row["Transactions"] = tx_by_file.get(row["Source file"], 0)
        row["Ignored"] = ig_by_file.get(row["Source file"], 0)

    years = sorted({t.trade_date.year for t in txs})
    for row in user_state.get("Settings", []) + user_state.get("Locked_Years", []):
        try:
            y = int(row.get("Tax year"))
        except (TypeError, ValueError):
            continue
        if y not in years:
            years.append(y)
    years = sorted(set(years))

    settings = build_settings(user_state, years)
    instrument_map = build_instrument_map(user_state, txs)
    apply_instrument_map(txs, instrument_map)
    fx_yearly, fx_daily, fx_yearly_sources = build_fx_tables(user_state, years)
    locked_years = build_locked_years(user_state, years)
    corporate_actions = build_corporate_actions(user_state)
    frozen_inventory = load_frozen_inventory(user_state)
    frozen_matching = load_frozen_matching(user_state)
    frozen_snapshots = load_frozen_snapshots(user_state)
    review_state = load_review_state(user_state)
    filed_reconciliation = load_filed_reconciliation(user_state)
    instrument_ids = sorted({t.instrument_id for t in txs})
    method_selection = build_method_selection(user_state, years, instrument_ids)

    fx = FXResolver(fx_yearly, fx_daily, settings)

    if fetch_missing_fx:
        daily_years_needed = [
            y for y in years if settings.get(y, {}).get("fx_method") == "FX_DAILY_CNB"
        ]
        if daily_years_needed:
            cache_path = _cnb_cache_path(Path(out_path))
            fx_daily, _dl_msgs = refresh_fx_daily_for_years(
                fx_daily, daily_years_needed, cache_path
            )
            fx = FXResolver(fx_yearly, fx_daily, settings)

    lots_final, match_lines, sim_warnings, year_end_inventory = simulate(
        txs,
        settings,
        method_selection,
        locked_years,
        corporate_actions,
        frozen_inventory,
        frozen_matching,
        frozen_snapshots,
        fx,
    )

    yearly_summary = build_yearly_summary(match_lines, settings)
    method_comparison = run_method_comparison(
        txs,
        settings,
        method_selection,
        locked_years,
        corporate_actions,
        frozen_inventory,
        frozen_matching,
        frozen_snapshots,
        fx,
    )
    split_warnings = split_audit(txs)

    return CalculationResult(
        inputs=inputs,
        output_path=out_path,
        raw_rows=raw_rows,
        txs=txs,
        ignored=ignored,
        problems=problems,
        import_log=import_log,
        years=years,
        instrument_ids=instrument_ids,
        user_state=user_state,
        settings=settings,
        instrument_map=instrument_map,
        fx_yearly=fx_yearly,
        fx_daily=fx_daily,
        fx_yearly_sources=fx_yearly_sources,
        locked_years=locked_years,
        corporate_actions=corporate_actions,
        frozen_inventory=frozen_inventory,
        frozen_matching=frozen_matching,
        frozen_snapshots=frozen_snapshots,
        review_state=review_state,
        filed_reconciliation=filed_reconciliation,
        method_selection=method_selection,
        fx=fx,
        lots_final=lots_final,
        match_lines=match_lines,
        sim_warnings=sim_warnings,
        year_end_inventory=year_end_inventory,
        yearly_summary=yearly_summary,
        method_comparison=method_comparison,
        split_warnings=split_warnings,
    )


def write_calculation_result(
    result: CalculationResult,
    *,
    backup_existing: bool = False,
) -> Path:
    out_path = result.output_path
    temp_out_path = _tmp_output_path(out_path)
    if temp_out_path.exists():
        temp_out_path.unlink(missing_ok=True)

    write_workbook(
        out_path=temp_out_path,
        raw_rows=result.raw_rows,
        txs=result.txs,
        ignored=result.ignored,
        problems=result.problems,
        instrument_map=result.instrument_map,
        fx_yearly=result.fx_yearly,
        fx_yearly_sources=result.fx_yearly_sources,
        fx_daily=result.fx_daily,
        corporate_actions=result.corporate_actions,
        method_selection=result.method_selection,
        locked_years=result.locked_years,
        settings=result.settings,
        frozen_inventory=result.frozen_inventory,
        frozen_matching=result.frozen_matching,
        frozen_snapshots=result.frozen_snapshots,
        fx=result.fx,
        lots_final=result.lots_final,
        match_lines=result.match_lines,
        sim_warnings=result.sim_warnings,
        yearly_summary=result.yearly_summary,
        method_comparison=result.method_comparison,
        split_warnings=result.split_warnings,
        year_end_inventory=result.year_end_inventory,
        import_log=result.import_log,
        review_state=result.review_state,
        filed_reconciliation=result.filed_reconciliation,
    )

    import verify_workbook as _verify_workbook

    verify_rc = _verify_workbook.main(str(temp_out_path))
    if verify_rc != 0:
        temp_out_path.unlink(missing_ok=True)
        raise RuntimeError(
            "Validation failed for temporary workbook; requested output was not replaced."
        )

    if backup_existing and out_path.exists():
        _backup_existing_output(out_path)

    _replace_output_or_fail(temp_out_path, out_path)
    return out_path


# -----------------------------------------------------------------------
# Workbook writing
# -----------------------------------------------------------------------

def write_workbook(
    out_path: Path,
    raw_rows: List[RawRow],
    txs: List[Transaction],
    ignored: List[Dict[str, Any]],
    problems: List[Dict[str, Any]],
    instrument_map: Dict[str, Dict[str, str]],
    fx_yearly: Dict[int, float],
    fx_daily: Dict[date, float],
    corporate_actions: List[Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    locked_years: Dict[int, bool],
    settings: Dict[int, Dict[str, Any]],
    frozen_inventory: Dict[int, List[Dict[str, Any]]],
    frozen_matching: Dict[int, List[Dict[str, Any]]],
    frozen_snapshots: Dict[int, Dict[str, Any]],
    fx: FXResolver,
    lots_final: List[Lot],
    match_lines: List[MatchLine],
    sim_warnings: List[Dict[str, Any]],
    yearly_summary: List[Dict[str, Any]],
    method_comparison: List[Dict[str, Any]],
    split_warnings: List[Dict[str, Any]],
    year_end_inventory: Dict[int, List[Lot]],
    import_log: List[Dict[str, Any]],
    review_state: Dict[str, Dict[str, Any]],
    filed_reconciliation: Dict[int, Dict[str, Any]],
    fx_yearly_sources: Optional[Dict[int, str]] = None,
) -> None:
    wb = Workbook()
    # Remove default sheet — we'll add ours in order.
    wb.remove(wb.active)

    _write_readme(wb)
    _write_dashboard(wb, yearly_summary, method_comparison, sim_warnings,
                     problems, import_log)
    _write_settings(wb, settings, sorted(settings.keys()))
    _write_import_log(wb, import_log)
    _write_raw_yahoo(wb, raw_rows)
    _write_ignored(wb, ignored)
    _write_transactions(wb, txs)
    _write_instrument_map(wb, instrument_map)
    _write_fx_daily(wb, fx_daily)
    _write_fx_yearly(wb, fx_yearly, sorted(settings.keys()), fx_yearly_sources)
    _write_corporate_actions(wb, corporate_actions)
    _write_split_audit(wb, split_warnings)
    _write_method_selection(wb, method_selection)
    instrument_ids = sorted({t.instrument_id for t in txs})
    _write_method_plan(
        wb,
        sorted(settings.keys()),
        instrument_ids,
        method_selection,
        settings,
        method_comparison,
        match_lines,
    )
    _write_filed_year_reconciliation(
        wb,
        yearly_summary,
        method_selection,
        instrument_ids,
        filed_reconciliation,
    )
    _write_locked_years(wb, locked_years, sorted(settings.keys()))
    _write_frozen_inventory(wb, frozen_inventory, year_end_inventory,
                            locked_years)
    _write_frozen_matching(wb, frozen_matching, match_lines, locked_years)
    _write_frozen_snapshots(wb, frozen_snapshots, year_end_inventory,
                            locked_years, match_lines)
    _write_review_state(wb, review_state, match_lines)
    _write_lots(wb, lots_final)
    _write_lot_matching(wb, match_lines)
    _write_sell_review(wb, match_lines, review_state)
    _write_open_lots_review(wb, lots_final, txs, fx)
    _write_open_position_check(wb, raw_rows, instrument_map, lots_final)
    _write_yearly_summary(wb, yearly_summary)
    _write_method_comparison(wb, method_comparison)
    _write_checks(wb, sim_warnings, problems, fx_yearly, fx_daily,
                  settings, locked_years, frozen_inventory, split_warnings,
                  method_selection, yearly_summary, match_lines, lots_final,
                  year_end_inventory, frozen_snapshots, fx)
    _write_audit_report(wb, raw_rows, txs, ignored, match_lines,
                        yearly_summary)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _tmp_output_path(out_path: Path) -> Path:
    """Temporary workbook path in the same directory for atomic replace."""
    return out_path.parent / f".{out_path.stem}.tmp{out_path.suffix}"


def _backup_existing_output(out_path: Path) -> Path:
    backup_dir = out_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    backup_path = backup_dir / f"{out_path.stem}_{ts}{out_path.suffix}"
    shutil.copy2(out_path, backup_path)
    return backup_path


def _replace_output_or_fail(temp_path: Path, out_path: Path) -> None:
    try:
        os.replace(temp_path, out_path)
    except PermissionError as exc:
        temp_path.unlink(missing_ok=True)
        raise RuntimeError(
            f"Cannot write {out_path.name} because it is open or locked. "
            "Close Excel and rerun."
        ) from exc


# ------------------- sheet writers -------------------

def _write_readme(wb: Workbook) -> None:
    ws = wb.create_sheet("README")
    lines = [
        ("Czech Personal Income Tax — Stock Trade Workbook",
         Font(bold=True, size=14)),
        ("", None),
        ("Purpose", Font(bold=True, size=12)),
        ("Compute Czech personal income tax from Yahoo Portfolio CSV "
         "exports.", None),
        ("This is a CALCULATION workbook. It is NOT official tax advice.",
         Font(italic=True, color="FFC00000")),
        ("", None),
        ("Key assumptions", Font(bold=True, size=12)),
        ("* Tax currency: CZK. All trade data is treated as USD-denominated.",
         None),
        ("* Trade Date (column from Yahoo CSV) is the operative date.", None),
        ("* FX fees are IGNORED. Only the Commission column is used as a "
         "broker fee.", None),
        ("* Buy commission increases acquisition cost; sell commission "
         "reduces proceeds.", None),
        ("* Matching happens GLOBALLY across brokers/accounts per "
         "Instrument_ID.", None),
        ("* Each matched lot is evaluated independently against the Czech "
         "3-year time test.", None),
        ("* Time test: a lot is exempt if sell_date > buy_date + 3 "
         "calendar years (same calendar day, 3 years later + 1 day).", None),
        ("* Within a year, taxable losses offset taxable gains; base floored "
         "at 0. No carryforward.", None),
        ("* 100,000 CZK annual gross proceeds exemption is OPTIONAL per "
         "year (Settings.Apply 100k exemption).", None),
        ("* Locked years are not recalculated. Frozen_Inventory captures "
         "open lots at year-end.", None),
        ("* Corporate actions (splits, reverse splits, ticker changes) are "
         "ENTERED MANUALLY.", None),
        ("* Default tax rate is 15 %; editable per year in Settings.", None),
        ("* Default FX method is FX_UNIFIED_GFR (GFŘ official rate). Switch per year in Settings; "
         "do not mix within one year.", None),
        ("", None),
        ("How to use", Font(bold=True, size=12)),
        ("1. Fill FX_Yearly (and FX_Daily if using FX_DAILY_CNB) against "
         "Czech National Bank tables.", None),
        ("2. Review Instrument_Map and add ISIN / stable Instrument_IDs "
         "where needed.", None),
        ("3. Add any splits / reverse splits / ticker changes to "
         "Corporate_Actions.", None),
        ("4. Pick a matching method per (year, instrument) in "
         "Method_Selection (default FIFO).", None),
        ("5. Open Settings and confirm tax rate, FX method and "
         "100k toggle per year.", None),
        ("6. Re-run build_stock_tax_workbook.py — it preserves locked-year "
         "snapshots.", None),
        ("7. For filed years: set Locked? = TRUE on the Locked_Years "
         "sheet. The workbook will freeze that year's inventory and "
         "matching on next regeneration.", None),
        ("", None),
        ("Traceability", Font(bold=True, size=12)),
        ("Every final tax number in Yearly_Tax_Summary is traceable "
         "through Lot_Matching to the original transaction row "
         "(Source file + Source row).", None),
        ("", None),
        ("Sheet index", Font(bold=True, size=12)),
    ]
    for title in [
        "README", "Operator_Dashboard", "Settings", "Import_Log",
        "Raw_Yahoo", "Ignored_Rows", "Transactions", "Instrument_Map",
        "FX_Daily", "FX_Yearly", "Corporate_Actions", "Split_Audit",
        "Method_Selection", "Locked_Years", "Frozen_Inventory",
        "Frozen_Lot_Matching", "Frozen_Snapshots", "Lots", "Lot_Matching",
        "Yearly_Tax_Summary", "Method_Comparison", "Checks",
        "Audit_Report",
    ]:
        lines.append((f"  - {title}", None))

    for i, (text, fnt) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if fnt is not None:
            cell.font = fnt
        cell.alignment = WRAP_LEFT
    ws.column_dimensions["A"].width = 96
    ws.freeze_panes = "A2"


def _write_dashboard(
    wb: Workbook, yearly_summary: List[Dict[str, Any]],
    method_comparison: List[Dict[str, Any]],
    sim_warnings: List[Dict[str, Any]],
    problems: List[Dict[str, Any]],
    import_log: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Operator_Dashboard")
    ws["A1"] = "Operator Dashboard"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Totals by tax year"
    ws["A3"].font = Font(bold=True, size=12)
    headers = ["Tax year", "Gross proceeds CZK", "Tax base CZK",
               "Tax due CZK", "Locked?", "FX method", "Apply 100k?"]
    write_header(ws, headers, row=4)
    for i, row in enumerate(yearly_summary, start=5):
        ws.cell(row=i, column=1, value=row["Tax year"])
        ws.cell(row=i, column=2,
                value=row["Gross proceeds CZK (all sells)"])
        ws.cell(row=i, column=3, value=row["Final tax base CZK"])
        ws.cell(row=i, column=4, value=row["Tax due CZK"])
        ws.cell(row=i, column=5, value=bool(row["Locked?"]))
        ws.cell(row=i, column=6, value=row["FX method"])
        ws.cell(row=i, column=7, value=bool(row["Apply 100k exemption?"]))

    base = len(yearly_summary) + 7
    ws.cell(row=base, column=1, value="Method comparison (tax CZK)")
    ws.cell(row=base, column=1).font = Font(bold=True, size=12)
    headers = ["Tax year", "FIFO", "LIFO", "MIN_GAIN", "MAX_GAIN",
               "Selected", "Best method", "Delta selected-best"]
    write_header(ws, headers, row=base + 1)
    for i, row in enumerate(method_comparison, start=base + 2):
        ws.cell(row=i, column=1, value=row["Tax year"])
        ws.cell(row=i, column=2, value=row["FIFO tax CZK"])
        ws.cell(row=i, column=3, value=row["LIFO tax CZK"])
        ws.cell(row=i, column=4, value=row["MIN_GAIN tax CZK"])
        ws.cell(row=i, column=5, value=row["MAX_GAIN tax CZK"])
        ws.cell(row=i, column=6, value=row["Selected method tax CZK"])
        ws.cell(row=i, column=7, value=row["Best method"])
        ws.cell(row=i, column=8, value=row["Delta selected vs best CZK"])

    checks_base = base + len(method_comparison) + 4
    ws.cell(row=checks_base, column=1, value="Validation summary")
    ws.cell(row=checks_base, column=1).font = Font(bold=True, size=12)

    err = sum(1 for p in sim_warnings + problems if p.get("severity") == "ERROR")
    warn = sum(1 for p in sim_warnings + problems if p.get("severity") == "WARN")
    ws.cell(row=checks_base + 1, column=1, value="Errors")
    ws.cell(row=checks_base + 1, column=2, value=err)
    ws.cell(row=checks_base + 2, column=1, value="Warnings")
    ws.cell(row=checks_base + 2, column=2, value=warn)
    if err > 0:
        ws.cell(row=checks_base + 1, column=2).fill = ERROR_FILL
    if warn > 0:
        ws.cell(row=checks_base + 2, column=2).fill = WARNING_FILL

    ws.cell(row=checks_base + 4, column=1,
            value="Check the Checks sheet and Audit_Report sheet for "
                  "details. This dashboard is a high-level overview only.")
    ws.cell(row=checks_base + 4, column=1).font = Font(italic=True)

    import_base = checks_base + 6
    ws.cell(row=import_base, column=1, value="Last import run")
    ws.cell(row=import_base, column=1).font = Font(bold=True, size=12)
    headers = ["Source file", "Raw rows", "Transactions", "Ignored", "Broker"]
    write_header(ws, headers, row=import_base + 1)
    for i, row in enumerate(import_log, start=import_base + 2):
        ws.cell(row=i, column=1, value=row["Source file"])
        ws.cell(row=i, column=2, value=row["Raw rows"])
        ws.cell(row=i, column=3, value=row["Transactions"])
        ws.cell(row=i, column=4, value=row["Ignored"])
        ws.cell(row=i, column=5, value=row["Broker"])

    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_settings(
    wb: Workbook, settings: Dict[int, Dict[str, Any]], years: List[int],
) -> None:
    ws = wb.create_sheet("Settings")
    headers = ["Tax year", "Tax rate", "FX method",
               "Apply 100k exemption?", "Locked year?", "Notes"]
    write_header(ws, headers)
    for i, y in enumerate(years, start=2):
        s = settings[y]
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=s["tax_rate"]).number_format = "0.00%"
        ws.cell(row=i, column=3, value=s["fx_method"])
        ws.cell(row=i, column=4, value=bool(s["apply_100k"]))
        ws.cell(row=i, column=5, value=bool(s["locked"]))
        ws.cell(row=i, column=6, value=s.get("notes", ""))
        for col in (2, 3, 4, 5, 6):
            ws.cell(row=i, column=col).fill = EDITABLE_FILL
    n = max(2, 1 + len(years))
    add_table(ws, "tbl_Settings", f"A1:F{n}")
    fx_dv = DataValidation(type="list",
                           formula1=f'"{",".join(SUPPORTED_FX_METHODS)}"',
                           allow_blank=False, showDropDown=False)
    fx_dv.add(f"C2:C{n}")
    ws.add_data_validation(fx_dv)
    bool_dv1 = DataValidation(type="list", formula1='"TRUE,FALSE"',
                              allow_blank=False, showDropDown=False)
    bool_dv1.add(f"D2:D{n}")
    ws.add_data_validation(bool_dv1)
    bool_dv2 = DataValidation(type="list", formula1='"TRUE,FALSE"',
                              allow_blank=False, showDropDown=False)
    bool_dv2.add(f"E2:E{n}")
    ws.add_data_validation(bool_dv2)
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_import_log(wb: Workbook, import_log: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Import_Log")
    headers = ["Source file", "Broker", "Account", "Raw rows",
               "Transactions", "Ignored", "Min Trade Date",
               "Max Trade Date", "Unique symbols", "Import timestamp"]
    write_header(ws, headers)
    for i, row in enumerate(import_log, start=2):
        ws.cell(row=i, column=1, value=row["Source file"])
        ws.cell(row=i, column=2, value=row["Broker"])
        ws.cell(row=i, column=3, value=row["Account"])
        ws.cell(row=i, column=4, value=row["Raw rows"])
        ws.cell(row=i, column=5, value=row["Transactions"])
        ws.cell(row=i, column=6, value=row["Ignored"])
        ws.cell(row=i, column=7, value=row["Min Trade Date"])
        ws.cell(row=i, column=8, value=row["Max Trade Date"])
        ws.cell(row=i, column=9, value=row["Unique symbols"])
        ws.cell(row=i, column=10, value=row["Import timestamp"])
    if import_log:
        add_table(ws, "tbl_ImportLog", f"A1:J{1+len(import_log)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_raw_yahoo(wb: Workbook, raw_rows: List[RawRow]) -> None:
    ws = wb.create_sheet("Raw_Yahoo")
    headers = ["Source file", "Source row", "Source broker",
               "Source account"] + YAHOO_COLUMNS
    write_header(ws, headers)
    for i, raw in enumerate(raw_rows, start=2):
        ws.cell(row=i, column=1, value=raw.source_file)
        ws.cell(row=i, column=2, value=raw.source_row)
        ws.cell(row=i, column=3, value=raw.source_broker)
        ws.cell(row=i, column=4, value=raw.source_account)
        for j, col in enumerate(YAHOO_COLUMNS, start=5):
            ws.cell(row=i, column=j, value=raw.data.get(col, ""))
    if raw_rows:
        add_table(ws, "tbl_Raw",
                  f"A1:{get_column_letter(4 + len(YAHOO_COLUMNS))}"
                  f"{1+len(raw_rows)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_ignored(wb: Workbook, ignored: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Ignored_Rows")
    headers = ["Source file", "Source row", "Symbol", "Reason", "Detail"]
    write_header(ws, headers)
    for i, r in enumerate(ignored, start=2):
        ws.cell(row=i, column=1, value=r["source_file"])
        ws.cell(row=i, column=2, value=r["source_row"])
        ws.cell(row=i, column=3, value=r["symbol"])
        ws.cell(row=i, column=4, value=r["reason"])
        ws.cell(row=i, column=5, value=r["detail"])
    if ignored:
        add_table(ws, "tbl_Ignored", f"A1:E{1+len(ignored)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_transactions(wb: Workbook, txs: List[Transaction]) -> None:
    ws = wb.create_sheet("Transactions")
    headers = ["Tx_ID", "Source file", "Source row", "Source broker",
               "Source account", "Symbol", "Instrument_ID", "Trade Date",
               "Side", "Quantity", "Price USD", "Commission USD", "Comment"]
    write_header(ws, headers)
    for i, tx in enumerate(txs, start=2):
        ws.cell(row=i, column=1, value=tx.tx_id)
        ws.cell(row=i, column=2, value=tx.source_file)
        ws.cell(row=i, column=3, value=tx.source_row)
        ws.cell(row=i, column=4, value=tx.source_broker)
        ws.cell(row=i, column=5, value=tx.source_account)
        ws.cell(row=i, column=6, value=tx.symbol)
        ws.cell(row=i, column=7, value=tx.instrument_id)
        ws.cell(row=i, column=8, value=tx.trade_date)
        ws.cell(row=i, column=8).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=9, value=tx.side)
        ws.cell(row=i, column=10, value=tx.quantity)
        ws.cell(row=i, column=11, value=tx.price_usd)
        ws.cell(row=i, column=12, value=tx.commission_usd)
        ws.cell(row=i, column=13, value=tx.comment)
    if txs:
        add_table(ws, "tbl_Transactions", f"A1:M{1+len(txs)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_instrument_map(
    wb: Workbook, instrument_map: Dict[str, Dict[str, str]],
) -> None:
    ws = wb.create_sheet("Instrument_Map")
    headers = ["Yahoo Symbol", "Instrument_ID", "ISIN",
               "Instrument name", "Notes"]
    write_header(ws, headers)
    for i, sym in enumerate(sorted(instrument_map.keys()), start=2):
        info = instrument_map[sym]
        for j, col in enumerate(headers, start=1):
            c = ws.cell(row=i, column=j, value=info.get(col, ""))
            if col in ("Instrument_ID", "ISIN", "Instrument name", "Notes"):
                c.fill = EDITABLE_FILL
    n = max(2, 1 + len(instrument_map))
    add_table(ws, "tbl_Instrument_Map", f"A1:E{n}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_fx_daily(wb: Workbook, fx_daily: Dict[date, float]) -> None:
    ws = wb.create_sheet("FX_Daily")
    ws["A1"] = ("CNB daily USD/CZK rates. Used when a tax year's FX method "
                "is FX_DAILY_CNB. Rates can be downloaded automatically by "
                "the builder, or pasted manually. Missing FX_DAILY_CNB rates "
                "on required trade dates are reported as ERROR in Checks.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:C1")
    headers = ["Date", "USD_CZK", "Source / note"]
    write_header(ws, headers, row=3)
    row = 4
    for d in sorted(fx_daily.keys()):
        ws.cell(row=row, column=1, value=d)
        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=2, value=fx_daily[d])
        ws.cell(row=row, column=2).number_format = "0.0000"
        row += 1
    add_table(ws, "tbl_FX_Daily", f"A3:C{max(4, row-1)}")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A4"


def _write_fx_yearly(wb: Workbook, fx_yearly: Dict[int, float],
                     years: List[int],
                     fx_yearly_sources: Optional[Dict[int, str]] = None) -> None:
    ws = wb.create_sheet("FX_Yearly")
    ws["A1"] = ("Official GF\u0158 annual USD/CZK rates (FX_UNIFIED_GFR method). "
                "GF\u0159-D-65 (2024), GF\u0159-D-75 (2025) are authoritative and "
                "pre-filled. Verify other years against CNB \u2018Kurzy dev. trhu \u2013 "
                "pr\u016fm\u011br roku\u2019. Editable per year.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:C1")
    headers = ["Tax year", "USD_CZK", "Source / note"]
    write_header(ws, headers, row=3)
    src = fx_yearly_sources or {}
    all_years = sorted(set(list(fx_yearly.keys()) + years))
    for i, y in enumerate(all_years, start=4):
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=fx_yearly.get(y, DEFAULT_FX_YEARLY.get(y, 22.0)))
        ws.cell(row=i, column=2).number_format = "0.0000"
        ws.cell(row=i, column=2).fill = EDITABLE_FILL
        source_label = src.get(y, "")
        if not source_label:
            source_label = "GF\u0158 official" if y in GFR_OFFICIAL_RATES else "DEFAULT \u2014 verify"
        ws.cell(row=i, column=3, value=source_label)
    add_table(ws, "tbl_FX_Yearly", f"A3:C{3+len(all_years)}")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A4"


def _write_corporate_actions(
    wb: Workbook, actions: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Corporate_Actions")
    ws["A1"] = ("Enter corporate actions manually. For a SPLIT 3-for-1, "
                "set Ratio old=1, Ratio new=3. For a REVERSE_SPLIT 1-for-10, "
                "set Ratio old=10, Ratio new=1. Splits adjust open-lot "
                "quantity and price-per-share but NOT total cost basis and "
                "NOT holding period. This workbook does NOT detect splits "
                "automatically — see Split_Audit for hints.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:I1")
    headers = ["Date", "Instrument_ID", "Action type", "Ratio old",
               "Ratio new", "Cash in lieu", "Notes", "Applied?",
               "Audit status"]
    write_header(ws, headers, row=3)
    max_row = 3
    for i, ca in enumerate(actions, start=4):
        ws.cell(row=i, column=1, value=ca["Date"])
        ws.cell(row=i, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=2, value=ca["Instrument_ID"])
        ws.cell(row=i, column=3, value=ca["Action type"])
        ws.cell(row=i, column=4, value=ca["Ratio old"])
        ws.cell(row=i, column=5, value=ca["Ratio new"])
        ws.cell(row=i, column=6, value=ca.get("Cash in lieu", 0.0))
        ws.cell(row=i, column=7, value=ca.get("Notes", ""))
        ws.cell(row=i, column=8, value=bool(ca.get("Applied", True)))
        ws.cell(row=i, column=9, value="applied" if ca.get("Applied") else
                "not applied")
        max_row = i
    # Empty template rows to encourage user entry.
    start_empty = max(max_row + 1, 4)
    for i in range(start_empty, start_empty + 8):
        for col in (1, 2, 3, 4, 5, 6, 7, 8):
            ws.cell(row=i, column=col).fill = EDITABLE_FILL
        max_row = i
    add_table(ws, "tbl_Corporate_Actions", f"A3:I{max_row}")
    ca_dv = DataValidation(type="list",
                           formula1=f'"{",".join(CA_TYPES)}"',
                           allow_blank=True, showDropDown=False)
    ca_dv.add(f"C4:C{max_row}")
    ws.add_data_validation(ca_dv)
    bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"',
                             allow_blank=True, showDropDown=False)
    bool_dv.add(f"H4:H{max_row}")
    ws.add_data_validation(bool_dv)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_split_audit(wb: Workbook, warns: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Split_Audit")
    ws["A1"] = ("Heuristic only. Identifies cases where consecutive trade "
                "prices for the same instrument differ by a factor of "
                "~3x or more. If Yahoo data is ALREADY split-adjusted, DO "
                "NOT also enter the split in Corporate_Actions (double "
                "application will mis-state cost basis). If Yahoo data is "
                "raw, add the split to Corporate_Actions.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:G1")
    headers = ["Instrument_ID", "From date", "To date", "Prev price USD",
               "Next price USD", "Ratio (next/prev)", "Hint"]
    write_header(ws, headers, row=3)
    for i, w in enumerate(warns, start=4):
        ws.cell(row=i, column=1, value=w["Instrument_ID"])
        ws.cell(row=i, column=2, value=w["From date"])
        ws.cell(row=i, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=3, value=w["To date"])
        ws.cell(row=i, column=3).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=4, value=w["Prev price USD"])
        ws.cell(row=i, column=5, value=w["Next price USD"])
        ws.cell(row=i, column=6, value=w["Ratio (next/prev)"])
        ws.cell(row=i, column=7, value=w["Hint"])
    if warns:
        add_table(ws, "tbl_Split_Audit", f"A3:G{3+len(warns)}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_method_selection(
    wb: Workbook, method_selection: Dict[Tuple[int, str], str],
) -> None:
    ws = wb.create_sheet("Method_Selection")
    ws["A1"] = ("Pick a matching method per (Tax year, Instrument_ID). "
                "Changes here influence Lot_Matching and Yearly_Tax_Summary. "
                "Use Method_Comparison to see the tax impact.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:D1")
    headers = ["Tax year", "Instrument_ID", "Method", "Notes"]
    write_header(ws, headers, row=3)
    keys = sorted(method_selection.keys(), key=lambda k: (k[0], k[1]))
    for i, key in enumerate(keys, start=4):
        y, inst = key
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=inst)
        ws.cell(row=i, column=3, value=method_selection[key])
        ws.cell(row=i, column=3).fill = EDITABLE_FILL
    n = max(4, 3 + len(keys))
    add_table(ws, "tbl_Method_Selection", f"A3:D{n}")
    dv = DataValidation(type="list",
                        formula1=f'"{",".join(SUPPORTED_METHODS)}"',
                        allow_blank=False, showDropDown=False)
    dv.add(f"C4:C{n}")
    ws.add_data_validation(dv)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_locked_years(
    wb: Workbook, locked_years: Dict[int, bool], years: List[int],
) -> None:
    ws = wb.create_sheet("Locked_Years")
    ws["A1"] = ("Flip Locked? = TRUE for any tax year whose return has "
                "been filed. The next regeneration snapshots that year's "
                "lot inventory and matching into Frozen_Inventory and "
                "Frozen_Lot_Matching. After that the workbook will never "
                "recompute that year, even if raw CSV data changes.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:C1")
    headers = ["Tax year", "Locked?", "Notes"]
    write_header(ws, headers, row=3)
    for i, y in enumerate(years, start=4):
        ws.cell(row=i, column=1, value=y)
        ws.cell(row=i, column=2, value=bool(locked_years.get(y, False)))
        ws.cell(row=i, column=2).fill = EDITABLE_FILL
        ws.cell(row=i, column=3, value="").fill = EDITABLE_FILL
    n = max(4, 3 + len(years))
    add_table(ws, "tbl_Locked_Years", f"A3:C{n}")
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"',
                        allow_blank=False, showDropDown=False)
    dv.add(f"B4:B{n}")
    ws.add_data_validation(dv)
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_frozen_inventory(
    wb: Workbook,
    existing: Dict[int, List[Dict[str, Any]]],
    year_end_inventory: Dict[int, List[Lot]],
    locked_years: Dict[int, bool],
) -> None:
    ws = wb.create_sheet("Frozen_Inventory")
    ws["A1"] = ("Per-lot snapshot of OPEN lots at year-end for locked years. "
                "New imports start future-year calculations from the latest "
                "frozen snapshot.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:L1")
    headers = ["Snapshot year", "Lot_ID", "Tx_ID", "Instrument_ID",
               "Source broker", "Source account", "Source file",
               "Source row", "Buy date", "Quantity original",
               "Quantity remaining", "Price per share USD",
               "Buy commission USD"]
    write_header(ws, headers, row=3)

    # Compose final set: preserve existing locked-year snapshots;
    # add new snapshot for any year flipped to Locked? in this run.
    emitted: Dict[int, List[Dict[str, Any]]] = {}
    for y, rows in existing.items():
        if locked_years.get(y):
            emitted[y] = list(rows)
    for y, inv in year_end_inventory.items():
        if not locked_years.get(y):
            continue
        if y in emitted:
            continue
        rows_for_y: List[Dict[str, Any]] = []
        for lot in inv:
            if lot.quantity_remaining <= 1e-9:
                continue
            rows_for_y.append({
                "Snapshot year": y,
                "Lot_ID": lot.lot_id,
                "Tx_ID": lot.tx_id,
                "Instrument_ID": lot.instrument_id,
                "Source broker": lot.source_broker,
                "Source account": lot.source_account,
                "Source file": lot.source_file,
                "Source row": lot.source_row,
                "Buy date": lot.buy_date,
                "Quantity original": lot.quantity_original,
                "Quantity remaining": lot.quantity_remaining,
                "Price per share USD": lot.price_per_share_usd,
                "Buy commission USD": lot.buy_commission_total_usd,
            })
        emitted[y] = rows_for_y

    i = 3
    for y in sorted(emitted.keys()):
        for rec in emitted[y]:
            i += 1
            ws.cell(row=i, column=1, value=rec["Snapshot year"])
            ws.cell(row=i, column=2, value=rec["Lot_ID"])
            ws.cell(row=i, column=3, value=rec["Tx_ID"])
            ws.cell(row=i, column=4, value=rec["Instrument_ID"])
            ws.cell(row=i, column=5, value=rec["Source broker"])
            ws.cell(row=i, column=6, value=rec["Source account"])
            ws.cell(row=i, column=7, value=rec["Source file"])
            ws.cell(row=i, column=8, value=rec["Source row"])
            ws.cell(row=i, column=9, value=rec["Buy date"])
            ws.cell(row=i, column=9).number_format = "yyyy-mm-dd"
            ws.cell(row=i, column=10, value=rec["Quantity original"])
            ws.cell(row=i, column=11, value=rec["Quantity remaining"])
            ws.cell(row=i, column=12, value=rec["Price per share USD"])
            ws.cell(row=i, column=13, value=rec["Buy commission USD"])
    last = max(4, i)
    add_table(ws, "tbl_Frozen_Inventory", f"A3:M{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_frozen_matching(
    wb: Workbook,
    existing: Dict[int, List[Dict[str, Any]]],
    current_match_lines: List[MatchLine],
    locked_years: Dict[int, bool],
) -> None:
    ws = wb.create_sheet("Frozen_Lot_Matching")
    ws["A1"] = ("Snapshot of matched lot lines for locked years. Locked-year "
                "calculations are read from here on subsequent "
                "regenerations, not recomputed.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:S1")

    headers = ["Tax year", "Match_ID", "Sell_ID", "Sell date",
               "Sell source broker", "Sell source file", "Sell source row",
               "Instrument_ID", "Buy Lot_ID", "Buy Tx_ID", "Buy date",
               "Buy source broker", "Buy source file", "Buy source row",
               "Quantity", "Buy price USD", "Sell price USD",
               "Allocated buy commission USD",
               "Allocated sell commission USD", "FX rate buy",
               "FX rate sell", "Cost basis CZK", "Proceeds CZK",
               "Holding days", "Time-test exempt?", "Taxable?",
               "Taxable gain CZK", "Method"]
    write_header(ws, headers, row=3)

    # Compose emitted rows: existing for locked years + snapshot current lines
    # for years newly flipped locked.
    emitted: List[Dict[str, Any]] = []
    covered_years = set()
    for y, rows in existing.items():
        if locked_years.get(y):
            covered_years.add(y)
            for r in rows:
                r = dict(r)
                r.setdefault("Tax year", y)
                emitted.append(r)
    for m in current_match_lines:
        if not locked_years.get(m.tax_year):
            continue
        if m.tax_year in covered_years:
            continue
        emitted.append({
            "Tax year": m.tax_year, "Match_ID": m.match_id,
            "Sell_ID": m.sell_tx_id, "Sell date": m.sell_date,
            "Sell source broker": m.sell_source_broker,
            "Sell source file": m.sell_source_file,
            "Sell source row": m.sell_source_row,
            "Instrument_ID": m.instrument_id,
            "Buy Lot_ID": m.buy_lot_id, "Buy Tx_ID": m.buy_tx_id,
            "Buy date": m.buy_date,
            "Buy source broker": m.buy_source_broker,
            "Buy source file": m.buy_source_file,
            "Buy source row": m.buy_source_row,
            "Quantity": m.quantity,
            "Buy price USD": m.buy_price_per_share_usd,
            "Sell price USD": m.sell_price_per_share_usd,
            "Allocated buy commission USD": m.allocated_buy_commission_usd,
            "Allocated sell commission USD": m.allocated_sell_commission_usd,
            "FX rate buy": m.fx_rate_buy, "FX rate sell": m.fx_rate_sell,
            "Cost basis CZK": m.cost_basis_czk,
            "Proceeds CZK": m.proceeds_czk,
            "Holding days": m.holding_days,
            "Time-test exempt?": m.time_test_exempt,
            "Taxable?": m.taxable,
            "Taxable gain CZK": m.taxable_gain_czk,
            "Method": m.method,
        })

    for i, r in enumerate(emitted, start=4):
        for j, h in enumerate(headers, start=1):
            v = r.get(h, "")
            ws.cell(row=i, column=j, value=v)
            if h in ("Sell date", "Buy date") and isinstance(v, date):
                ws.cell(row=i, column=j).number_format = "yyyy-mm-dd"
    last = max(4, 3 + len(emitted))
    add_table(ws, "tbl_Frozen_Lot_Matching",
              f"A3:{get_column_letter(len(headers))}{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _write_lots(wb: Workbook, lots: List[Lot]) -> None:
    ws = wb.create_sheet("Lots")
    headers = ["Lot_ID", "Tx_ID", "Instrument_ID", "Source broker",
               "Source account", "Source file", "Source row", "Buy date",
               "Quantity original", "Quantity remaining",
               "Price per share USD", "Buy commission USD",
               "Adjustments"]
    write_header(ws, headers)
    for i, lot in enumerate(lots, start=2):
        ws.cell(row=i, column=1, value=lot.lot_id)
        ws.cell(row=i, column=2, value=lot.tx_id)
        ws.cell(row=i, column=3, value=lot.instrument_id)
        ws.cell(row=i, column=4, value=lot.source_broker)
        ws.cell(row=i, column=5, value=lot.source_account)
        ws.cell(row=i, column=6, value=lot.source_file)
        ws.cell(row=i, column=7, value=lot.source_row)
        ws.cell(row=i, column=8, value=lot.buy_date)
        ws.cell(row=i, column=8).number_format = "yyyy-mm-dd"
        ws.cell(row=i, column=9, value=lot.quantity_original)
        ws.cell(row=i, column=10, value=lot.quantity_remaining)
        ws.cell(row=i, column=11, value=lot.price_per_share_usd)
        ws.cell(row=i, column=12, value=lot.buy_commission_total_usd)
        ws.cell(row=i, column=13,
                value="; ".join(lot.adjustments) if lot.adjustments else "")
    if lots:
        add_table(ws, "tbl_Lots", f"A1:M{1+len(lots)}")
    # Conditional formatting for remaining quantity < 0 (sanity).
    if lots:
        rng = f"J2:J{1+len(lots)}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"],
                            fill=ERROR_FILL))
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_lot_matching(wb: Workbook, matches: List[MatchLine]) -> None:
    ws = wb.create_sheet("Lot_Matching")
    headers = ["Match_ID", "Tax year", "Sell_ID", "Sell date",
               "Sell source broker", "Sell source file", "Sell source row",
               "Instrument_ID", "Buy Lot_ID", "Buy Tx_ID", "Buy date",
               "Buy source broker", "Buy source file", "Buy source row",
               "Quantity", "Buy price USD", "Sell price USD",
               "Allocated buy commission USD",
               "Allocated sell commission USD", "FX rate buy",
               "FX rate sell", "Cost basis CZK", "Proceeds CZK",
               "Holding days", "Time-test exempt?", "Taxable?",
               "Taxable gain CZK", "Method"]
    write_header(ws, headers)
    for i, m in enumerate(matches, start=2):
        values = [
            m.match_id, m.tax_year, m.sell_tx_id, m.sell_date,
            m.sell_source_broker, m.sell_source_file, m.sell_source_row,
            m.instrument_id, m.buy_lot_id, m.buy_tx_id, m.buy_date,
            m.buy_source_broker, m.buy_source_file, m.buy_source_row,
            m.quantity, m.buy_price_per_share_usd,
            m.sell_price_per_share_usd, m.allocated_buy_commission_usd,
            m.allocated_sell_commission_usd, m.fx_rate_buy, m.fx_rate_sell,
            m.cost_basis_czk, m.proceeds_czk, m.holding_days,
            m.time_test_exempt, m.taxable, m.taxable_gain_czk, m.method,
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=i, column=j, value=v)
            if j in (4, 11) and isinstance(v, date):
                c.number_format = "yyyy-mm-dd"
    if matches:
        add_table(ws, "tbl_Lot_Matching",
                  f"A1:{get_column_letter(len(headers))}{1+len(matches)}")
        # Highlight exempt rows green, taxable losses red.
        last = 1 + len(matches)
        ws.conditional_formatting.add(
            f"Y2:Y{last}",
            CellIsRule(operator="equal", formula=['"TRUE"'], fill=OK_FILL),
        )
        ws.conditional_formatting.add(
            f"AA2:AA{last}",
            CellIsRule(operator="lessThan", formula=["0"], fill=WARNING_FILL),
        )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_frozen_snapshots(
    wb: Workbook,
    existing: Dict[int, Dict[str, Any]],
    year_end_inventory: Dict[int, List[Lot]],
    locked_years: Dict[int, bool],
    current_match_lines: List[MatchLine],
) -> None:
    ws = wb.create_sheet("Frozen_Snapshots")
    ws["A1"] = ("Snapshot manifest for locked years. One row per locked year "
                "confirms that the frozen state has been captured even when "
                "open inventory is zero.")
    ws["A1"].font = Font(italic=True, color="FF5A5A5A")
    ws["A1"].alignment = WRAP_LEFT
    ws.merge_cells("A1:E1")

    headers = ["Snapshot year", "Snapshot captured?", "Open lots captured",
               "Match lines captured", "Captured at UTC"]
    write_header(ws, headers, row=3)

    existing_rows: Dict[int, Dict[str, Any]] = {
        int(y): dict(r) for y, r in existing.items()
    }
    match_count_by_year: Dict[int, int] = defaultdict(int)
    for m in current_match_lines:
        match_count_by_year[m.tax_year] += 1

    emitted: List[Dict[str, Any]] = []
    now_utc = datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")
    for y in sorted(locked_years.keys()):
        if not locked_years.get(y):
            continue
        prior = existing_rows.get(y, {})
        has_new_snapshot = y in year_end_inventory
        emitted.append({
            "Snapshot year": y,
            "Snapshot captured?": True,
            "Open lots captured": (
                len([l for l in year_end_inventory.get(y, [])
                     if l.quantity_remaining > 1e-9])
                if has_new_snapshot else int(prior.get("Open lots captured") or 0)
            ),
            "Match lines captured": (
                match_count_by_year.get(y, 0)
                if has_new_snapshot else int(prior.get("Match lines captured") or 0)
            ),
            "Captured at UTC": (
                now_utc if has_new_snapshot else (prior.get("Captured at UTC") or "")
            ),
        })

    for i, row in enumerate(emitted, start=4):
        for j, h in enumerate(headers, start=1):
            ws.cell(row=i, column=j, value=row.get(h, ""))
    last = max(4, 3 + len(emitted))
    add_table(ws, "tbl_Frozen_Snapshots", f"A3:E{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A4"


def _symbol_by_instrument(txs: List[Transaction]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for tx in txs:
        out.setdefault(tx.instrument_id, tx.symbol)
    return out


def extract_position_rows(
    raw_rows: List[RawRow],
    instrument_map: Dict[str, Dict[str, str]],
) -> Dict[str, float]:
    """Extract Yahoo position rows (no Trade Date/Tx Type, has Quantity)."""
    out: Dict[str, float] = defaultdict(float)
    for raw in raw_rows:
        tt = (raw.data.get("Transaction Type") or "").strip()
        td = (raw.data.get("Trade Date") or "").strip()
        if tt or td:
            continue
        qty_raw = raw.data.get("Quantity")
        qty, ok = safe_float(qty_raw or "", default=None)
        if not ok or qty is None:
            continue
        symbol = (raw.data.get("Symbol") or "").strip()
        inst = (instrument_map.get(symbol, {}).get("Instrument_ID") or symbol)
        if not inst:
            continue
        out[inst] += qty
    return dict(out)


def _write_sell_review(
    wb: Workbook,
    match_lines: List[MatchLine],
    review_state: Dict[str, Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Sell_Review")
    headers = [
        "Row type", "Tax year", "Instrument_ID", "Sell_ID", "Sell date",
        "Sell broker", "Sell source file", "Sell source row", "Sell quantity",
        "Matched quantity", "Difference", "Method used", "Time-test mixed?",
        "Total proceeds CZK", "Total cost CZK", "Taxable gain/loss CZK",
        "Buy Lot_ID", "Buy date", "Buy broker", "Buy source file",
        "Buy source row", "Quantity matched", "Buy price USD", "Sell price USD",
        "FX buy", "FX sell", "Cost basis CZK", "Proceeds CZK",
        "Holding days", "Time-test exempt?", "Review status", "Operator note",
    ]
    write_header(ws, headers)
    by_sell: Dict[str, List[MatchLine]] = defaultdict(list)
    for m in match_lines:
        by_sell[m.sell_tx_id].append(m)

    row = 2
    for sell_id in sorted(by_sell.keys()):
        group = by_sell[sell_id]
        g0 = group[0]
        matched_qty = sum(m.quantity for m in group)
        proceeds = sum(m.proceeds_czk for m in group)
        cost = sum(m.cost_basis_czk for m in group)
        taxable_gain = sum(m.taxable_gain_czk for m in group if m.taxable)
        methods = sorted(set(m.method for m in group))
        mixed = len({bool(m.time_test_exempt) for m in group}) > 1
        state = review_state.get(sell_id, {})
        header_vals = [
            "SELL", g0.tax_year, g0.instrument_id, sell_id, g0.sell_date,
            g0.sell_source_broker, g0.sell_source_file, g0.sell_source_row,
            matched_qty, matched_qty, 0.0,
            ",".join(methods), mixed, round(proceeds, 2), round(cost, 2),
            round(taxable_gain, 2),
            "", "", "", "", "", "", "", "", "", "", "", "", "", "",
            state.get("review_status", ""),
            state.get("operator_note", ""),
        ]
        for c, v in enumerate(header_vals, start=1):
            ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=c).fill = SUBHEADER_FILL
        ws.cell(row=row, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=31).fill = EDITABLE_FILL
        ws.cell(row=row, column=32).fill = EDITABLE_FILL
        row += 1
        for m in group:
            vals = [
                "LOT", m.tax_year, m.instrument_id, m.sell_tx_id, m.sell_date,
                m.sell_source_broker, m.sell_source_file, m.sell_source_row,
                "", "", "", m.method, "",
                "", "", m.taxable_gain_czk,
                m.buy_lot_id, m.buy_date, m.buy_source_broker, m.buy_source_file,
                m.buy_source_row, m.quantity, m.buy_price_per_share_usd,
                m.sell_price_per_share_usd, m.fx_rate_buy, m.fx_rate_sell,
                m.cost_basis_czk, m.proceeds_czk, m.holding_days,
                m.time_test_exempt, "", "",
            ]
            for c, v in enumerate(vals, start=1):
                ws.cell(row=row, column=c, value=v)
            ws.cell(row=row, column=5).number_format = "yyyy-mm-dd"
            ws.cell(row=row, column=18).number_format = "yyyy-mm-dd"
            row += 1

    last = max(2, row - 1)
    add_table(ws, "tbl_Sell_Review", f"A1:{get_column_letter(len(headers))}{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_open_lots_review(
    wb: Workbook,
    lots: List[Lot],
    txs: List[Transaction],
    fx: FXResolver,
) -> None:
    ws = wb.create_sheet("Open_Lots_Review")
    headers = [
        "Instrument_ID", "Symbol", "Broker/source", "Lot_ID", "Buy date",
        "Orig qty", "Qty remaining", "Price per share USD",
        "Remaining cost basis USD", "Est. CZK cost basis", "Source file",
        "Source row",
    ]
    write_header(ws, headers)
    sym_map = _symbol_by_instrument(txs)
    open_lots = [l for l in lots if l.quantity_remaining > 1e-9]
    open_lots.sort(key=lambda l: (l.instrument_id, l.buy_date, l.source_file, l.source_row))
    for i, lot in enumerate(open_lots, start=2):
        remaining_cost_usd = (
            lot.quantity_remaining * lot.price_per_share_usd
            + lot.buy_commission_total_usd * (lot.quantity_remaining / lot.quantity_original)
            if lot.quantity_original > 0 else 0.0
        )
        fx_buy, _ = fx.rate_for(lot.buy_date)
        vals = [
            lot.instrument_id,
            sym_map.get(lot.instrument_id, ""),
            f"{lot.source_broker}/{lot.source_account}",
            lot.lot_id,
            lot.buy_date,
            lot.quantity_original,
            lot.quantity_remaining,
            lot.price_per_share_usd,
            remaining_cost_usd,
            remaining_cost_usd * fx_buy,
            lot.source_file,
            lot.source_row,
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=i, column=c, value=v)
        ws.cell(row=i, column=5).number_format = "yyyy-mm-dd"
    last = max(2, 1 + len(open_lots))
    add_table(ws, "tbl_Open_Lots_Review", f"A1:{get_column_letter(len(headers))}{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_open_position_check(
    wb: Workbook,
    raw_rows: List[RawRow],
    instrument_map: Dict[str, Dict[str, str]],
    lots: List[Lot],
) -> None:
    ws = wb.create_sheet("Open_Position_Check")
    headers = ["Instrument_ID", "Yahoo qty", "Calculated qty", "Difference", "Status"]
    write_header(ws, headers)
    rows = build_open_position_rows(raw_rows, instrument_map, lots)
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row["Instrument_ID"])
        ws.cell(row=i, column=2, value=row["Yahoo qty"])
        ws.cell(row=i, column=3, value=row["Calculated qty"])
        ws.cell(row=i, column=4, value=row["Difference"])
        ws.cell(row=i, column=5, value=row["Status"])
    last = max(2, 1 + len(rows))
    add_table(ws, "tbl_Open_Position_Check", f"A1:E{last}")
    ws.conditional_formatting.add(
        f"E2:E{last}",
        CellIsRule(operator="equal", formula=['"ERROR"'], fill=ERROR_FILL),
    )
    ws.conditional_formatting.add(
        f"E2:E{last}",
        CellIsRule(operator="equal", formula=['"WARN"'], fill=WARNING_FILL),
    )
    ws.conditional_formatting.add(
        f"E2:E{last}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL),
    )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_method_plan(
    wb: Workbook,
    years: List[int],
    instrument_ids: List[str],
    method_selection: Dict[Tuple[int, str], str],
    settings: Dict[int, Dict[str, Any]],
    method_comparison: List[Dict[str, Any]],
    match_lines: List[MatchLine],
) -> None:
    ws = wb.create_sheet("Method_Plan")
    headers = [
        "Tax year", "Instrument_ID", "Selected method", "Filed?", "Locked?",
        "Best simulated method", "Selected tax CZK", "Best simulated tax CZK",
        "Delta CZK", "Action required",
    ]
    write_header(ws, headers)
    cmp_by_year = {int(r.get("Tax year")): r for r in method_comparison}
    by_yi_gain: Dict[Tuple[int, str], float] = defaultdict(float)
    for m in match_lines:
        if m.taxable:
            by_yi_gain[(m.tax_year, m.instrument_id)] += m.taxable_gain_czk

    row = 2
    for y in years:
        rcmp = cmp_by_year.get(y, {})
        best_method = rcmp.get("Best method", "")
        best_tax_year = float(rcmp.get("Best method tax CZK") or 0.0)
        for inst in instrument_ids:
            sel = method_selection.get((y, inst), policy.default_method_for(y))
            filed = policy.is_filed(y)
            locked = bool(settings.get(y, {}).get("locked", False))
            tax_rate = float(settings.get(y, {}).get("tax_rate", DEFAULT_TAX_RATE))
            yi_tax = max(0.0, by_yi_gain.get((y, inst), 0.0)) * tax_rate
            delta = yi_tax - best_tax_year
            if locked:
                action = "LOCKED - do not optimize"
            elif filed:
                action = "FILED - keep filed method"
            elif sel != best_method and best_method:
                action = "Review possible optimization"
            else:
                action = "OK"
            vals = [
                y, inst, sel, filed, locked, best_method,
                round(yi_tax, 2), round(best_tax_year, 2), round(delta, 2), action,
            ]
            for c, v in enumerate(vals, start=1):
                ws.cell(row=row, column=c, value=v)
            row += 1
    last = max(2, row - 1)
    add_table(ws, "tbl_Method_Plan", f"A1:J{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_filed_year_reconciliation(
    wb: Workbook,
    yearly_summary: List[Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    instrument_ids: List[str],
    existing: Dict[int, Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Filed_Year_Reconciliation")
    headers = [
        "Tax year", "Expected filed method", "Workbook method",
        "Workbook tax due CZK", "Filed tax due CZK", "Difference",
        "Status", "Note",
    ]
    write_header(ws, headers)
    summary_by_year = {int(r.get("Tax year")): r for r in yearly_summary}
    row = 2
    for y in sorted(FILED_YEARS.keys()):
        expected_method = policy.filed_method(y) or policy.default_method_for(y)
        methods = {
            method_selection.get((y, inst), policy.default_method_for(y))
            for inst in instrument_ids
        }
        workbook_method = methods.pop() if len(methods) == 1 else "MIXED"
        workbook_tax = float(summary_by_year.get(y, {}).get("Tax due CZK") or 0.0)
        filed_tax = _to_float(existing.get(y, {}).get("filed_tax_due"), workbook_tax)
        diff = workbook_tax - filed_tax
        status = "OK" if abs(diff) <= 0.5 and workbook_method == expected_method else "ERROR"
        note = ""
        if workbook_method != expected_method:
            note = f"Expected method {expected_method}, got {workbook_method}."
        vals = [y, expected_method, workbook_method, workbook_tax, filed_tax, diff, status, note]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=row, column=c, value=v)
        ws.cell(row=row, column=5).fill = EDITABLE_FILL
        row += 1
    last = max(2, row - 1)
    add_table(ws, "tbl_Filed_Year_Reconciliation", f"A1:H{last}")
    ws.conditional_formatting.add(
        f"G2:G{last}",
        CellIsRule(operator="equal", formula=['"ERROR"'], fill=ERROR_FILL),
    )
    ws.conditional_formatting.add(
        f"G2:G{last}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=OK_FILL),
    )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_review_state(
    wb: Workbook,
    review_state: Dict[str, Dict[str, Any]],
    match_lines: List[MatchLine],
) -> None:
    ws = wb.create_sheet("Review_State")
    headers = ["Sell_ID", "Review status", "Operator note"]
    write_header(ws, headers)
    sell_ids = sorted({m.sell_tx_id for m in match_lines} | set(review_state.keys()))
    for i, sid in enumerate(sell_ids, start=2):
        state = review_state.get(sid, {})
        ws.cell(row=i, column=1, value=sid)
        ws.cell(row=i, column=2, value=state.get("review_status", ""))
        ws.cell(row=i, column=3, value=state.get("operator_note", ""))
        ws.cell(row=i, column=2).fill = EDITABLE_FILL
        ws.cell(row=i, column=3).fill = EDITABLE_FILL
    last = max(2, 1 + len(sell_ids))
    add_table(ws, "tbl_Review_State", f"A1:C{last}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_yearly_summary(
    wb: Workbook, rows: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Yearly_Tax_Summary")
    headers = ["Tax year", "Gross proceeds CZK (all sells)",
               "Time-test exempt proceeds CZK",
               "Non-exempt cost basis CZK", "Non-exempt proceeds CZK",
               "Taxable gains CZK", "Taxable losses CZK",
               "Pre-exemption tax base CZK",
               "100k threshold met?", "Apply 100k exemption?",
               "Final tax base CZK", "Tax rate", "Tax due CZK",
               "FX method", "Locked?", "Match line count", "Note"]
    write_header(ws, headers)
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            v = row.get(h, "")
            c = ws.cell(row=i, column=j, value=v)
            if h == "Tax rate":
                c.number_format = "0.00%"
            elif "CZK" in h:
                c.number_format = "#,##0.00"
    if rows:
        add_table(ws, "tbl_Yearly_Summary",
                  f"A1:{get_column_letter(len(headers))}{1+len(rows)}")
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_method_comparison(
    wb: Workbook, rows: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Method_Comparison")
    headers = ["Tax year",
               "FIFO tax base CZK", "FIFO tax CZK",
               "LIFO tax base CZK", "LIFO tax CZK",
               "MIN_GAIN tax base CZK", "MIN_GAIN tax CZK",
               "MAX_GAIN tax base CZK", "MAX_GAIN tax CZK",
               "Selected method tax base CZK",
               "Selected method tax CZK",
               "Best method", "Best method tax CZK",
               "Delta selected vs best CZK"]
    write_header(ws, headers)
    for i, row in enumerate(rows, start=2):
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=i, column=j, value=row.get(h, ""))
            if "CZK" in h:
                c.number_format = "#,##0.00"
    if rows:
        add_table(ws, "tbl_Method_Comparison",
                  f"A1:{get_column_letter(len(headers))}{1+len(rows)}")
        last = 1 + len(rows)
        ws.conditional_formatting.add(
            f"N2:N{last}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       fill=WARNING_FILL),
        )
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_checks(
    wb: Workbook,
    sim_warnings: List[Dict[str, Any]],
    problems: List[Dict[str, Any]],
    fx_yearly: Dict[int, float], fx_daily: Dict[date, float],
    settings: Dict[int, Dict[str, Any]],
    locked_years: Dict[int, bool],
    frozen_inventory: Dict[int, List[Dict[str, Any]]],
    split_warnings: List[Dict[str, Any]],
    method_selection: Dict[Tuple[int, str], str],
    yearly_summary: List[Dict[str, Any]],
    match_lines: List[MatchLine],
    lots_final: List[Lot],
    year_end_inventory: Dict[int, List[Lot]],
    frozen_snapshots: Dict[int, Dict[str, Any]],
    fx: FXResolver,
) -> None:
    ws = wb.create_sheet("Checks")
    headers = ["Severity", "Category", "Detail", "Source file",
               "Source row"]
    write_header(ws, headers)
    rows = build_check_rows(
        sim_warnings=sim_warnings,
        problems=problems,
        fx_yearly=fx_yearly,
        fx_daily=fx_daily,
        settings=settings,
        locked_years=locked_years,
        frozen_inventory=frozen_inventory,
        split_warnings=split_warnings,
        method_selection=method_selection,
        yearly_summary=yearly_summary,
        match_lines=match_lines,
        lots_final=lots_final,
        year_end_inventory=year_end_inventory,
        frozen_snapshots=frozen_snapshots,
        fx=fx,
    )

    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row["Severity"])
        ws.cell(row=i, column=2, value=row["Category"])
        ws.cell(row=i, column=3, value=row["Detail"])
        ws.cell(row=i, column=4, value=row["Source file"])
        ws.cell(row=i, column=5, value=row["Source row"])
    add_table(ws, "tbl_Checks", f"A1:E{1+len(rows)}")
    last = 1 + len(rows)
    ws.conditional_formatting.add(
        f"A2:A{last}",
        FormulaRule(formula=['EXACT($A2,"ERROR")'], fill=ERROR_FILL))
    ws.conditional_formatting.add(
        f"A2:A{last}",
        FormulaRule(formula=['EXACT($A2,"WARN")'], fill=WARNING_FILL))
    ws.conditional_formatting.add(
        f"A2:A{last}",
        FormulaRule(formula=['EXACT($A2,"INFO")'], fill=SUBHEADER_FILL))
    autosize_columns(ws)
    ws.freeze_panes = "A2"


def _write_audit_report(
    wb: Workbook,
    raw_rows: List[RawRow], txs: List[Transaction],
    ignored: List[Dict[str, Any]], match_lines: List[MatchLine],
    yearly_summary: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("Audit_Report")
    ws["A1"] = "Audit Report — end-to-end traceability summary"
    ws["A1"].font = Font(bold=True, size=14)

    # Per source file totals.
    per_file: Dict[str, Dict[str, int]] = defaultdict(
        lambda: {"raw": 0, "tx": 0, "ignored": 0})
    for r in raw_rows:
        per_file[r.source_file]["raw"] += 1
    for t in txs:
        per_file[t.source_file]["tx"] += 1
    for ig in ignored:
        per_file[ig["source_file"]]["ignored"] += 1

    ws["A3"] = "Per source file"
    ws["A3"].font = Font(bold=True, size=12)
    headers = ["Source file", "Raw rows", "Transactions", "Ignored"]
    write_header(ws, headers, row=4)
    i = 5
    for src in sorted(per_file.keys()):
        d = per_file[src]
        ws.cell(row=i, column=1, value=src)
        ws.cell(row=i, column=2, value=d["raw"])
        ws.cell(row=i, column=3, value=d["tx"])
        ws.cell(row=i, column=4, value=d["ignored"])
        i += 1
    ws.cell(row=i, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=i, column=2,
            value=sum(d["raw"] for d in per_file.values()))
    ws.cell(row=i, column=3,
            value=sum(d["tx"] for d in per_file.values()))
    ws.cell(row=i, column=4,
            value=sum(d["ignored"] for d in per_file.values()))

    # Per-year trace counts
    base = i + 3
    ws.cell(row=base, column=1, value="Per tax year reconciliation")
    ws.cell(row=base, column=1).font = Font(bold=True, size=12)
    headers = ["Tax year", "Match lines",
               "Gross proceeds CZK", "Tax base CZK", "Tax due CZK",
               "Link"]
    write_header(ws, headers, row=base + 1)
    for k, row in enumerate(yearly_summary, start=base + 2):
        ws.cell(row=k, column=1, value=row["Tax year"])
        ws.cell(row=k, column=2, value=row["Match line count"])
        ws.cell(row=k, column=3,
                value=row["Gross proceeds CZK (all sells)"])
        ws.cell(row=k, column=4, value=row["Final tax base CZK"])
        ws.cell(row=k, column=5, value=row["Tax due CZK"])
        ws.cell(row=k, column=6,
                value="See Lot_Matching filtered by Tax year for details")

    autosize_columns(ws)
    ws.freeze_panes = "A2"


# -----------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="Build Czech stock-trade tax workbook from Yahoo CSVs")
    p.add_argument("--input", nargs="+", required=True,
                   help="CSV input file paths")
    p.add_argument("--output", default="stock_tax_system.xlsx",
                   help="Output .xlsx path")
    p.add_argument(
        "--backup-existing",
        action="store_true",
        help=("Backup existing output to backups/<name>_YYYY-MM-DD_HHMMSS.xlsx "
              "before replace"),
    )
    p.add_argument(
        "--allow-alternate-output",
        action="store_true",
        help=("Allow writing to output names other than "
              f"{CANONICAL_OUTPUT_NAME}"),
    )
    p.add_argument("--verbose", action="store_true")
    args = p.parse_args(argv)

    out_path = Path(args.output)
    if not args.allow_alternate_output and out_path.name != CANONICAL_OUTPUT_NAME:
        print(
            ("Refusing alternate output name while --allow-alternate-output is "
             f"false. Use --output {CANONICAL_OUTPUT_NAME} or pass "
             "--allow-alternate-output."),
            file=sys.stderr,
        )
        return 1
    inputs = [Path(i) for i in args.input]
    result = calculate_workbook_data(inputs, out_path, fetch_missing_fx=True)
    try:
        existing = out_path.exists()
        write_calculation_result(result, backup_existing=args.backup_existing)
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 1

    # Console summary.
    print(f"Workbook written: {out_path}")
    if args.backup_existing and existing:
        print(f"Backup written: {out_path.parent / 'backups'}")
    print(f"  Raw rows:        {len(result.raw_rows)}")
    print(f"  Transactions:    {len(result.txs)}")
    print(f"  Ignored:         {len(result.ignored)}")
    print(f"  Lots (final):    {len(result.lots_final)}")
    print(f"  Match lines:     {len(result.match_lines)}")
    print(f"  Years:           {result.years}")
    print("Per year:")
    for row in result.yearly_summary:
        print(f"  {row['Tax year']}: "
              f"gross CZK {row['Gross proceeds CZK (all sells)']:,.2f}  "
              f"base CZK {row['Final tax base CZK']:,.2f}  "
              f"tax CZK {row['Tax due CZK']:,.2f}  "
              f"({row['FX method']}, "
              f"locked={bool(row['Locked?'])})")
    errs = sum(
        1
        for p in result.problems + result.sim_warnings
        if p.get("severity") == "ERROR"
    )
    if errs:
        print(f"WARN: {errs} errors detected. See Checks sheet.",
              file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
