from __future__ import annotations

import json
import shutil
from datetime import date
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
import pytest

import build_stock_tax_workbook as workbook_module
from stock_tax_app.backend.main import create_app
from stock_tax_app.engine import policy
from stock_tax_app.engine.core import run
from stock_tax_app.state import project_store
from stock_tax_app.state.models import ProjectState, ProjectStateMetadata, SCHEMA_VERSION


ROOT = Path(__file__).resolve().parent


def _items(payload: dict) -> list[dict]:
    return payload["items"]


def _copy_project_fixture(tmp_path: Path) -> Path:
    project = tmp_path / "project"
    project.mkdir()
    shutil.copytree(ROOT / ".csv", project / ".csv")
    return project


def _workbook_path(project: Path) -> Path:
    # Tests pre-seed a workbook at the same path the backend runtime
    # uses as its export target. The runtime default is no longer the
    # legacy `stock_tax_system.xlsx` name.
    return project / "exports" / "stock_tax_export.xlsx"


def _ensure_test_workbook(project: Path) -> Path:
    workbook_path = _workbook_path(project)
    if workbook_path.exists():
        return workbook_path
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=workbook_path,
        project_dir=project,
        fetch_missing_fx=False,
    )
    workbook_module.write_workbook(
        calc.output_path,
        calc.raw_rows,
        calc.txs,
        calc.ignored,
        calc.problems,
        calc.instrument_map,
        calc.fx_yearly,
        calc.fx_daily,
        calc.fx_daily_sources,
        calc.corporate_actions,
        calc.method_selection,
        calc.locked_years,
        calc.settings,
        calc.frozen_inventory,
        calc.frozen_matching,
        calc.frozen_snapshots,
        calc.fx,
        calc.lots_final,
        calc.match_lines,
        calc.sim_warnings,
        calc.yearly_summary,
        calc.method_comparison,
        calc.split_warnings,
        calc.year_end_inventory,
        calc.import_log,
        calc.review_state,
        calc.filed_reconciliation,
        fx_yearly_sources=calc.fx_yearly_sources,
    )
    return workbook_path


def _set_workbook_tax_rate(project: Path, year: int, tax_rate: float) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == year:
            ws.cell(row=row, column=2, value=tax_rate)
            wb.save(workbook_path)
            return
    raise AssertionError(f"Year {year} not found in Settings")


def _set_workbook_method_selection(project: Path, year: int, instrument_id: str, method: str) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Method_Selection"]
    for row in range(4, ws.max_row + 1):
        if (
            ws.cell(row=row, column=1).value == year
            and ws.cell(row=row, column=2).value == instrument_id
        ):
            ws.cell(row=row, column=3, value=method)
            wb.save(workbook_path)
            return
    raise AssertionError(f"Method_Selection row not found for {(year, instrument_id)}")


def _set_workbook_fx_yearly(project: Path, year: int, rate: float, source_note: str = "") -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["FX_Yearly"]
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == year:
            ws.cell(row=row, column=2, value=rate)
            ws.cell(row=row, column=3, value=source_note)
            wb.save(workbook_path)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=year)
    ws.cell(row=row, column=2, value=rate)
    ws.cell(row=row, column=3, value=source_note)
    wb.save(workbook_path)


def _set_workbook_fx_daily(project: Path, day: str, rate: float, source_note: str = "") -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["FX_Daily"]
    for row in range(4, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        current = value.date().isoformat() if hasattr(value, "date") else str(value or "")
        if current == day:
            ws.cell(row=row, column=1, value=day)
            ws.cell(row=row, column=2, value=rate)
            ws.cell(row=row, column=3, value=source_note)
            wb.save(workbook_path)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=day)
    ws.cell(row=row, column=2, value=rate)
    ws.cell(row=row, column=3, value=source_note)
    wb.save(workbook_path)


def _set_workbook_instrument_map(
    project: Path,
    symbol: str,
    *,
    instrument_id: str,
    isin: str = "",
    instrument_name: str = "",
    notes: str = "",
) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Instrument_Map"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == symbol:
            ws.cell(row=row, column=2, value=instrument_id)
            ws.cell(row=row, column=3, value=isin)
            ws.cell(row=row, column=4, value=instrument_name)
            ws.cell(row=row, column=5, value=notes)
            wb.save(workbook_path)
            return
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=symbol)
    ws.cell(row=row, column=2, value=instrument_id)
    ws.cell(row=row, column=3, value=isin)
    ws.cell(row=row, column=4, value=instrument_name)
    ws.cell(row=row, column=5, value=notes)
    wb.save(workbook_path)


def _remove_workbook_instrument_map(project: Path, symbol: str) -> None:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Instrument_Map"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == symbol:
            for col in range(1, 6):
                ws.cell(row=row, column=col, value=None)
            break
    wb.save(workbook_path)


def _set_workbook_corporate_action(
    project: Path,
    *,
    effective_date: str,
    instrument_id: str,
    action_type: str,
    ratio_old: float = 1.0,
    ratio_new: float = 1.0,
    notes: str = "",
    applied: bool = True,
) -> int:
    workbook_path = _ensure_test_workbook(project)
    wb = load_workbook(workbook_path)
    ws = wb["Corporate_Actions"]
    target_row = None
    for row in range(4, ws.max_row + 1):
        if ws.cell(row=row, column=1).value in (None, "") and ws.cell(row=row, column=2).value in (None, ""):
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
    ws.cell(row=target_row, column=1, value=effective_date)
    ws.cell(row=target_row, column=2, value=instrument_id)
    ws.cell(row=target_row, column=3, value=action_type)
    ws.cell(row=target_row, column=4, value=ratio_old)
    ws.cell(row=target_row, column=5, value=ratio_new)
    ws.cell(row=target_row, column=7, value=notes)
    ws.cell(row=target_row, column=8, value=applied)
    ws.cell(row=target_row, column=9, value="applied" if applied else "not applied")
    wb.save(workbook_path)
    return target_row


def _read_workbook_corporate_actions(project: Path) -> list[dict[str, object]]:
    wb = load_workbook(_ensure_test_workbook(project))
    ws = wb["Corporate_Actions"]
    rows: list[dict[str, object]] = []
    for row in range(4, ws.max_row + 1):
        date_value = ws.cell(row=row, column=1).value
        instrument_id = ws.cell(row=row, column=2).value
        action_type = ws.cell(row=row, column=3).value
        if not (date_value or instrument_id or action_type):
            continue
        rows.append(
            {
                "Date": date_value,
                "Instrument_ID": instrument_id,
                "Action type": action_type,
                "Ratio old": ws.cell(row=row, column=4).value,
                "Ratio new": ws.cell(row=row, column=5).value,
                "Notes": ws.cell(row=row, column=7).value,
                "Applied?": ws.cell(row=row, column=8).value,
            }
        )
    return rows


def _find_settings_row(project: Path, year: int) -> int | None:
    wb = load_workbook(_ensure_test_workbook(project))
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == year:
            return row
    return None


def _find_method_selection_row(project: Path, year: int, instrument_id: str) -> int | None:
    wb = load_workbook(_ensure_test_workbook(project))
    ws = wb["Method_Selection"]
    for row in range(4, ws.max_row + 1):
        if (
            ws.cell(row=row, column=1).value == year
            and ws.cell(row=row, column=2).value == instrument_id
        ):
            return row
    return None


def _find_instrument_map_row(project: Path, symbol: str) -> int | None:
    wb = load_workbook(_ensure_test_workbook(project))
    ws = wb["Instrument_Map"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == symbol:
            return row
    return None


def test_project_state_missing_file_load_returns_default_state(tmp_path):
    state = project_store.load_project_state(tmp_path)

    assert state == ProjectState()
    assert not project_store.state_path_for(tmp_path).exists()


def test_project_state_roundtrip(tmp_path):
    state = ProjectState(
        metadata=ProjectStateMetadata(schema_version=SCHEMA_VERSION),
        year_settings={
            2025: {
                "method": "MIN_GAIN",
                "tax_rate": 0.2,
                "fx_method": "FX_UNIFIED_GFR",
                "apply_100k": True,
                "notes": "from state",
            }
        },
        method_selection={2025: {"AAPL": "MAX_GAIN"}},
        fx_yearly={2025: {"currency_pair": "USD/CZK", "rate": 21.84, "source_note": "manual", "manual": True}},
        fx_daily={"2025-01-02": {"currency_pair": "USD/CZK", "rate": 21.0, "source_note": "manual", "manual": True}},
        instrument_map={
            "AAPL": {
                "yahoo_symbol": "AAPL",
                "instrument_id": "AAPL",
                "isin": "US0378331005",
                "instrument_name": "Apple Inc.",
                "notes": "state map",
            }
        },
        corporate_actions=[
            {
                "action_id": "ca-1",
                "action_type": "split",
                "effective_date": "2025-01-01",
                "instrument_id": "AAPL",
                "source_symbol": "AAPL",
                "target_instrument_id": "",
                "target_symbol": "",
                "ratio_numerator": 2.0,
                "ratio_denominator": 1.0,
                "source": "unit-test",
                "note": "roundtrip",
                "enabled": True,
            }
        ],
        locked_years={2024: True},
        frozen_inventory={2024: [{"Lot_ID": "L1"}]},
        frozen_lot_matching={2024: [{"Match_ID": "M1"}]},
        frozen_snapshots={2024: {"Snapshot year": 2024}},
        filed_year_reconciliation={2024: {"filed_tax_due": 123.45}},
    )

    project_store.save_project_state(tmp_path, state)
    reloaded = project_store.load_project_state(tmp_path)

    assert reloaded == state
    payload = json.loads(project_store.state_path_for(tmp_path).read_text(encoding="utf-8"))
    assert payload["metadata"]["schema_version"] == SCHEMA_VERSION
    assert sorted(payload.keys()) == [
        "corporate_actions",
        "filed_year_reconciliation",
        "frozen_inventory",
        "frozen_lot_matching",
        "frozen_snapshots",
        "fx_daily",
        "fx_yearly",
        "instrument_map",
        "locked_years",
        "metadata",
        "method_selection",
        "year_settings",
    ]


def test_project_state_unsupported_version_errors(tmp_path):
    path = project_store.state_path_for(tmp_path)
    path.write_text(
        json.dumps({"metadata": {"schema_version": 999}}, indent=2),
        encoding="utf-8",
    )

    with pytest.raises(project_store.UnsupportedProjectStateVersionError):
        project_store.load_project_state(tmp_path)


def test_project_state_beats_workbook_fallback_for_wired_domains(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_tax_rate(project, 2025, 0.19)
    _set_workbook_method_selection(project, first_sale.year, first_sale.instrument_id, "LIFO")

    state = ProjectState(
        year_settings={
            2025: {
                "method": "MAX_GAIN",
                "tax_rate": 0.2,
                "fx_method": "FX_UNIFIED_GFR",
                "apply_100k": False,
                "notes": "state-owned",
            }
        },
        method_selection={first_sale.year: {first_sale.instrument_id: "FIFO"}},
    )
    project_store.save_project_state(project, state)

    result = run(project_dir=project, write_workbook=False)
    sale_after = next(sale for sale in result.sales.items if sale.id == first_sale.id)
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)

    assert sale_after.method == "FIFO"
    assert year_2025.method == "MAX_GAIN"
    assert year_2025.tax_rate == 0.2


def test_year_method_default_applies_without_known_instruments(tmp_path):
    project = _copy_project_fixture(tmp_path)
    project_store.save_project_state(
        project,
        ProjectState(
            year_settings={
                2030: {
                    "method": "LIFO",
                    "tax_rate": 0.15,
                    "fx_method": "FX_UNIFIED_GFR",
                    "apply_100k": False,
                }
            }
        ),
    )

    result = run(project_dir=project, write_workbook=False)
    year_2030 = next(year for year in result.tax_years.items if year.year == 2030)

    assert year_2030.method == "LIFO"
    assert year_2030.tax_rate == pytest.approx(0.15)


def test_per_instrument_method_override_beats_year_method_default(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    project_store.save_project_state(
        project,
        ProjectState(
            year_settings={
                first_sale.year: {
                    "method": "MIN_GAIN",
                }
            },
            method_selection={first_sale.year: {first_sale.instrument_id: "FIFO"}},
        ),
    )

    result = run(project_dir=project, write_workbook=False)
    sale_after = next(sale for sale in result.sales.items if sale.id == first_sale.id)
    year_row = next(year for year in result.tax_years.items if year.year == first_sale.year)

    assert sale_after.method == "FIFO"
    assert year_row.method == "MIN_GAIN"


def test_legacy_workbook_settings_fallback_retired_p3_2(tmp_path):
    """P3.2: Workbook Settings.Tax rate / FX method / Apply 100k / Notes must NOT
    flow into normal runtime.  Runtime must use policy/generated defaults when
    ProjectState has no entry for those fields."""
    project = _copy_project_fixture(tmp_path)

    _set_workbook_tax_rate(project, 2025, 0.21)

    # No ProjectState written — runtime must ignore workbook Settings.Tax rate.
    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)
    import build_stock_tax_workbook as _wbm
    assert year_2025.tax_rate == pytest.approx(_wbm.DEFAULT_TAX_RATE)
    assert year_2025.settings_source == "generated_default"


def test_legacy_workbook_settings_adoption_still_works_after_p3_2(tmp_path):
    """P3.2: Explicit adoption via adopt_legacy_workbook_state still migrates
    Settings rows into ProjectState.  This is the only supported path."""
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_tax_rate(project, 2025, 0.21)
    _set_workbook_method_selection(project, first_sale.year, first_sale.instrument_id, "MAX_GAIN")

    # Method selection is retired (P3.1) — runtime must use policy/default until adoption.
    result = run(project_dir=project, write_workbook=False)
    sale_after = next(sale for sale in result.sales.items if sale.id == first_sale.id)
    assert sale_after.method == policy.default_method_for(first_sale.year)
    year_first = next(year for year in result.tax_years.items if year.year == first_sale.year)
    assert year_first.method_source != "workbook_fallback"

    legacy_state = workbook_module.load_existing_user_state(_workbook_path(project))
    adopted = project_store.adopt_legacy_workbook_state(project, legacy_state)
    assert project_store.state_path_for(project).exists()
    assert adopted.year_settings[2025]["tax_rate"] == 0.21
    assert adopted.method_selection[first_sale.year][first_sale.instrument_id] == "MAX_GAIN"


def test_project_state_yearly_fx_beats_workbook_fallback(tmp_path):
    project = _copy_project_fixture(tmp_path)
    _set_workbook_fx_yearly(project, 2025, 19.99, "workbook yearly")
    project_store.save_project_state(
        project,
        ProjectState(
            fx_yearly={
                2025: {
                    "currency_pair": "USD/CZK",
                    "rate": 24.44,
                    "source_note": "state yearly",
                    "manual": True,
                }
            }
        ),
    )

    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.fx_years.items if year.year == 2025)

    assert year_2025.unified_rate == 24.44
    assert year_2025.source_label == "state yearly"
    assert year_2025.manual_override is True


def test_project_state_daily_fx_beats_workbook_fallback(tmp_path):
    project = _copy_project_fixture(tmp_path)
    target_day = "2020-02-03"
    _set_workbook_fx_daily(project, target_day, 19.11, "workbook daily")
    project_store.save_project_state(
        project,
        ProjectState(
            fx_daily={
                target_day: {
                    "currency_pair": "USD/CZK",
                    "rate": 24.88,
                    "source_note": "state daily",
                    "manual": True,
                }
            }
        ),
    )

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=_workbook_path(project),
        project_dir=project,
        fetch_missing_fx=False,
    )

    assert calc.fx_daily[date.fromisoformat(target_day)] == 24.88
    assert calc.fx_daily_sources[date.fromisoformat(target_day)] == "state daily"


def test_workbook_fx_fallback_still_works_when_project_state_missing(tmp_path):
    project = _copy_project_fixture(tmp_path)
    _set_workbook_fx_yearly(project, 2025, 20.75, "workbook fallback")

    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.fx_years.items if year.year == 2025)

    assert year_2025.unified_rate == 20.75
    assert year_2025.source_label == "workbook fallback"


def test_explicit_legacy_adoption_migrates_fx_without_overwriting_existing_entries(tmp_path):
    project = _copy_project_fixture(tmp_path)
    _set_workbook_fx_yearly(project, 2025, 20.31, "workbook yearly")
    _set_workbook_fx_daily(project, "2020-02-03", 23.45, "workbook daily")
    project_store.save_project_state(
        project,
        ProjectState(
            fx_yearly={
                2025: {
                    "currency_pair": "USD/CZK",
                    "rate": 26.01,
                    "source_note": "keep state yearly",
                    "manual": True,
                }
            }
        ),
    )

    legacy_state = workbook_module.load_existing_user_state(_workbook_path(project))
    adopted = project_store.adopt_legacy_workbook_state(project, legacy_state)
    reloaded = project_store.load_project_state(project)

    assert adopted.fx_yearly[2025]["rate"] == 26.01
    assert adopted.fx_yearly[2025]["source_note"] == "keep state yearly"
    assert adopted.fx_daily["2020-02-03"]["rate"] == 23.45
    assert reloaded.fx_yearly[2025]["rate"] == 26.01
    assert reloaded.fx_daily["2020-02-03"]["source_note"] == "workbook daily"


def test_project_state_instrument_map_beats_workbook_fallback(tmp_path):
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_WORKBOOK",
        isin="WB000SHOP",
        instrument_name="Workbook SHOP",
        notes="workbook map",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            instrument_map={
                symbol: {
                    "yahoo_symbol": symbol,
                    "instrument_id": "SHOP_STATE",
                    "isin": "STATE000SHOP",
                    "instrument_name": "State SHOP",
                    "notes": "state map",
                }
            }
        ),
    )

    result = run(project_dir=project, write_workbook=False)
    sale = next(sale for sale in result.sales.items if sale.ticker == symbol)

    assert sale.instrument_id == "SHOP_STATE"

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=_workbook_path(project),
        project_dir=project,
        fetch_missing_fx=False,
    )
    assert calc.instrument_map[symbol]["Instrument_ID"] == "SHOP_STATE"
    assert calc.instrument_map[symbol]["ISIN"] == "STATE000SHOP"


def test_p3_3_workbook_instrument_map_no_longer_fallback(tmp_path):
    """P3.3: Workbook-only Instrument_Map must not influence normal runtime.

    When ProjectState has no instrument map entry for a symbol, runtime must use
    generated/default identity (symbol itself as instrument_id), not any value
    stored in the workbook Instrument_Map sheet.
    """
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_WORKBOOK",
        isin="WB000SHOP",
        instrument_name="Workbook SHOP",
        notes="workbook map",
    )

    # No ProjectState instrument map entry — only workbook has a row.
    result = run(project_dir=project, write_workbook=False)
    sale = next(sale for sale in result.sales.items if sale.ticker == symbol)
    # After P3.3: generated default, not workbook value.
    assert sale.instrument_id == symbol
    assert sale.instrument_map_source == "generated_default"

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=_workbook_path(project),
        project_dir=project,
        fetch_missing_fx=False,
    )
    # Merged user_state Instrument_Map must not carry the workbook-only row.
    assert calc.instrument_map[symbol]["Instrument_ID"] == symbol


def test_default_generated_instrument_map_still_works(tmp_path):
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _remove_workbook_instrument_map(project, symbol)

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=_workbook_path(project),
        project_dir=project,
        fetch_missing_fx=False,
    )

    assert calc.instrument_map[symbol]["Yahoo Symbol"] == symbol
    assert calc.instrument_map[symbol]["Instrument_ID"] == symbol
    assert calc.instrument_map[symbol]["ISIN"] == ""


def test_explicit_legacy_adoption_migrates_instrument_map_without_overwriting_by_default(tmp_path):
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    copied_symbol = "AMD"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_WORKBOOK",
        isin="WB000SHOP",
        instrument_name="Workbook SHOP",
        notes="workbook map",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            instrument_map={
                symbol: {
                    "yahoo_symbol": symbol,
                    "instrument_id": "SHOP_STATE",
                    "isin": "STATE000SHOP",
                    "instrument_name": "State SHOP",
                    "notes": "state map",
                }
            }
        ),
    )

    legacy_state = workbook_module.load_existing_user_state(_workbook_path(project))
    adopted = project_store.adopt_legacy_workbook_state(project, legacy_state)
    assert copied_symbol in adopted.instrument_map
    assert adopted.instrument_map[copied_symbol]["instrument_id"] == copied_symbol
    assert adopted.instrument_map[symbol]["instrument_id"] == "SHOP_STATE"
    assert adopted.instrument_map[symbol]["notes"] == "state map"

    overwritten = project_store.adopt_legacy_workbook_state(project, legacy_state, overwrite=True)
    assert overwritten.instrument_map[symbol]["instrument_id"] == "SHOP_WORKBOOK"
    assert overwritten.instrument_map[symbol]["isin"] == "WB000SHOP"


def test_workbook_export_reflects_project_state_for_migrated_domains(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_tax_rate(project, 2025, 0.19)
    _set_workbook_method_selection(project, first_sale.year, first_sale.instrument_id, "LIFO")

    project_store.save_project_state(
        project,
        ProjectState(
            year_settings={
                2025: {
                    "tax_rate": 0.2,
                    "fx_method": "FX_UNIFIED_GFR",
                    "apply_100k": False,
                    "notes": "exported from state",
                }
            },
            method_selection={first_sale.year: {first_sale.instrument_id: "FIFO"}},
        ),
    )

    client = TestClient(create_app(project_dir=project))
    result = client.app.state.runtime.calculate(write_workbook=True)
    assert result.tax_years.items

    settings_row = _find_settings_row(project, 2025)
    method_row = _find_method_selection_row(project, first_sale.year, first_sale.instrument_id)
    assert settings_row is not None
    assert method_row is not None

    wb = load_workbook(_workbook_path(project))
    assert wb["Settings"].cell(row=settings_row, column=2).value == 0.2
    assert wb["Method_Selection"].cell(row=method_row, column=3).value == "FIFO"


def test_workbook_export_reflects_project_state_instrument_map(tmp_path):
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_WORKBOOK",
        isin="WB000SHOP",
        instrument_name="Workbook SHOP",
        notes="workbook map",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            instrument_map={
                symbol: {
                    "yahoo_symbol": symbol,
                    "instrument_id": "SHOP_STATE",
                    "isin": "STATE000SHOP",
                    "instrument_name": "State SHOP",
                    "notes": "state map",
                }
            }
        ),
    )

    client = TestClient(create_app(project_dir=project))
    result = client.app.state.runtime.calculate(write_workbook=True)
    assert result.sales.items

    row = _find_instrument_map_row(project, symbol)
    assert row is not None

    wb = load_workbook(_workbook_path(project))
    ws = wb["Instrument_Map"]
    assert ws.cell(row=row, column=2).value == "SHOP_STATE"
    assert ws.cell(row=row, column=3).value == "STATE000SHOP"
    assert ws.cell(row=row, column=4).value == "State SHOP"
    assert ws.cell(row=row, column=5).value == "state map"


def test_project_state_corporate_actions_beat_workbook_fallback(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    instrument_id = next(
        pos.instrument_id for pos in baseline.open_positions.items if float(pos.calculated_qty) > 0
    )
    _set_workbook_corporate_action(
        project,
        effective_date="2025-01-01",
        instrument_id=instrument_id,
        action_type="SPLIT",
        ratio_old=1.0,
        ratio_new=2.0,
        notes="workbook fallback",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "state-split",
                    "action_type": "split",
                    "effective_date": "2100-01-01",
                    "instrument_id": instrument_id,
                    "ratio_numerator": 3.0,
                    "ratio_denominator": 1.0,
                    "source": "project_state",
                    "note": "state override",
                    "enabled": True,
                }
            ]
        ),
    )

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=_workbook_path(project),
        project_dir=project,
        fetch_missing_fx=False,
    )
    assert calc.corporate_actions
    state_row = next(row for row in calc.corporate_actions if row["Instrument_ID"] == instrument_id)
    assert state_row["Ratio new"] == pytest.approx(3.0)
    assert state_row["Notes"] == "state override"


def test_workbook_corporate_action_fallback_still_works_without_project_state(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    instrument_id = next(
        pos.instrument_id for pos in baseline.open_positions.items if float(pos.calculated_qty) > 0
    )
    _set_workbook_corporate_action(
        project,
        effective_date="2100-01-01",
        instrument_id=instrument_id,
        action_type="SPLIT",
        ratio_old=1.0,
        ratio_new=2.0,
        notes="workbook only",
    )

    calc = workbook_module.calculate_workbook_data(
        inputs=sorted((project / ".csv").glob("*.csv")),
        out_path=_workbook_path(project),
        project_dir=project,
        fetch_missing_fx=False,
    )
    assert any(
        row["Instrument_ID"] == instrument_id
        and row["Ratio new"] == pytest.approx(2.0)
        and row["Notes"] == "workbook only"
        for row in calc.corporate_actions
    )


def test_explicit_legacy_adoption_migrates_corporate_actions_without_overwriting(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    instrument_id = next(
        pos.instrument_id for pos in baseline.open_positions.items if float(pos.calculated_qty) > 0
    )
    _set_workbook_corporate_action(
        project,
        effective_date="2100-01-01",
        instrument_id=instrument_id,
        action_type="SPLIT",
        ratio_old=1.0,
        ratio_new=2.0,
        notes="legacy ca",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "state-keep",
                    "action_type": "split",
                    "effective_date": "2099-01-01",
                    "instrument_id": instrument_id,
                    "ratio_numerator": 5.0,
                    "ratio_denominator": 1.0,
                    "source": "project_state",
                    "note": "keep me",
                    "enabled": True,
                }
            ]
        ),
    )

    legacy_state = workbook_module.load_existing_user_state(_workbook_path(project))
    adopted = project_store.adopt_legacy_workbook_state(project, legacy_state)
    assert any(action.get("action_id") == "state-keep" for action in adopted.corporate_actions)
    assert any(action.get("note") == "legacy ca" for action in adopted.corporate_actions)


def test_workbook_export_reflects_project_state_corporate_actions(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    instrument_id = next(
        pos.instrument_id for pos in baseline.open_positions.items if float(pos.calculated_qty) > 0
    )
    _set_workbook_corporate_action(
        project,
        effective_date="2100-01-01",
        instrument_id=instrument_id,
        action_type="SPLIT",
        ratio_old=1.0,
        ratio_new=2.0,
        notes="workbook stale",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "state-export",
                    "action_type": "split",
                    "effective_date": "2100-01-01",
                    "instrument_id": instrument_id,
                    "ratio_numerator": 4.0,
                    "ratio_denominator": 1.0,
                    "source": "project_state",
                    "note": "state export",
                    "enabled": True,
                }
            ]
        ),
    )

    client = TestClient(create_app(project_dir=project))
    result = client.app.state.runtime.calculate(write_workbook=True)
    assert result.sales.items

    rows = _read_workbook_corporate_actions(project)
    row = next(item for item in rows if item["Instrument_ID"] == instrument_id)
    assert row["Ratio new"] == pytest.approx(4.0)
    assert row["Notes"] == "state export"


def test_valid_split_action_changes_open_inventory_quantities(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    target = next(pos for pos in baseline.open_positions.items if float(pos.calculated_qty) > 0)

    project_store.save_project_state(
        project,
        ProjectState(
            corporate_actions=[
                {
                    "action_id": "split-impact",
                    "action_type": "split",
                    "effective_date": "2100-01-01",
                    "instrument_id": target.instrument_id,
                    "ratio_numerator": 2.0,
                    "ratio_denominator": 1.0,
                    "source": "project_state",
                    "note": "inventory impact",
                    "enabled": True,
                }
            ]
        ),
    )

    after = run(project_dir=project, write_workbook=False)
    updated = next(pos for pos in after.open_positions.items if pos.instrument_id == target.instrument_id)
    assert float(updated.calculated_qty) == pytest.approx(float(target.calculated_qty) * 2.0)


# ---------------------------------------------------------------------------
# P3.1 — Method_Selection workbook fallback retirement
# ---------------------------------------------------------------------------


def test_runtime_ignores_workbook_method_selection_when_project_state_missing(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_method_selection(
        project, first_sale.year, first_sale.instrument_id, "MAX_GAIN"
    )

    assert not project_store.state_path_for(project).exists()

    result = run(project_dir=project, write_workbook=False)
    sale_after = next(sale for sale in result.sales.items if sale.id == first_sale.id)
    year_row = next(year for year in result.tax_years.items if year.year == first_sale.year)

    expected_default = policy.default_method_for(first_sale.year)
    assert sale_after.method == expected_default
    assert year_row.method_source != "workbook_fallback"
    if not policy.is_filed(first_sale.year):
        assert year_row.method_source == "generated_default"
    # Runtime must not silently materialize a project state file.
    assert not project_store.state_path_for(project).exists()


def test_explicit_method_selection_adoption_migrates_workbook_rows(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_method_selection(
        project, first_sale.year, first_sale.instrument_id, "MAX_GAIN"
    )

    summary = workbook_module.adopt_legacy_workbook_method_selection(
        project, _workbook_path(project)
    )
    assert summary["adopted"] >= 1
    assert summary["skipped_conflicts"] == 0
    assert summary["legacy_rows"] >= 1

    state = project_store.load_project_state(project)
    assert (
        state.method_selection[first_sale.year][first_sale.instrument_id] == "MAX_GAIN"
    )

    result = run(project_dir=project, write_workbook=False)
    sale_after = next(sale for sale in result.sales.items if sale.id == first_sale.id)
    year_row = next(year for year in result.tax_years.items if year.year == first_sale.year)
    assert sale_after.method == "MAX_GAIN"
    assert year_row.method_source == "project_state"


def test_project_state_method_beats_workbook_method_selection_conflict(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_method_selection(
        project, first_sale.year, first_sale.instrument_id, "MAX_GAIN"
    )
    project_store.save_project_state(
        project,
        ProjectState(
            method_selection={first_sale.year: {first_sale.instrument_id: "FIFO"}},
        ),
    )

    result = run(project_dir=project, write_workbook=False)
    sale_after = next(sale for sale in result.sales.items if sale.id == first_sale.id)
    year_row = next(year for year in result.tax_years.items if year.year == first_sale.year)
    assert sale_after.method == "FIFO"
    assert year_row.method_source == "project_state"


def test_explicit_method_selection_adoption_overwrite_false_preserves_existing(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_method_selection(
        project, first_sale.year, first_sale.instrument_id, "MAX_GAIN"
    )
    project_store.save_project_state(
        project,
        ProjectState(
            method_selection={first_sale.year: {first_sale.instrument_id: "FIFO"}},
        ),
    )

    summary = workbook_module.adopt_legacy_workbook_method_selection(
        project, _workbook_path(project), overwrite=False
    )
    assert summary["skipped_conflicts"] >= 1
    assert summary["overwritten"] == 0

    state = project_store.load_project_state(project)
    assert (
        state.method_selection[first_sale.year][first_sale.instrument_id] == "FIFO"
    )


def test_explicit_method_selection_adoption_overwrite_true_replaces_conflicts(tmp_path):
    project = _copy_project_fixture(tmp_path)
    baseline = run(project_dir=project, write_workbook=False)
    first_sale = baseline.sales.items[0]

    _set_workbook_method_selection(
        project, first_sale.year, first_sale.instrument_id, "MAX_GAIN"
    )
    project_store.save_project_state(
        project,
        ProjectState(
            method_selection={first_sale.year: {first_sale.instrument_id: "FIFO"}},
        ),
    )

    summary = workbook_module.adopt_legacy_workbook_method_selection(
        project, _workbook_path(project), overwrite=True
    )
    assert summary["overwritten"] >= 1

    state = project_store.load_project_state(project)
    assert (
        state.method_selection[first_sale.year][first_sale.instrument_id] == "MAX_GAIN"
    )


def test_explicit_method_selection_adoption_skips_invalid_methods(tmp_path):
    project = _copy_project_fixture(tmp_path)
    legacy = {
        "Method_Selection": [
            {"Tax year": 2025, "Instrument_ID": "AAPL", "Method": "BOGUS"},
            {"Tax year": "not-a-year", "Instrument_ID": "AAPL", "Method": "FIFO"},
            {"Tax year": 2025, "Instrument_ID": "", "Method": "FIFO"},
            {"Tax year": 2025, "Instrument_ID": "MSFT", "Method": "lifo"},
        ]
    }

    summary = project_store.adopt_legacy_workbook_method_selection(project, legacy)
    assert summary["skipped_invalid"] == 3
    assert summary["adopted"] == 1

    state = project_store.load_project_state(project)
    assert state.method_selection[2025]["MSFT"] == "LIFO"


# ---------------------------------------------------------------------------
# P3.2 — Year-settings workbook fallback retirement tests
# ---------------------------------------------------------------------------

def test_p3_2_runtime_ignores_workbook_tax_rate_when_project_state_missing(tmp_path):
    """A: Runtime must NOT read workbook Settings.Tax rate as truth."""
    project = _copy_project_fixture(tmp_path)
    _set_workbook_tax_rate(project, 2025, 0.19)

    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)

    assert year_2025.tax_rate == pytest.approx(workbook_module.DEFAULT_TAX_RATE)
    assert year_2025.settings_source == "generated_default"


def test_p3_2_runtime_ignores_workbook_fx_method_when_project_state_missing(tmp_path):
    """B: Runtime must NOT read workbook Settings.FX method as truth."""
    project = _copy_project_fixture(tmp_path)
    _ensure_test_workbook(project)

    # Manually write FX_DAILY_CNB into workbook Settings sheet
    from openpyxl import load_workbook as _lw
    wb = _lw(_workbook_path(project))
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 2025:
            ws.cell(row=row, column=3, value="FX_DAILY_CNB")
            break
    wb.save(_workbook_path(project))

    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)

    # Runtime must ignore the workbook FX method and use the policy default.
    assert year_2025.fx_method == workbook_module.DEFAULT_FX_METHOD
    assert year_2025.settings_source == "generated_default"


def test_p3_2_runtime_ignores_workbook_apply_100k_when_project_state_missing(tmp_path):
    """C: Runtime must NOT read workbook Settings.Apply 100k exemption? as truth."""
    project = _copy_project_fixture(tmp_path)
    _ensure_test_workbook(project)

    from openpyxl import load_workbook as _lw
    wb = _lw(_workbook_path(project))
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 2025:
            ws.cell(row=row, column=4, value=True)
            break
    wb.save(_workbook_path(project))

    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)

    assert year_2025.exemption_100k is False
    assert year_2025.settings_source == "generated_default"


def test_p3_2_get_years_does_not_report_workbook_fallback_for_settings_fields(tmp_path):
    """E: GET /api/years must never report workbook_fallback for year settings source."""
    project = _copy_project_fixture(tmp_path)
    _set_workbook_tax_rate(project, 2025, 0.19)

    client = TestClient(create_app(project_dir=project))
    body = client.get("/api/years").json()

    for item in body["items"]:
        assert item["settings_source"] != "workbook_fallback", (
            f"Year {item['year']} reported settings_source='workbook_fallback' after P3.2 retirement"
        )
    # reconciliation_source for filed years may still show workbook_fallback (not retired in P3.2)
    # but settings_source must not contribute workbook_fallback to the sources list.
    settings_sources = {item["settings_source"] for item in body["items"]}
    assert "workbook_fallback" not in settings_sources


def test_p3_2_adopt_legacy_workbook_year_settings_migrates_into_project_state(tmp_path):
    """F: Explicit adoption migrates workbook Settings into ProjectState."""
    project = _copy_project_fixture(tmp_path)
    _set_workbook_tax_rate(project, 2025, 0.21)

    summary = project_store.adopt_legacy_workbook_year_settings(
        project,
        _workbook_path(project),
    )

    assert summary["fields_adopted"] >= 1
    state = project_store.load_project_state(project)
    assert state.year_settings[2025]["tax_rate"] == pytest.approx(0.21)

    # After adoption, runtime uses ProjectState value.
    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)
    assert year_2025.tax_rate == pytest.approx(0.21)
    assert year_2025.settings_source == "project_state"


def test_p3_2_project_state_wins_over_workbook_settings_after_adoption_attempt(tmp_path):
    """G: ProjectState always wins in normal runtime over workbook Settings."""
    project = _copy_project_fixture(tmp_path)
    _set_workbook_tax_rate(project, 2025, 0.19)

    project_store.save_project_state(
        project,
        ProjectState(year_settings={2025: {"tax_rate": 0.20}}),
    )

    result = run(project_dir=project, write_workbook=False)
    year_2025 = next(year for year in result.tax_years.items if year.year == 2025)
    assert year_2025.tax_rate == pytest.approx(0.20)
    assert year_2025.settings_source == "project_state"


def test_p3_2_adopt_year_settings_overwrite_false_fills_only_missing_fields(tmp_path):
    """H: overwrite=False fills missing fields but skips existing ones field-by-field."""
    project = _copy_project_fixture(tmp_path)
    # ProjectState owns tax_rate but not fx_method
    project_store.save_project_state(
        project,
        ProjectState(year_settings={2025: {"tax_rate": 0.20}}),
    )
    # Workbook has both tax_rate (different value) and fx_method
    _set_workbook_tax_rate(project, 2025, 0.19)

    summary = project_store.adopt_legacy_workbook_year_settings(
        project,
        _workbook_path(project),
        overwrite=False,
    )

    state = project_store.load_project_state(project)
    # tax_rate must NOT be overwritten (PS owns it)
    assert state.year_settings[2025]["tax_rate"] == pytest.approx(0.20)
    assert summary["fields_skipped"] >= 1


def test_p3_2_adopt_year_settings_overwrite_true_replaces_existing_fields(tmp_path):
    """I: overwrite=True replaces existing fields from workbook."""
    project = _copy_project_fixture(tmp_path)
    project_store.save_project_state(
        project,
        ProjectState(year_settings={2025: {"tax_rate": 0.20}}),
    )
    _set_workbook_tax_rate(project, 2025, 0.18)

    summary = project_store.adopt_legacy_workbook_year_settings(
        project,
        _workbook_path(project),
        overwrite=True,
    )

    state = project_store.load_project_state(project)
    assert state.year_settings[2025]["tax_rate"] == pytest.approx(0.18)
    assert summary["fields_overwritten"] >= 1


def test_p3_2_adopt_year_settings_skips_invalid_tax_rate(tmp_path):
    """Adoption skips tax_rate values outside [0.0, 1.0] and non-numeric values."""
    project = _copy_project_fixture(tmp_path)
    _ensure_test_workbook(project)

    from openpyxl import load_workbook as _lw
    wb = _lw(_workbook_path(project))
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 2025:
            ws.cell(row=row, column=2, value=99.9)  # invalid: > 1.0
            break
    wb.save(_workbook_path(project))

    summary = project_store.adopt_legacy_workbook_year_settings(
        project,
        _workbook_path(project),
    )

    # tax_rate should NOT have been written
    state = project_store.load_project_state(project)
    assert state.year_settings.get(2025, {}).get("tax_rate") is None
    _ = summary  # summary counts may vary; just verify no tax_rate was stored


def test_p3_2_adopt_year_settings_skips_invalid_fx_method(tmp_path):
    """Adoption skips fx_method values not in SUPPORTED_FX_METHODS."""
    project = _copy_project_fixture(tmp_path)
    _ensure_test_workbook(project)

    from openpyxl import load_workbook as _lw
    wb = _lw(_workbook_path(project))
    ws = wb["Settings"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 2025:
            ws.cell(row=row, column=3, value="BOGUS_METHOD")
            break
    wb.save(_workbook_path(project))

    summary = project_store.adopt_legacy_workbook_year_settings(
        project,
        _workbook_path(project),
    )

    state = project_store.load_project_state(project)
    assert state.year_settings.get(2025, {}).get("fx_method") is None
    _ = summary


# ---------------------------------------------------------------------------
# P3.3 — Instrument_Map workbook fallback retirement
# ---------------------------------------------------------------------------


def test_p3_3_adopt_instrument_map_migrates_entries(tmp_path):
    """B: adopt_legacy_workbook_instrument_map writes workbook rows into ProjectState."""
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_LEGACY",
        isin="WB000SHOP",
        instrument_name="Legacy SHOP",
        notes="legacy note",
    )

    summary = project_store.adopt_legacy_workbook_instrument_map(
        project,
        _workbook_path(project),
    )

    assert summary["legacy_rows"] >= 1
    assert summary["adopted"] >= 1
    assert summary["skipped_conflicts"] == 0

    state = project_store.load_project_state(project)
    entry = state.instrument_map.get(symbol)
    assert entry is not None
    assert entry["instrument_id"] == "SHOP_LEGACY"
    assert entry["isin"] == "WB000SHOP"

    # Runtime now reflects adopted mapping.
    result = run(project_dir=project, write_workbook=False)
    sale = next(sale for sale in result.sales.items if sale.ticker == symbol)
    assert sale.instrument_id == "SHOP_LEGACY"
    assert sale.instrument_map_source == "project_state"


def test_p3_3_adopt_overwrite_false_skips_conflicts(tmp_path):
    """D: adopt with overwrite=False skips entries already in ProjectState."""
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_LEGACY",
        isin="WB000SHOP",
        instrument_name="Legacy SHOP",
        notes="legacy note",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            instrument_map={
                symbol: {
                    "yahoo_symbol": symbol,
                    "instrument_id": "SHOP_STATE",
                    "isin": "STATE000SHOP",
                    "instrument_name": "State SHOP",
                    "notes": "state note",
                }
            }
        ),
    )

    summary = project_store.adopt_legacy_workbook_instrument_map(
        project,
        _workbook_path(project),
        overwrite=False,
    )

    assert summary["skipped_conflicts"] >= 1
    assert summary["overwritten"] == 0

    state = project_store.load_project_state(project)
    assert state.instrument_map[symbol]["instrument_id"] == "SHOP_STATE"


def test_p3_3_adopt_overwrite_true_replaces_existing(tmp_path):
    """E: adopt with overwrite=True replaces existing ProjectState entries."""
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_LEGACY",
        isin="WB000SHOP",
        instrument_name="Legacy SHOP",
        notes="legacy note",
    )
    project_store.save_project_state(
        project,
        ProjectState(
            instrument_map={
                symbol: {
                    "yahoo_symbol": symbol,
                    "instrument_id": "SHOP_STATE",
                    "isin": "STATE000SHOP",
                    "instrument_name": "State SHOP",
                    "notes": "state note",
                }
            }
        ),
    )

    summary = project_store.adopt_legacy_workbook_instrument_map(
        project,
        _workbook_path(project),
        overwrite=True,
    )

    assert summary["overwritten"] >= 1

    state = project_store.load_project_state(project)
    assert state.instrument_map[symbol]["instrument_id"] == "SHOP_LEGACY"
    assert state.instrument_map[symbol]["isin"] == "WB000SHOP"


def test_p3_3_adopt_invalid_rows_skipped_with_counters(tmp_path):
    """F: Rows with missing Yahoo Symbol are skipped and counted."""
    project = _copy_project_fixture(tmp_path)
    _ensure_test_workbook(project)

    # Inject a blank-symbol row directly into the workbook Instrument_Map sheet.
    from openpyxl import load_workbook as _lw
    wb = _lw(_workbook_path(project))
    ws = wb["Instrument_Map"]
    blank_row = ws.max_row + 1
    ws.cell(row=blank_row, column=1, value=None)  # no Yahoo Symbol
    ws.cell(row=blank_row, column=2, value="BOGUS_ID")
    wb.save(_workbook_path(project))

    summary = project_store.adopt_legacy_workbook_instrument_map(
        project,
        _workbook_path(project),
    )

    assert summary["skipped_invalid"] >= 1


def test_p3_3_project_state_instrument_map_survives_recalc(tmp_path):
    """G: Existing ProjectState instrument mapping is preserved through recalc/reload."""
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    project_store.save_project_state(
        project,
        ProjectState(
            instrument_map={
                symbol: {
                    "yahoo_symbol": symbol,
                    "instrument_id": "SHOP_PERSISTENT",
                    "isin": "PERSIST000",
                    "instrument_name": "Persistent SHOP",
                    "notes": "persisted",
                }
            }
        ),
    )

    # First calculation.
    result1 = run(project_dir=project, write_workbook=False)
    sale1 = next(sale for sale in result1.sales.items if sale.ticker == symbol)
    assert sale1.instrument_id == "SHOP_PERSISTENT"

    # Second calculation — state must survive.
    result2 = run(project_dir=project, write_workbook=False)
    sale2 = next(sale for sale in result2.sales.items if sale.ticker == symbol)
    assert sale2.instrument_id == "SHOP_PERSISTENT"
    assert sale2.instrument_map_source == "project_state"

    reloaded = project_store.load_project_state(project)
    assert reloaded.instrument_map[symbol]["instrument_id"] == "SHOP_PERSISTENT"


def test_p3_3_instrument_map_source_never_workbook_fallback(tmp_path):
    """I: instrument_map_source must never report workbook_fallback after P3.3."""
    project = _copy_project_fixture(tmp_path)
    symbol = "SHOP"
    _set_workbook_instrument_map(
        project,
        symbol,
        instrument_id="SHOP_WORKBOOK",
        isin="WB999SHOP",
        instrument_name="Workbook SHOP",
        notes="wb note",
    )
    # No ProjectState instrument map entry.
    result = run(project_dir=project, write_workbook=False)

    for sale in result.sales.items:
        assert sale.instrument_map_source != "workbook_fallback", (
            f"Sale {sale.id} (ticker={sale.ticker}) reported instrument_map_source='workbook_fallback' "
            "after P3.3 retirement"
        )
    for pos in result.open_positions.items:
        assert pos.instrument_map_source != "workbook_fallback", (
            f"Open position {pos.instrument_id} reported instrument_map_source='workbook_fallback' "
            "after P3.3 retirement"
        )
